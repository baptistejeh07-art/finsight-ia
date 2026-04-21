"""Writer Excel générique pour TOUS les secteurs (hors Énergie qui a son
template dédié SECTOR_ENERGIE_TEMPLATE.xlsx).

Produit un XLSX avec 3 feuilles auto-calculées :
  - Dashboard : portrait secteur (HHI, P/E cycle, ROIC dispersion, compo top 5)
  - Screening : tableau de 1 ligne par société avec ratios clés
  - Analytics : statistiques agrégées (médianes, moyennes, écart-types)

Ce writer est appelé en fallback quand il n'existe pas de template dédié
pour le secteur analysé. But : que Baptiste ait TOUJOURS un XLSX, peu importe
le secteur.

Usage :
    from outputs.sector_generic_xlsx_writer import write_generic_sector_xlsx
    path = write_generic_sector_xlsx(
        sector="Healthcare",
        universe="S&P 500",
        tickers_data=[...],
        sector_analytics={...},
        output_path="outputs/generated/cli_tests/secteur_Healthcare_S&P_500.xlsx",
    )
"""
from __future__ import annotations
import logging
from pathlib import Path
from typing import Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

NAVY = "1B2A4A"
NAVY_LIGHT = "E8EAF0"
INK_50 = "F9FAFB"
INK_500 = "6B7280"
WHITE = "FFFFFF"

_HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="1B2A4A")
_VALUE_FONT = Font(name="Calibri", size=11, bold=True, color="111827")
_HINT_FONT = Font(name="Calibri", size=8, italic=True, color=INK_500)
_BORDER = Border(
    left=Side(border_style="thin", color="E5E7EB"),
    right=Side(border_style="thin", color="E5E7EB"),
    top=Side(border_style="thin", color="E5E7EB"),
    bottom=Side(border_style="thin", color="E5E7EB"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")


def _safe_num(x) -> Optional[float]:
    try:
        v = float(x)
        return v if not (np.isnan(v) or np.isinf(v)) else None
    except (TypeError, ValueError):
        return None


def _median(values: list[float]) -> Optional[float]:
    clean = [v for v in values if v is not None]
    return round(float(np.median(clean)), 2) if clean else None


def _mean(values: list[float]) -> Optional[float]:
    clean = [v for v in values if v is not None]
    return round(float(np.mean(clean)), 2) if clean else None


def write_generic_sector_xlsx(
    sector: str,
    universe: str,
    tickers_data: list[dict],
    sector_analytics: dict,
    output_path: str | Path,
) -> Path:
    """Écrit un XLSX générique à 3 feuilles pour un secteur quelconque."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Supprime la feuille par défaut
    default = wb.active
    wb.remove(default)

    _build_dashboard(wb, sector, universe, tickers_data, sector_analytics)
    _build_screening(wb, sector, tickers_data)
    _build_analytics(wb, sector, tickers_data)

    wb.save(str(output_path))
    log.info(
        f"[sector_generic_xlsx] {output_path.name} | secteur={sector} "
        f"| {len(tickers_data)} sociétés | 3 feuilles"
    )
    return output_path


def _build_dashboard(wb: Workbook, sector: str, universe: str,
                     tickers: list[dict], analytics: dict) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 50

    ws["B2"] = f"Portrait secteur — {sector}"
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color=NAVY)
    ws["B3"] = f"Univers : {universe} · {len(tickers)} sociétés couvertes"
    ws["B3"].font = _HINT_FONT

    rows = [
        ("HHI (Herfindahl-Hirschman)", analytics.get("hhi"), analytics.get("hhi_label")),
        ("P/E médian LTM", analytics.get("pe_median_ltm"), analytics.get("pe_cycle_label")),
        ("P/E médian historique", analytics.get("pe_median_hist"), "5 ans"),
        ("ROIC moyen", analytics.get("roic_mean"), analytics.get("roic_label")),
        ("ROIC écart-type", analytics.get("roic_std"), "Dispersion qualité"),
        ("ROIC min / max",
         f"{analytics.get('roic_min')} / {analytics.get('roic_max')}"
         if analytics.get("roic_min") is not None else None,
         "Fourchette sectorielle"),
    ]
    r = 5
    ws.cell(row=r, column=2, value="Indicateur").font = _HEADER_FONT
    ws.cell(row=r, column=2).fill = _HEADER_FILL
    ws.cell(row=r, column=2).alignment = _LEFT
    ws.cell(row=r, column=3, value="Valeur").font = _HEADER_FONT
    ws.cell(row=r, column=3).fill = _HEADER_FILL
    ws.cell(row=r, column=3).alignment = _CENTER
    ws.cell(row=r, column=4, value="Lecture").font = _HEADER_FONT
    ws.cell(row=r, column=4).fill = _HEADER_FILL
    ws.cell(row=r, column=4).alignment = _LEFT
    ws.row_dimensions[r].height = 22

    for i, (label, value, hint) in enumerate(rows):
        rr = r + 1 + i
        ws.cell(row=rr, column=2, value=label).font = _LABEL_FONT
        ws.cell(row=rr, column=2).border = _BORDER
        ws.cell(row=rr, column=2).alignment = _LEFT
        val_cell = ws.cell(row=rr, column=3, value=value if value is not None else "—")
        val_cell.font = _VALUE_FONT
        val_cell.border = _BORDER
        val_cell.alignment = _CENTER
        hint_cell = ws.cell(row=rr, column=4, value=hint or "")
        hint_cell.font = _HINT_FONT
        hint_cell.border = _BORDER
        hint_cell.alignment = _LEFT
        ws.row_dimensions[rr].height = 20

    # Top 5 sociétés par market cap
    sorted_t = sorted(tickers, key=lambda x: x.get("market_cap") or 0, reverse=True)[:5]
    r2 = r + len(rows) + 3
    ws.cell(row=r2, column=2, value="Top 5 par capitalisation").font = Font(
        name="Calibri", size=12, bold=True, color=NAVY
    )
    hdrs = ["Ticker", "Société", "Market Cap (Mds)"]
    for j, h in enumerate(hdrs):
        c = ws.cell(row=r2 + 1, column=2 + j, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER if j > 0 else _LEFT
    ws.row_dimensions[r2 + 1].height = 20
    for i, t in enumerate(sorted_t):
        rr = r2 + 2 + i
        mc_b = (t.get("market_cap") or 0) / 1_000_000_000
        ws.cell(row=rr, column=2, value=t.get("ticker") or "").font = Font(
            name="Consolas", size=10, bold=True
        )
        ws.cell(row=rr, column=3, value=t.get("name") or t.get("company_name") or "")
        ws.cell(row=rr, column=4, value=round(mc_b, 2)).number_format = "#,##0.00"
        for c in range(2, 5):
            ws.cell(row=rr, column=c).border = _BORDER


def _build_screening(wb: Workbook, sector: str, tickers: list[dict]) -> None:
    ws = wb.create_sheet("Screening")
    cols = [
        ("Ticker", 12),
        ("Société", 28),
        ("Market Cap (Mds)", 16),
        ("P/E", 10),
        ("EV/EBITDA", 12),
        ("EV/Revenue", 12),
        ("Mg Brute %", 12),
        ("Mg EBITDA %", 13),
        ("Mg Nette %", 12),
        ("ROE %", 10),
        ("ROIC %", 10),
        ("Croissance rev %", 16),
        ("Dette nette / EBITDA", 20),
        ("Altman Z", 10),
    ]
    for j, (h, w) in enumerate(cols):
        ws.column_dimensions[get_column_letter(j + 1)].width = w
        c = ws.cell(row=1, column=j + 1, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
    ws.row_dimensions[1].height = 24

    for i, t in enumerate(tickers):
        r = 2 + i
        r_data = t.get("ratios") or {}
        row_vals = [
            t.get("ticker") or "",
            t.get("name") or t.get("company_name") or "",
            round((t.get("market_cap") or 0) / 1e9, 2) if t.get("market_cap") else None,
            _safe_num(r_data.get("pe_ratio") or t.get("pe_ratio")),
            _safe_num(r_data.get("ev_ebitda") or t.get("ev_ebitda")),
            _safe_num(r_data.get("ev_revenue") or t.get("ev_revenue")),
            _safe_num(r_data.get("gross_margin") or t.get("gross_margin")),
            _safe_num(r_data.get("ebitda_margin") or t.get("ebitda_margin")),
            _safe_num(r_data.get("net_margin") or t.get("net_margin")),
            _safe_num(r_data.get("roe") or t.get("roe")),
            _safe_num(r_data.get("roic") or t.get("roic")),
            _safe_num(r_data.get("revenue_growth") or t.get("revenue_growth")),
            _safe_num(r_data.get("net_debt_ebitda") or t.get("net_debt_ebitda")),
            _safe_num(r_data.get("altman_z") or t.get("altman_z")),
        ]
        for j, v in enumerate(row_vals):
            c = ws.cell(row=r, column=j + 1, value=v if v is not None else "—")
            c.border = _BORDER
            c.alignment = _RIGHT if j >= 2 else _LEFT
            if j in (6, 7, 8, 9, 10, 11) and isinstance(v, (int, float)):
                c.number_format = "0.0"
            elif j in (3, 4, 5, 12, 13) and isinstance(v, (int, float)):
                c.number_format = "0.00"

    # Fige la 1ère ligne
    ws.freeze_panes = "A2"


def _build_analytics(wb: Workbook, sector: str, tickers: list[dict]) -> None:
    ws = wb.create_sheet("Analytics")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    ws["B2"] = f"Statistiques agrégées — {sector}"
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    # Extract par métrique
    def _vals(key: str) -> list[float]:
        out = []
        for t in tickers:
            r = t.get("ratios") or {}
            v = _safe_num(r.get(key) or t.get(key))
            if v is not None:
                out.append(v)
        return out

    metrics = [
        ("P/E", "pe_ratio"),
        ("EV/EBITDA", "ev_ebitda"),
        ("EV/Revenue", "ev_revenue"),
        ("Marge brute %", "gross_margin"),
        ("Marge EBITDA %", "ebitda_margin"),
        ("Marge nette %", "net_margin"),
        ("ROE %", "roe"),
        ("ROIC %", "roic"),
        ("Croissance revenue %", "revenue_growth"),
        ("Dette nette / EBITDA", "net_debt_ebitda"),
        ("Altman Z", "altman_z"),
    ]

    r = 4
    hdrs = ["Métrique", "Médiane", "Moyenne", "N"]
    for j, h in enumerate(hdrs):
        c = ws.cell(row=r, column=2 + j, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER if j > 0 else _LEFT
    ws.row_dimensions[r].height = 22

    for i, (label, key) in enumerate(metrics):
        rr = r + 1 + i
        vs = _vals(key)
        ws.cell(row=rr, column=2, value=label).font = _LABEL_FONT
        med = _median(vs)
        mean = _mean(vs)
        ws.cell(row=rr, column=3, value=med if med is not None else "—").number_format = "0.00"
        ws.cell(row=rr, column=4, value=mean if mean is not None else "—").number_format = "0.00"
        ws.cell(row=rr, column=5, value=len(vs))
        for c in range(2, 6):
            ws.cell(row=rr, column=c).border = _BORDER
            if c > 2:
                ws.cell(row=rr, column=c).alignment = _CENTER
