"""
Excel writer PME — 5 feuilles :
  F1 : SIG détaillés (format expert-comptable)
  F2 : Ratios 5 ans + médianes sectorielles
  F3 : Benchmark peers (positionnement visuel)
  F4 : Scoring (santé + bankabilité + Altman)
  F5 : M&A placeholder (désactivé v1)
"""

from __future__ import annotations

import logging
from pathlib import Path

from core.pappers.analytics import PmeAnalysis, YearAccounts
from core.pappers.benchmark import BenchmarkResult
from core.pappers.bodacc_client import BodaccSummary

log = logging.getLogger(__name__)

NAVY = "1B2A4A"
NAVY_LIGHT = "3A5288"
INK_100 = "F0F0F0"
INK_200 = "E5E5E5"
GREEN = "15803D"
AMBER = "D97706"
RED = "DC2626"


def write_pme_xlsx(
    output_path: str | Path,
    yearly_accounts: list[YearAccounts],
    analysis: PmeAnalysis,
    benchmark: BenchmarkResult,
    bodacc: BodaccSummary | None,
    siren: str,
    denomination: str,
    language: str = "fr",
    currency: str = "EUR",
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError as e:
        raise RuntimeError("openpyxl requis") from e

    # i18n helper
    from core.i18n import t as _i18n_t, ratio_label as _i18n_ratio, sig_label as _i18n_sig, normalize_language
    _lang = normalize_language(language)
    def _t(key, default=None):
        return _i18n_t(_lang, key, default)
    def _rl(key):
        return _i18n_ratio(key, _lang)
    def _sl(key):
        return _i18n_sig(key, _lang)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Styles
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    label_font = Font(bold=True)
    thin = Side(style="thin", color=INK_200)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ─── F1 : SIG ───
    ws = wb.active
    ws.title = _t("report.sig", "SIG")[:31]  # max 31 chars sheet name
    ws["A1"] = f"{denomination} ({siren}) — {_t('report.sig')}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws.merge_cells(f"A1:{chr(ord('A') + len(analysis.sig_by_year))}1")

    years = sorted(analysis.sig_by_year.keys())
    ws["A3"] = "Indicateur"
    ws["A3"].font = header_font
    ws["A3"].fill = header_fill
    for i, y in enumerate(years, start=2):
        c = ws.cell(row=3, column=i, value=y)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    sig_lines = [
        (_sl("chiffre_affaires"), lambda y: next((a.chiffre_affaires for a in yearly_accounts if a.annee == y), None)),
        (_sl("production_exercice"), lambda y: analysis.sig_by_year[y].production_exercice),
        (_sl("valeur_ajoutee"), lambda y: analysis.sig_by_year[y].valeur_ajoutee),
        (_sl("ebe"), lambda y: analysis.sig_by_year[y].ebe),
        (_sl("resultat_exploitation"), lambda y: analysis.sig_by_year[y].resultat_exploitation),
        (_sl("resultat_courant"), lambda y: analysis.sig_by_year[y].resultat_courant_av_impots),
        (_sl("resultat_net"), lambda y: analysis.sig_by_year[y].resultat_net),
        (_sl("caf"), lambda y: analysis.sig_by_year[y].capacite_autofinancement),
        (_sl("charges_personnel"), lambda y: analysis.sig_by_year[y].charges_personnel_total),
        (_sl("consommations_externes"), lambda y: analysis.sig_by_year[y].consommations_externes),
    ]
    for row_idx, (label, extractor) in enumerate(sig_lines, start=4):
        c = ws.cell(row=row_idx, column=1, value=label)
        c.font = label_font
        for i, y in enumerate(years, start=2):
            val = extractor(y)
            cell = ws.cell(row=row_idx, column=i, value=val)
            cell.number_format = '#,##0 "€"'
            cell.alignment = Alignment(horizontal="right")
    ws.column_dimensions["A"].width = 42
    for i in range(len(years)):
        ws.column_dimensions[chr(ord("B") + i)].width = 16

    # ─── F2 : Ratios ───
    ws2 = wb.create_sheet(_t("report.ratios", "Ratios")[:31])
    ws2["A1"] = f"{denomination} — Ratios clés"
    ws2["A1"].font = title_font
    ws2["A1"].fill = title_fill
    ws2.merge_cells(f"A1:{chr(ord('A') + len(years))}1")

    ws2["A3"] = "Ratio"
    ws2["A3"].font = header_font
    ws2["A3"].fill = header_fill
    for i, y in enumerate(years, start=2):
        c = ws2.cell(row=3, column=i, value=y)
        c.font = header_font
        c.fill = header_fill

    ratio_lines = [
        (_rl("marge_ebitda"), "marge_ebitda", "pct"),
        (_rl("marge_nette"), "marge_nette", "pct"),
        (_rl("roce"), "roce", "pct"),
        (_rl("roe"), "roe", "pct"),
        (_rl("dette_nette_ebitda"), "dette_nette_ebitda", "x"),
        (_rl("couverture_interets"), "couverture_interets", "x"),
        (_rl("autonomie_financiere"), "autonomie_financiere", "pct"),
        (_rl("bfr_jours_ca"), "bfr_jours_ca", "days"),
        (_rl("dso_jours"), "dso_jours", "days"),
        (_rl("dpo_jours"), "dpo_jours", "days"),
        (_rl("rotation_stocks"), "rotation_stocks", "x"),
        (_rl("ca_par_employe"), "ca_par_employe", "eur"),
        (_rl("charges_perso_ca"), "charges_perso_ca", "pct"),
    ]
    for row_idx, (label, attr, kind) in enumerate(ratio_lines, start=4):
        ws2.cell(row=row_idx, column=1, value=label).font = label_font
        for i, y in enumerate(years, start=2):
            r = analysis.ratios_by_year.get(y)
            val = getattr(r, attr) if r else None
            cell = ws2.cell(row=row_idx, column=i, value=val)
            if kind == "pct":
                cell.number_format = "0.0%"
            elif kind == "x":
                cell.number_format = "0.00"
            elif kind == "days":
                cell.number_format = "0 \"j\""
            elif kind == "eur":
                cell.number_format = '#,##0 "€"'
            cell.alignment = Alignment(horizontal="right")
    ws2.column_dimensions["A"].width = 32
    for i in range(len(years)):
        ws2.column_dimensions[chr(ord("B") + i)].width = 14

    # ─── F3 : Benchmark ───
    ws3 = wb.create_sheet(_t("report.sector_benchmark", "Benchmark")[:31])
    ws3["A1"] = f"{denomination} — Benchmark sectoriel (source: {benchmark.source})"
    ws3["A1"].font = title_font
    ws3["A1"].fill = title_fill
    ws3.merge_cells("A1:E1")

    headers = ["Ratio", "Cible", "Q25", "Médiane (Q50)", "Q75", "Position"]
    for i, h in enumerate(headers, start=1):
        cell = ws3.cell(row=3, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
    rank_labels = {
        "top_25": "Top 25%", "above_median": "Au-dessus médiane",
        "below_median": "Sous médiane", "bottom_25": "Bottom 25%",
    }
    rank_colors = {
        "top_25": GREEN, "above_median": GREEN,
        "below_median": AMBER, "bottom_25": RED,
    }
    for row_idx, (name, q) in enumerate(benchmark.ratios.items(), start=4):
        ws3.cell(row=row_idx, column=1, value=name).font = label_font
        for col_idx, val in enumerate([q.value, q.q25, q.q50, q.q75], start=2):
            ws3.cell(row=row_idx, column=col_idx, value=val).number_format = "0.00"
        rank_cell = ws3.cell(row=row_idx, column=6, value=rank_labels.get(q.rank, "—"))
        rank_cell.font = Font(bold=True, color=rank_colors.get(q.rank, "000000"))
    ws3.column_dimensions["A"].width = 28
    for c in "BCDEF":
        ws3.column_dimensions[c].width = 16

    # ─── F4 : Scoring ───
    ws4 = wb.create_sheet(_t("report.scoring", "Scoring")[:31])
    ws4["A1"] = f"{denomination} — Scoring"
    ws4["A1"].font = title_font
    ws4["A1"].fill = title_fill
    ws4.merge_cells("A1:B1")

    scoring = [
        ("Altman Z-Score (non coté)", analysis.altman_z),
        ("Verdict Altman", analysis.altman_verdict),
        ("Score santé FinSight (0-100)", analysis.health_score),
        ("Score bankabilité (0-100)", analysis.bankability_score),
        ("Capacité dette additionnelle (€)", analysis.debt_capacity_estimate),
    ]
    if bodacc:
        scoring.extend([
            ("BODACC — Annonces totales", bodacc.total_annonces),
            ("BODACC — Procédures collectives", len(bodacc.procedures_collectives)),
            ("BODACC — Radiée", "Oui" if bodacc.radie else "Non"),
            ("BODACC — Pénalité scoring", bodacc.bodacc_score_penalty),
        ])
    for row_idx, (label, value) in enumerate(scoring, start=3):
        ws4.cell(row=row_idx, column=1, value=label).font = label_font
        ws4.cell(row=row_idx, column=2, value=value)
    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["B"].width = 18

    # ─── F5 : M&A placeholder ───
    ws5 = wb.create_sheet("M&A (v2)")
    ws5["A1"] = "Module M&A — disponible en version premium"
    ws5["A1"].font = title_font
    ws5["A1"].fill = title_fill
    ws5.merge_cells("A1:D1")
    ws5["A3"] = ("Le module M&A (valorisation multiples + DCF simplifié + LBO indicatif) "
                 "sera activé dans une future version. Il nécessite un scope utilisateur "
                 "explicite (acquisition ou cession).")
    ws5["A3"].alignment = Alignment(wrap_text=True)
    ws5.merge_cells("A3:D6")
    ws5.column_dimensions["A"].width = 25

    wb.save(output_path)
    return output_path
