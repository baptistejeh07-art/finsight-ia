# -*- coding: utf-8 -*-
"""
outputs/lbo_model.py — Module LBO Model FinSight.

Construit les feuilles "LBO MODEL" (visible) et "_LBO_CALC" (masquee) dans un
workbook openpyxl deja charge. Utilise par excel_writer.py a chaque generation
d'analyse societe.

API publique :
    build_lbo_visible(wb)  -> ajoute la feuille LBO MODEL
    build_lbo_calc(wb)     -> ajoute la feuille _LBO_CALC masquee
    build_lbo_sheets(wb)   -> les deux + cellules nommees (helper unique)

Style : replique fidele de la charte FinSight (couleurs, polices, layout DCF).
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle, Color
)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle


# ═════════════════════════════════════════════════════════════════════════════
# CHARTE VISUELLE FINSIGHT (extraite de DCF/RATIOS)
# ═════════════════════════════════════════════════════════════════════════════

# Couleurs (sans le préfixe FF)
NAVY_DARK    = "000C3F"   # bandeau header pleine largeur
NAVY_TEXT    = "132E57"   # texte navy data
BLUE_SECTION = "3271D2"   # texte titre section
BLUE_PALE    = "E7F2FF"   # fond section header
BLUE_TABLE   = "D9E5F7"   # fond table header
GREY_CALC    = "F2F2F2"   # fond cellules calculées
GREY_BORDER  = "BFBFBF"   # bordure fine
WHITE        = "FFFFFF"
INPUT_BLUE   = "0000FF"   # police formules paramétrables (cellules input jaunes)
INPUT_YELLOW = "FFF4D6"   # fond cellules paramétrables
EXPORT_PALE  = "EEF3FA"   # fond zone EXPORT
GREEN_OK     = "1A7A4A"
GREEN_PALE   = "E8F5EE"
RED_KO       = "A82020"
RED_PALE     = "FBEBEB"
AMBER_WARN   = "B06000"
AMBER_PALE   = "FDF3E5"

FONT_FAMILY = "Open Sans"


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS DE STYLE
# ═════════════════════════════════════════════════════════════════════════════

def font(size=10, bold=False, italic=False, color="000000"):
    return Font(name=FONT_FAMILY, size=size, bold=bold, italic=italic,
                color=Color(rgb="FF" + color))


def fill(rgb):
    return PatternFill(fill_type="solid",
                       start_color=Color(rgb="FF" + rgb),
                       end_color=Color(rgb="FF" + rgb))


def border(top=False, bottom=False, left=False, right=False,
           color=GREY_BORDER, style="thin"):
    side = Side(style=style, color=Color(rgb="FF" + color))
    none = Side(style=None)
    return Border(
        top=side if top else none,
        bottom=side if bottom else none,
        left=side if left else none,
        right=side if right else none,
    )


def box_border(color=GREY_BORDER, style="thin"):
    side = Side(style=style, color=Color(rgb="FF" + color))
    return Border(top=side, bottom=side, left=side, right=side)


def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left_align(wrap=False, indent=0):
    return Alignment(horizontal="left", vertical="center",
                     wrap_text=wrap, indent=indent)


def right_align():
    return Alignment(horizontal="right", vertical="center")


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS DE REMPLISSAGE CELLULES
# ═════════════════════════════════════════════════════════════════════════════

def set_cell(ws, ref, value, *, font_obj=None, fill_obj=None, border_obj=None,
             align=None, number_format=None):
    cell = ws[ref]
    cell.value = value
    if font_obj is not None:
        cell.font = font_obj
    if fill_obj is not None:
        cell.fill = fill_obj
    if border_obj is not None:
        cell.border = border_obj
    if align is not None:
        cell.alignment = align
    if number_format is not None:
        cell.number_format = number_format
    return cell


def merge_and_set(ws, range_str, value, *, font_obj=None, fill_obj=None,
                  border_obj=None, align=None):
    ws.merge_cells(range_str)
    first = range_str.split(":")[0]
    set_cell(ws, first, value, font_obj=font_obj, fill_obj=fill_obj,
             border_obj=border_obj, align=align)


# ═════════════════════════════════════════════════════════════════════════════
# CONSTRUCTION ONGLET VISIBLE "LBO MODEL"
# ═════════════════════════════════════════════════════════════════════════════

def build_lbo_visible(wb):
    """Construit l'onglet visible LBO MODEL (style cmp DCF FinSight)."""
    if "LBO MODEL" in wb.sheetnames:
        del wb["LBO MODEL"]
    ws = wb.create_sheet("LBO MODEL")

    # ─── Largeurs colonnes (mimic DCF) ───
    widths = {
        "A": 2.0, "B": 42.0, "C": 12.0, "D": 13.0, "E": 13.0,
        "F": 13.0, "G": 13.0, "H": 13.0, "I": 13.0, "J": 13.0,
        "K": 13.0, "L": 13.0, "M": 13.0,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ─── Hauteurs lignes ───
    ws.row_dimensions[1].height = 54.95   # bandeau header
    ws.row_dimensions[3].height = 24.0    # titre LBO
    ws.row_dimensions[4].height = 16.0    # sous-titre

    # ═══════════════════════════════════════════════════════════════════
    # LIGNE 1 : Bandeau navy "FinSight"
    # ═══════════════════════════════════════════════════════════════════
    merge_and_set(
        ws, "B1:M1", "FinSight",
        font_obj=font(size=18, bold=True, color=WHITE),
        fill_obj=fill(NAVY_DARK),
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )

    # ═══════════════════════════════════════════════════════════════════
    # LIGNE 3 : Titre principal section
    # ═══════════════════════════════════════════════════════════════════
    merge_and_set(
        ws, "B3:M3", "MODÈLE LBO",
        font_obj=font(size=14, bold=True, color=BLUE_SECTION),
        fill_obj=fill(BLUE_PALE),
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    merge_and_set(
        ws, "B4:M4",
        "Modèle institutionnel mid-market  ·  calculs sur 5 ans  ·  toutes hypothèses paramétrables (cellules jaunes)",
        font_obj=font(size=9, italic=True, color="555555"),
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )

    # ═══════════════════════════════════════════════════════════════════
    # SECTION A — PROFIL & ÉLIGIBILITÉ
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 6, "A.  PROFIL & ÉLIGIBILITÉ")

    # Ligne 7 : Société + EV
    # Société et ticker extraits depuis STOCK_DATA!A1 qui contient
    # "STOCK_DATA — Apple Inc. (AAPL) — Trailing 52 Weeks" (em-dash U+2014)
    # Parsing position-fixe : nom commence position 14, parenthèse ouvrante = fin
    set_cell(ws, "B7", "Société",
             font_obj=font(bold=False, color="555555"))
    set_cell(ws, "C7",
             '=IFERROR(TRIM(MID(STOCK_DATA!A1,14,FIND("(",STOCK_DATA!A1)-15)),"")',
             font_obj=font(bold=True, color=NAVY_TEXT))
    set_cell(ws, "G7", "Enterprise Value LTM",
             font_obj=font(bold=False, color="555555"))
    set_cell(ws, "H7", "=INPUT!H98",
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0_);[Red](#,##0);"—"')

    # Ligne 8 : Ticker + EBITDA
    set_cell(ws, "B8", "Ticker",
             font_obj=font(bold=False, color="555555"))
    # Parsing : extrait le ticker entre parenthèses
    set_cell(ws, "C8",
             '=IFERROR(MID(STOCK_DATA!A1,'
             'SEARCH("(",STOCK_DATA!A1)+1,'
             'SEARCH(")",STOCK_DATA!A1)-SEARCH("(",STOCK_DATA!A1)-1),"")',
             font_obj=font(bold=True, color=NAVY_TEXT))
    set_cell(ws, "G8", "EBITDA LTM",
             font_obj=font(bold=False, color="555555"))
    set_cell(ws, "H8", "=INPUT!H30",
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0_);[Red](#,##0);"—"')

    # Ligne 9 : Revenue LTM + ND/EBITDA
    set_cell(ws, "B9", "Revenue LTM",
             font_obj=font(bold=False, color="555555"))
    set_cell(ws, "C9", "=INPUT!H9",
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0_);[Red](#,##0);"—"')
    set_cell(ws, "G9", "Net Debt / EBITDA",
             font_obj=font(bold=False, color="555555"))
    set_cell(ws, "H9", "=IFERROR((INPUT!H54+INPUT!H48-INPUT!H34)/INPUT!H30,0)",
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='0.0"x"')

    # Ligne 10 : Marge EBITDA + FCF/EBITDA
    set_cell(ws, "B10", "Marge EBITDA",
             font_obj=font(bold=False, color="555555"))
    set_cell(ws, "C10", "=IFERROR(INPUT!H30/INPUT!H9,0)",
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='0.0%')
    set_cell(ws, "G10", "Cash Conversion (FCF/EBITDA)",
             font_obj=font(bold=False, color="555555"))
    set_cell(ws, "H10", "=IFERROR(DCF!H49/INPUT!H30,0)",
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='0.0%')

    # ─── Statut éligibilité (lignes 12-14) ───
    # Ligne 12 : grosse cellule éligibilité
    ws.row_dimensions[12].height = 28.0
    merge_and_set(
        ws, "B12:M12",
        '=IF(AND(INPUT!H30>0,'
        'IFERROR(INPUT!H30/INPUT!H9,0)>0.15,'
        'IFERROR(DCF!H49/INPUT!H30,0)>0.6,'
        'IFERROR((INPUT!H54+INPUT!H48-INPUT!H34)/INPUT!H30,99)<3.5),'
        '"  ✓  PROFIL LBO-ABLE  —  Société éligible aux critères d\'un LBO mid-market standard",'
        '"  ✗  PROFIL NON ÉLIGIBLE  —  Voir critères ci-dessous pour les freins identifiés")',
        font_obj=font(size=12, bold=True, color=NAVY_TEXT),
        fill_obj=fill(GREEN_PALE),
        align=Alignment(horizontal="left", vertical="center"),
    )

    # Ligne 13 : critères individuels (4 colonnes)
    set_cell(ws, "B13",
             '=IF(IFERROR(INPUT!H30/INPUT!H9,0)>0.15,"✓ Marge EBITDA > 15%","✗ Marge EBITDA < 15%")',
             font_obj=font(size=9, color=NAVY_TEXT))
    set_cell(ws, "E13",
             '=IF(IFERROR(DCF!H49/INPUT!H30,0)>0.6,"✓ Cash conv > 60%","✗ Cash conv < 60%")',
             font_obj=font(size=9, color=NAVY_TEXT))
    set_cell(ws, "H13",
             '=IF(IFERROR((INPUT!H54+INPUT!H48-INPUT!H34)/INPUT!H30,99)<3.5,"✓ Levier < 3,5x","✗ Levier > 3,5x")',
             font_obj=font(size=9, color=NAVY_TEXT))
    set_cell(ws, "K13",
             '=IF(INPUT!H30>0,"✓ EBITDA positif","✗ EBITDA négatif")',
             font_obj=font(size=9, color=NAVY_TEXT))

    # Ligne 14 : Mega-deal flag (3 niveaux)
    ws.row_dimensions[14].height = 22.0
    merge_and_set(
        ws, "B14:M14",
        '=IF(INPUT!H98>100000000,'
        '"  ⚠⚠  SCÉNARIO THÉORIQUE  —  EV > 100 Md$, taille hors marché LBO historique",'
        'IF(INPUT!H98>50000000,'
        '"  ⚠  MEGA-DEAL  —  EV entre 50 et 100 Md$, syndicate sponsorship requis (type Twitter/X, Heinz)",'
        '""))',
        font_obj=font(size=10, bold=True, color=AMBER_WARN),
        fill_obj=fill(AMBER_PALE),
        align=Alignment(horizontal="left", vertical="center"),
    )

    # ═══════════════════════════════════════════════════════════════════
    # SECTION B — HYPOTHÈSES (cellules jaunes paramétrables)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 16, "B.  HYPOTHÈSES (cellules jaunes paramétrables)")

    # Cellules paramétrables (fond jaune, police bleue)
    hypotheses_input = [
        ("B17", "Multiple d'entrée (× EBITDA LTM)", "C17", 14.0, '0.0"x"'),
        ("B19", "Leverage total au closing (× EBITDA)", "C19", 5.0, '0.0"x"'),
        ("B20", "  · dont Senior Term Loan B", "C20", 3.5, '0.0"x"'),
        ("B21", "  · dont Mezzanine", "C21", 1.5, '0.0"x"'),
        ("B22", "Coût Senior TLB (SOFR + 400 bps)", "C22", 0.08, '0.0%'),
        ("B23", "Coût Mezzanine (cash 5% + PIK 5%)", "C23", 0.10, '0.0%'),
        ("B24", "Multiple de sortie (× EBITDA Y5)", "C24", 14.0, '0.0"x"'),
        ("B25", "Période de détention (années)", "C25", 5, '0'),
        ("B26", "Transaction fees (% EV deal)", "C26", 0.025, '0.0%'),
        ("B27", "Cash minimum BS (% EV)", "C27", 0.05, '0.0%'),
    ]
    for label_ref, label, val_ref, val, fmt in hypotheses_input:
        set_cell(ws, label_ref, label,
                 font_obj=font(color=NAVY_TEXT),
                 align=left_align())
        set_cell(ws, val_ref, val,
                 font_obj=font(bold=True, color=INPUT_BLUE),
                 fill_obj=fill(INPUT_YELLOW),
                 border_obj=box_border(color="C9A227"),
                 align=center(),
                 number_format=fmt)

    # Cellule calculée (display) : Premium implicite vs Market Cap actuelle
    set_cell(ws, "B18", "Premium implicite vs Market Cap (calculé)",
             font_obj=font(color="555555", italic=True),
             align=left_align())
    set_cell(ws, "C18", "=IFERROR((L28-INPUT!H97)/INPUT!H97,0)",
             font_obj=font(bold=True, color=NAVY_TEXT),
             fill_obj=fill(GREY_CALC),
             align=center(),
             number_format='+0.0%;-0.0%;0.0%')

    # ─── Panneau droite : Multiples sectoriels & contexte marché ───
    # Aide l'utilisateur à voir si son multiple d'entrée est réaliste
    # On utilise des cellules à partir de F (libres car la section "Hypothèses"
    # B/C n'occupe que les colonnes B-C). Le header section B a déjà mergé B:M
    # ligne 16, donc on utilise F18 comme première ligne du panneau droite.

    set_cell(ws, "F17", "EV / EBITDA actuel marché",
             font_obj=font(size=9, color="555555"),
             align=left_align())
    set_cell(ws, "L17", "=IFERROR(INPUT!H98/INPUT!H30,0)",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             fill_obj=fill(GREY_CALC),
             align=right_align(),
             number_format='0.0"x"')

    set_cell(ws, "F18", "EV / EBITDA peers (méd. COMPARABLES)",
             font_obj=font(size=9, color="555555"),
             align=left_align())
    set_cell(ws, "L18", "=IFERROR(MEDIAN(COMPARABLES!H4:H20),0)",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             fill_obj=fill(GREY_CALC),
             align=right_align(),
             number_format='0.0"x"')

    set_cell(ws, "F19", "Différence vs multiple d'entrée saisi",
             font_obj=font(size=9, color="555555"),
             align=left_align())
    set_cell(ws, "L19", "=IFERROR(C17-L17,0)",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             fill_obj=fill(GREY_CALC),
             align=right_align(),
             number_format='+0.0"x";-0.0"x";0.0"x"')

    set_cell(ws, "F21", "Recommandation multiple d'entrée",
             font_obj=font(size=9, bold=True, color="555555"),
             align=left_align())
    set_cell(ws, "L21",
             '=IFERROR(ROUND(MAX(C17,L17*1.1),1),C17)',
             font_obj=font(size=9, bold=True, color=GREEN_OK),
             fill_obj=fill(GREEN_PALE),
             align=right_align(),
             number_format='0.0"x"')

    merge_and_set(ws, "F22:M22",
                  "Note : pour un LBO réaliste, le multiple d'entrée doit "
                  "être >= multiple actuel + 10% (premium d'acquisition).",
                  font_obj=font(size=8, italic=True, color="888888"),
                  align=left_align(indent=1))

    # ═══════════════════════════════════════════════════════════════════
    # SECTION C — SOURCES & USES
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 29, "C.  SOURCES & USES")

    # Sub-headers SOURCES (gauche) / USES (droite)
    set_cell(ws, "B30", "SOURCES (financement)",
             font_obj=font(size=10, bold=True, color=BLUE_SECTION),
             fill_obj=fill(BLUE_TABLE),
             align=center())
    merge_and_set(ws, "B30:E30", "SOURCES (financement)",
                  font_obj=font(size=10, bold=True, color=BLUE_SECTION),
                  fill_obj=fill(BLUE_TABLE),
                  align=center())
    merge_and_set(ws, "G30:M30", "USES (utilisation)",
                  font_obj=font(size=10, bold=True, color=BLUE_SECTION),
                  fill_obj=fill(BLUE_TABLE),
                  align=center())

    # EV deal MULTIPLE-DRIVEN : C17 (multiple entrée) × EBITDA LTM
    set_cell(ws, "K28", "EV_deal:", font_obj=font(size=8, color="999999"),
             align=right_align())
    set_cell(ws, "L28", "=C17*INPUT!H30",
             font_obj=font(size=8, bold=True, color="555555"),
             align=right_align(),
             number_format='#,##0')

    # SOURCES — Sponsor Equity = TOTAL USES - Senior - Mezz
    # (équilibre par construction Sources = Uses)
    sources = [
        ("B31", "Sponsor Equity", "D31", "=L35-D32-D33", "E31", "=IFERROR(D31/L35,0)"),
        ("B32", "Senior Term Loan B", "D32", "=C20*INPUT!H30", "E32", "=IFERROR(D32/L35,0)"),
        ("B33", "Mezzanine", "D33", "=C21*INPUT!H30", "E33", "=IFERROR(D33/L35,0)"),
    ]
    for lbl_ref, lbl, val_ref, val_formula, pct_ref, pct_formula in sources:
        set_cell(ws, lbl_ref, lbl, font_obj=font(color=NAVY_TEXT))
        set_cell(ws, val_ref, val_formula,
                 font_obj=font(bold=True, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0_);[Red](#,##0);"—"')
        set_cell(ws, pct_ref, pct_formula,
                 font_obj=font(color="555555"),
                 align=right_align(),
                 number_format='0.0%')

    # TOTAL Sources — = SUM des sources, équivaut à L35 par construction
    set_cell(ws, "B35", "TOTAL SOURCES",
             font_obj=font(bold=True, color=BLUE_SECTION),
             fill_obj=fill(BLUE_PALE),
             border_obj=border(top=True, bottom=True))
    set_cell(ws, "D35", "=SUM(D31:D33)",
             font_obj=font(bold=True, color=BLUE_SECTION),
             fill_obj=fill(BLUE_PALE),
             border_obj=border(top=True, bottom=True),
             align=right_align(),
             number_format='#,##0_);[Red](#,##0);"—"')
    set_cell(ws, "E35", "=IFERROR(D35/L35,0)",
             font_obj=font(bold=True, color=BLUE_SECTION),
             fill_obj=fill(BLUE_PALE),
             border_obj=border(top=True, bottom=True),
             align=right_align(),
             number_format='0.0%')

    # USES
    # Purchase price equity = EV_deal - Net Debt assumée (refi)
    # On rachète l'equity à : EV_deal - dette existante - cash
    # = (multiple × EBITDA) - (Senior+Mezz pre-existing) + cash existant
    uses = [
        ("G31", "Purchase price equity",  "L31",
         "=L28-(INPUT!H54+INPUT!H48-INPUT!H34)"),
        ("G32", "Refinancement dette existante", "L32",
         "=INPUT!H54+INPUT!H48"),
        ("G33", "Transaction fees", "L33", "=L28*C26"),
        ("G34", "Cash minimum sur BS", "L34", "=L28*C27"),
    ]
    for lbl_ref, lbl, val_ref, val_formula in uses:
        set_cell(ws, lbl_ref, lbl, font_obj=font(color=NAVY_TEXT))
        set_cell(ws, val_ref, val_formula,
                 font_obj=font(bold=True, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0_);[Red](#,##0);"—"')

    # TOTAL Uses
    set_cell(ws, "G35", "TOTAL USES",
             font_obj=font(bold=True, color=BLUE_SECTION),
             fill_obj=fill(BLUE_PALE),
             border_obj=border(top=True, bottom=True))
    set_cell(ws, "L35", "=SUM(L31:L34)",
             font_obj=font(bold=True, color=BLUE_SECTION),
             fill_obj=fill(BLUE_PALE),
             border_obj=border(top=True, bottom=True),
             align=right_align(),
             number_format='#,##0_);[Red](#,##0);"—"')

    # ═══════════════════════════════════════════════════════════════════
    # SECTION D — RETURNS BASE CASE (KPIs principaux)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 37, "D.  RETURNS BASE CASE  —  Sortie Y5")

    # 4 KPIs en grand : IRR, MOIC, Equity exit, Money-back
    set_cell(ws, "B38", "Equity entry sponsor",
             font_obj=font(color="555555"))
    set_cell(ws, "C38", "=D31",
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    set_cell(ws, "B39", "Equity exit (Y5)",
             font_obj=font(color="555555"))
    set_cell(ws, "C39", "=_LBO_CALC!N50",   # référence vers calc
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    # Grands chiffres KPI : IRR, MOIC
    ws.row_dimensions[41].height = 30.0
    merge_and_set(
        ws, "B41:E41", "IRR Sponsor",
        font_obj=font(size=10, bold=True, color="555555"),
        fill_obj=fill(BLUE_TABLE),
        align=center(),
    )
    merge_and_set(
        ws, "G41:J41", "MOIC",
        font_obj=font(size=10, bold=True, color="555555"),
        fill_obj=fill(BLUE_TABLE),
        align=center(),
    )

    ws.row_dimensions[42].height = 36.0
    merge_and_set(
        ws, "B42:E42", "=_LBO_CALC!N51",
        font_obj=font(size=22, bold=True, color=NAVY_TEXT),
        fill_obj=fill(BLUE_PALE),
        align=center(),
    )
    ws["B42"].number_format = '0.0%'
    merge_and_set(
        ws, "G42:J42", "=_LBO_CALC!N52",
        font_obj=font(size=22, bold=True, color=NAVY_TEXT),
        fill_obj=fill(BLUE_PALE),
        align=center(),
    )
    ws["G42"].number_format = '0.00"x"'

    # Money-back period (garde si IRR <= 0)
    set_cell(ws, "B44", "Money-back period",
             font_obj=font(color="555555"))
    set_cell(ws, "C44",
             '=IF(_LBO_CALC!N51<=0,"N/A",IF(LN(2)/LN(1+_LBO_CALC!N51)>C25*2,'
             '">"&TEXT(C25*2,"0")&" ans",ROUND(LN(2)/LN(1+_LBO_CALC!N51),1)&" ans"))',
             font_obj=font(bold=True, color=NAVY_TEXT),
             align=right_align())

    # ═══════════════════════════════════════════════════════════════════
    # SECTION E — SENSIBILITÉ IRR (heatmap 5×5)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 46, "E.  SENSIBILITÉ IRR  —  Multiple d'entrée × Multiple de sortie")

    # Header axe X (multiple de sortie)
    set_cell(ws, "B47", "↓ Mult. entrée  /  Mult. sortie →",
             font_obj=font(size=9, bold=True, color="555555"),
             fill_obj=fill(BLUE_TABLE),
             align=center())

    # Multiples sortie en colonnes E-I (centrés sur base)
    exit_mults = [-2, -1, 0, 1, 2]   # offsets vs base
    for i, off in enumerate(exit_mults):
        col = chr(ord("E") + i)
        set_cell(ws, f"{col}47", f"=C24+({off})",
                 font_obj=font(size=9, bold=True, color="555555"),
                 fill_obj=fill(BLUE_TABLE),
                 align=center(),
                 number_format='0.0"x"')

    # Multiples entrée en lignes 48-52
    # Formule sensibilité directe :
    # IRR = ((EBITDA_Y5 * mult_sortie - NetDebt_Y5) / Sponsor_Equity)^(1/n) - 1
    # avec :
    #   EBITDA_Y5 = _LBO_CALC!M7  (déjà projeté via DCF)
    #   NetDebt_Y5 = _LBO_CALC!M30 (déjà calculé via debt schedule)
    #   Sponsor_Equity = (mult_entree × EBITDA_LTM) - Senior - Mezz
    # Hypothèse simplifiée : la dette absolue ne dépend pas du multiple d'entrée,
    # c'est une approximation didactique pour la heatmap.
    entry_mults = [-2, -1, 0, 1, 2]
    for i, off_e in enumerate(entry_mults):
        row = 48 + i
        set_cell(ws, f"B{row}", f"=C17+({off_e})",
                 font_obj=font(size=9, bold=True, color="555555"),
                 fill_obj=fill(BLUE_TABLE),
                 align=center(),
                 number_format='0.0"x"')
        for j, off_x in enumerate(exit_mults):
            col = chr(ord("E") + j)
            cell_ref = f"{col}{row}"
            # Sponsor equity = (mult_entree × EBITDA_LTM × 1.075) - new_debt
            #   le 1.075 = fees 2.5% + cash min 5% chargés au sponsor
            # EV exit = EBITDA_Y5 × (mult_sortie)
            # Equity exit = EV exit - Net Debt Y5
            # IRR = (Equity exit / Sponsor equity)^(1/n) - 1
            formula = (
                f"=IFERROR(("
                f"(_LBO_CALC!$M$7*(C24+({off_x}))-_LBO_CALC!$M$30)/"
                f"MAX(1,(INPUT!H30*(C17+({off_e}))*1.075-C19*INPUT!H30))"
                f")^(1/C25)-1,0)"
            )
            set_cell(ws, cell_ref, formula,
                     font_obj=font(size=10, color=NAVY_TEXT),
                     fill_obj=fill(WHITE),
                     border_obj=box_border(color="DDDDDD"),
                     align=center(),
                     number_format='0.0%')

    # Conditional formatting : heatmap rouge → jaune → vert
    color_scale = ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + "F8696B",
        mid_type="num", mid_value=0.15, mid_color="FF" + "FFEB84",
        end_type="num", end_value=0.30, end_color="FF" + "63BE7B",
    )
    ws.conditional_formatting.add("E48:I52", color_scale)

    # ═══════════════════════════════════════════════════════════════════
    # SECTION F — SCÉNARIOS BULL / BASE / BEAR
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 54, "F.  SCÉNARIOS  —  Bull / Base / Bear")

    # Headers
    set_cell(ws, "B55", "Métrique",
             font_obj=font(size=10, bold=True, color="555555"),
             fill_obj=fill(BLUE_TABLE),
             align=left_align(indent=1))
    set_cell(ws, "E55", "BULL",
             font_obj=font(size=10, bold=True, color=GREEN_OK),
             fill_obj=fill(GREEN_PALE),
             align=center())
    set_cell(ws, "G55", "BASE",
             font_obj=font(size=10, bold=True, color=NAVY_TEXT),
             fill_obj=fill(BLUE_PALE),
             align=center())
    set_cell(ws, "I55", "BEAR",
             font_obj=font(size=10, bold=True, color=RED_KO),
             fill_obj=fill(RED_PALE),
             align=center())

    # Lignes scénarios — référencent _LBO_CALC
    scenarios = [
        ("B56", "IRR Sponsor",       "E56", "=_LBO_CALC!P55", "G56", "=_LBO_CALC!P56", "I56", "=_LBO_CALC!P57", '0.0%'),
        ("B57", "MOIC",              "E57", "=_LBO_CALC!Q55", "G57", "=_LBO_CALC!Q56", "I57", "=_LBO_CALC!Q57", '0.00"x"'),
        ("B58", "Leverage exit",     "E58", "=_LBO_CALC!R55", "G58", "=_LBO_CALC!R56", "I58", "=_LBO_CALC!R57", '0.0"x"'),
        ("B59", "ICR exit (EBIT/Int)", "E59", "=_LBO_CALC!S55", "G59", "=_LBO_CALC!S56", "I59", "=_LBO_CALC!S57", '0.0"x"'),
    ]
    for lbl_ref, lbl, ec, ef, gc, gf, ic, ifr, fmt in scenarios:
        set_cell(ws, lbl_ref, lbl, font_obj=font(color=NAVY_TEXT),
                 align=left_align(indent=1))
        set_cell(ws, ec, ef, font_obj=font(bold=True, color=GREEN_OK),
                 align=center(), number_format=fmt)
        set_cell(ws, gc, gf, font_obj=font(bold=True, color=NAVY_TEXT),
                 align=center(), number_format=fmt)
        set_cell(ws, ic, ifr, font_obj=font(bold=True, color=RED_KO),
                 align=center(), number_format=fmt)

    # ═══════════════════════════════════════════════════════════════════
    # SECTION G — ZONE EXPORT (cellules nommées pour Python)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 61, "G.  EXPORT  —  Données reprises dans le PPTX et le PDF")

    exports = [
        ("B62", "LBO_ELIGIBLE", "C62",
         '=IF(AND(INPUT!H30>0,'
         'IFERROR(INPUT!H30/INPUT!H9,0)>0.15,'
         'IFERROR(DCF!H49/INPUT!H30,0)>0.6,'
         'IFERROR((INPUT!H54+INPUT!H48-INPUT!H34)/INPUT!H30,99)<3.5),'
         '"OUI","NON")',
         "@"),
        ("B63", "LBO_MEGA_FLAG", "C63",
         '=IF(INPUT!H98>100000000,"theorique",IF(INPUT!H98>50000000,"mega","standard"))',
         "@"),
        ("B64", "LBO_IRR_BASE",        "C64", "=_LBO_CALC!N51", "0.000%"),
        ("B65", "LBO_MOIC_BASE",       "C65", "=_LBO_CALC!N52", "0.000"),
        ("B66", "LBO_IRR_BULL",        "C66", "=_LBO_CALC!P55", "0.000%"),
        ("B67", "LBO_IRR_BEAR",        "C67", "=_LBO_CALC!P57", "0.000%"),
        ("B68", "LBO_LEVERAGE_EXIT",   "C68", "=_LBO_CALC!R56", "0.000"),
        ("B69", "LBO_EQUITY_ENTRY",    "C69", "=D31",            "#,##0"),
        ("B70", "LBO_EQUITY_EXIT",     "C70", "=_LBO_CALC!N50", "#,##0"),
    ]
    for lbl_ref, lbl, val_ref, val_formula, fmt in exports:
        set_cell(ws, lbl_ref, lbl,
                 font_obj=font(size=9, color="555555"),
                 fill_obj=fill(EXPORT_PALE),
                 align=left_align(indent=1))
        set_cell(ws, val_ref, val_formula,
                 font_obj=font(size=9, bold=True, color=NAVY_TEXT),
                 fill_obj=fill(EXPORT_PALE),
                 align=right_align(),
                 number_format=fmt)

    # ─── Création des cellules nommées (defined names) ───
    cell_name_map = {
        "LBO_ELIGIBLE":      "C62",
        "LBO_MEGA_FLAG":     "C63",
        "LBO_IRR_BASE":      "C64",
        "LBO_MOIC_BASE":     "C65",
        "LBO_IRR_BULL":      "C66",
        "LBO_IRR_BEAR":      "C67",
        "LBO_LEVERAGE_EXIT": "C68",
        "LBO_EQUITY_ENTRY":  "C69",
        "LBO_EQUITY_EXIT":   "C70",
    }
    for name, cell_ref in cell_name_map.items():
        defn = DefinedName(name=name, attr_text=f"'LBO MODEL'!${cell_ref[0]}${cell_ref[1:]}")
        wb.defined_names[name] = defn

    # ═══════════════════════════════════════════════════════════════════
    # FOOTER
    # ═══════════════════════════════════════════════════════════════════
    merge_and_set(
        ws, "B73:M73",
        "Méthodologie complète : voir slide méthodologie société. "
        "Calculs détaillés dans l'onglet _LBO_CALC (masqué — démasquer pour audit).",
        font_obj=font(size=8, italic=True, color="888888"),
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )

    # Hauteur ligne section headers
    for r in [6, 16, 29, 37, 46, 54, 61]:
        ws.row_dimensions[r].height = 20.0

    return ws


def section_header(ws, row, title, end_col="M"):
    """Header de section style FinSight (bleu sur fond bleu pâle)."""
    merge_and_set(
        ws, f"B{row}:{end_col}{row}", title,
        font_obj=font(size=10, bold=True, color=BLUE_SECTION),
        fill_obj=fill(BLUE_PALE),
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )


# ═════════════════════════════════════════════════════════════════════════════
# CONSTRUCTION ONGLET MASQUÉ "_LBO_CALC"
# ═════════════════════════════════════════════════════════════════════════════

def build_lbo_calc(wb):
    """Onglet masqué : tous les calculs lourds JPM/GS-grade."""
    if "_LBO_CALC" in wb.sheetnames:
        del wb["_LBO_CALC"]
    ws = wb.create_sheet("_LBO_CALC")

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 2.0
    ws.column_dimensions["B"].width = 36.0
    for col in "CDEFGHIJKLMN":
        ws.column_dimensions[col].width = 11.0
    for col in "OPQRS":
        ws.column_dimensions[col].width = 12.0

    # ─── Header ───
    merge_and_set(
        ws, "B1:N1", "FinSight  —  LBO Calculation Engine (JPM/GS-grade)",
        font_obj=font(size=14, bold=True, color=WHITE),
        fill_obj=fill(NAVY_DARK),
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[1].height = 28.0

    merge_and_set(
        ws, "B2:N2",
        "Onglet masqué — démasquer pour audit complet du modèle.",
        font_obj=font(size=9, italic=True, color="555555"),
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )

    # ═══════════════════════════════════════════════════════════════════
    # PROJECTIONS Y0-Y5 (référencent DCF)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 4, "1.  PROJECTIONS DÉTAILLÉES Y0-Y5  (référence DCF)", end_col="L")

    # Headers années
    set_cell(ws, "B5", "($000)",
             font_obj=font(size=9, italic=True, color="888888"))
    years_label = ["Y0 (LTM)", "Y1", "Y2", "Y3", "Y4", "Y5"]
    cols_dcf = ["H", "I", "J", "K", "L", "M"]   # colonnes DCF correspondantes
    cols_calc = ["H", "I", "J", "K", "L", "M"]
    for i, lbl in enumerate(years_label):
        set_cell(ws, f"{cols_calc[i]}5", lbl,
                 font_obj=font(size=10, bold=True, color="555555"),
                 fill_obj=fill(BLUE_TABLE),
                 align=center())

    proj_lines = [
        ("Revenue",            "B6",  "DCF!{c}9"),
        ("EBITDA",             "B7",  "DCF!{c}15"),
        ("D&A",                "B8",  "DCF!{c}17"),
        ("EBIT",               "B9",  "DCF!{c}18"),
        ("CapEx",              "B10", "DCF!{c}47"),
        ("Δ Working Capital",  "B11", "DCF!{c}48"),
        ("FCF Unlevered (Pre-Debt)", "B12", "DCF!{c}49"),
    ]
    for label, lbl_ref, formula_template in proj_lines:
        set_cell(ws, lbl_ref, label,
                 font_obj=font(size=10, color=NAVY_TEXT))
        for i, c_dcf in enumerate(cols_dcf):
            ref = f"{cols_calc[i]}{lbl_ref[1:]}"
            set_cell(ws, ref, "=" + formula_template.format(c=c_dcf),
                     font_obj=font(size=10, color=NAVY_TEXT),
                     align=right_align(),
                     number_format='#,##0_);[Red](#,##0);"—"',
                     fill_obj=fill(GREY_CALC))

    # ═══════════════════════════════════════════════════════════════════
    # DEBT SCHEDULE (lignes 14-32)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 14, "2.  DEBT SCHEDULE COMPLET (Senior TLB + Mezzanine)", end_col="L")

    # Sub-section : Senior TLB
    set_cell(ws, "B15", "SENIOR TERM LOAN B",
             font_obj=font(size=9, bold=True, color=BLUE_SECTION))

    set_cell(ws, "B16", "Beginning balance",
             font_obj=font(size=9, color=NAVY_TEXT))
    # Y0 : initial = leverage senior * EBITDA
    set_cell(ws, "H16", "='LBO MODEL'!C20*INPUT!H30",
             font_obj=font(size=9, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')
    # Y1-Y5 : ref année précédente
    for i in range(1, 6):
        c_prev = cols_calc[i - 1]
        c_curr = cols_calc[i]
        set_cell(ws, f"{c_curr}16", f"={c_prev}19",  # ending balance previous
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0')

    set_cell(ws, "B17", "Mandatory amortization (1%)",
             font_obj=font(size=9, color="555555"))
    for i in range(1, 6):  # Y1-Y5
        c = cols_calc[i]
        set_cell(ws, f"{c}17", f"=-MIN({c}16*0.01,{c}16)",
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0')

    set_cell(ws, "B18", "Cash sweep (75% excess FCF)",
             font_obj=font(size=9, color="555555"))
    for i in range(1, 6):
        c = cols_calc[i]
        # FCF after interest = FCF unlev - interest - mandatory amort
        # Cash sweep = 75% du FCF positif disponible
        formula = (
            f"=-MIN({c}16+{c}17,"
            f"MAX(0,({c}12+{c}20)*0.75))"
        )
        set_cell(ws, f"{c}18", formula,
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0')

    set_cell(ws, "B19", "Ending balance",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT))
    set_cell(ws, "H19", "=H16",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             border_obj=border(top=True),
             number_format='#,##0')
    for i in range(1, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}19", f"={c}16+{c}17+{c}18",
                 font_obj=font(size=9, bold=True, color=NAVY_TEXT),
                 align=right_align(),
                 border_obj=border(top=True),
                 number_format='#,##0')

    set_cell(ws, "B20", "Interest expense (TLB)",
             font_obj=font(size=9, color="555555"))
    for i in range(1, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}20", f"=-{c}16*'LBO MODEL'!$C$22",
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0')

    # Sub-section : Mezzanine
    set_cell(ws, "B22", "MEZZANINE",
             font_obj=font(size=9, bold=True, color=BLUE_SECTION))

    set_cell(ws, "B23", "Beginning balance",
             font_obj=font(size=9, color=NAVY_TEXT))
    set_cell(ws, "H23", "='LBO MODEL'!C21*INPUT!H30",
             font_obj=font(size=9, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')
    for i in range(1, 6):
        c_prev = cols_calc[i - 1]
        c_curr = cols_calc[i]
        set_cell(ws, f"{c_curr}23", f"={c_prev}26",
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0')

    set_cell(ws, "B24", "PIK interest (5% capitalisé)",
             font_obj=font(size=9, color="555555"))
    for i in range(1, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}24", f"={c}23*0.05",
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0')

    set_cell(ws, "B25", "Bullet repayment (Y5)",
             font_obj=font(size=9, color="555555"))
    set_cell(ws, "M25", "=-(M23+M24)",
             font_obj=font(size=9, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    set_cell(ws, "B26", "Ending balance",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT))
    set_cell(ws, "H26", "=H23",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             border_obj=border(top=True),
             number_format='#,##0')
    for i in range(1, 6):
        c = cols_calc[i]
        if i < 5:
            set_cell(ws, f"{c}26", f"={c}23+{c}24",
                     font_obj=font(size=9, bold=True, color=NAVY_TEXT),
                     align=right_align(),
                     border_obj=border(top=True),
                     number_format='#,##0')
        else:
            set_cell(ws, f"{c}26", f"={c}23+{c}24+{c}25",
                     font_obj=font(size=9, bold=True, color=NAVY_TEXT),
                     align=right_align(),
                     border_obj=border(top=True),
                     number_format='#,##0')

    set_cell(ws, "B27", "Interest expense Mezz (cash 5%)",
             font_obj=font(size=9, color="555555"))
    for i in range(1, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}27", f"=-{c}23*0.05",
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='#,##0')

    # Total interest + Total Net Debt
    set_cell(ws, "B29", "TOTAL INTEREST EXPENSE",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT))
    for i in range(1, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}29", f"={c}20+{c}27",
                 font_obj=font(size=9, bold=True, color=NAVY_TEXT),
                 align=right_align(),
                 fill_obj=fill(GREY_CALC),
                 number_format='#,##0')

    set_cell(ws, "B30", "Net Debt (Senior + Mezz - Cash)",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT))
    set_cell(ws, "H30", "=H19+H26",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             fill_obj=fill(GREY_CALC),
             number_format='#,##0')
    for i in range(1, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}30", f"={c}19+{c}26",
                 font_obj=font(size=9, bold=True, color=NAVY_TEXT),
                 align=right_align(),
                 fill_obj=fill(GREY_CALC),
                 number_format='#,##0')

    set_cell(ws, "B31", "Total Leverage (× EBITDA)",
             font_obj=font(size=9, color="555555"))
    for i in range(0, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}31", f"=IFERROR({c}30/{c}7,0)",
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='0.00"x"')

    set_cell(ws, "B32", "Interest Coverage (EBIT/Int)",
             font_obj=font(size=9, color="555555"))
    for i in range(1, 6):
        c = cols_calc[i]
        set_cell(ws, f"{c}32", f"=IFERROR(-{c}9/{c}29,0)",
                 font_obj=font(size=9, color=NAVY_TEXT),
                 align=right_align(),
                 number_format='0.0"x"')

    # ═══════════════════════════════════════════════════════════════════
    # EXIT WATERFALL & RETURNS (lignes 34-52)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 34, "3.  EXIT WATERFALL & RETURNS BASE CASE", end_col="L")

    set_cell(ws, "B35", "EBITDA exit (Y5)",
             font_obj=font(size=9, color="555555"))
    set_cell(ws, "M35", "=M7",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    set_cell(ws, "B36", "× Multiple de sortie",
             font_obj=font(size=9, color="555555"))
    set_cell(ws, "M36", "='LBO MODEL'!$C$24",
             font_obj=font(size=9, color=NAVY_TEXT),
             align=right_align(),
             number_format='0.0"x"')

    set_cell(ws, "B37", "= EV exit",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT))
    set_cell(ws, "M37", "=M35*M36",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             border_obj=border(top=True),
             number_format='#,##0')

    set_cell(ws, "B38", "− Net Debt at exit",
             font_obj=font(size=9, color="555555"))
    set_cell(ws, "M38", "=-M30",
             font_obj=font(size=9, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0;(#,##0)')

    set_cell(ws, "B39", "= Equity exit",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT))
    set_cell(ws, "M39", "=M37+M38",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             border_obj=border(top=True, bottom=True),
             fill_obj=fill(GREY_CALC),
             number_format='#,##0')

    # Summary KPIs en N48-N52
    set_cell(ws, "M45", "EQUITY ENTRY",
             font_obj=font(size=9, bold=True, color="555555"),
             align=center())
    set_cell(ws, "N45", "='LBO MODEL'!D31",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    set_cell(ws, "M48", "EBITDA Y5",
             font_obj=font(size=9, bold=True, color="555555"))
    set_cell(ws, "N48", "=M35",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    set_cell(ws, "M49", "NET DEBT Y5",
             font_obj=font(size=9, bold=True, color="555555"))
    set_cell(ws, "N49", "=M30",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    set_cell(ws, "M50", "EQUITY EXIT",
             font_obj=font(size=9, bold=True, color="555555"))
    set_cell(ws, "N50", "=M39",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='#,##0')

    set_cell(ws, "M51", "IRR BASE",
             font_obj=font(size=9, bold=True, color="555555"))
    set_cell(ws, "N51", "=IFERROR((N50/N45)^(1/'LBO MODEL'!C25)-1,0)",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='0.0%')

    set_cell(ws, "M52", "MOIC BASE",
             font_obj=font(size=9, bold=True, color="555555"))
    set_cell(ws, "N52", "=IFERROR(N50/N45,0)",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(),
             number_format='0.00"x"')

    # ═══════════════════════════════════════════════════════════════════
    # SCÉNARIOS BULL/BASE/BEAR (cellules P55-S57)
    # ═══════════════════════════════════════════════════════════════════
    section_header(ws, 54, "4.  SCÉNARIOS BULL / BASE / BEAR (formules paramétrées)", end_col="L")

    set_cell(ws, "M54", "Scénario", font_obj=font(size=9, bold=True, color="555555"),
             fill_obj=fill(BLUE_TABLE), align=center())
    set_cell(ws, "P54", "IRR", font_obj=font(size=9, bold=True, color="555555"),
             fill_obj=fill(BLUE_TABLE), align=center())
    set_cell(ws, "Q54", "MOIC", font_obj=font(size=9, bold=True, color="555555"),
             fill_obj=fill(BLUE_TABLE), align=center())
    set_cell(ws, "R54", "Lvg exit", font_obj=font(size=9, bold=True, color="555555"),
             fill_obj=fill(BLUE_TABLE), align=center())
    set_cell(ws, "S54", "ICR exit", font_obj=font(size=9, bold=True, color="555555"),
             fill_obj=fill(BLUE_TABLE), align=center())

    # BULL : revenue +15%, marge +200bps, exit +1x
    # On approxime IRR bull = IRR base + delta lié aux ajustements
    set_cell(ws, "M55", "BULL (rev +15%, marge +200bps, exit +1x)",
             font_obj=font(size=9, color=GREEN_OK))
    set_cell(ws, "P55", "=IFERROR(((N48*1.15*(M36+1)-N49*0.7)/N45)^(1/'LBO MODEL'!C25)-1,0)",
             font_obj=font(size=9, bold=True, color=GREEN_OK),
             align=right_align(), number_format='0.0%')
    set_cell(ws, "Q55", "=IFERROR((N48*1.15*(M36+1)-N49*0.7)/N45,0)",
             font_obj=font(size=9, bold=True, color=GREEN_OK),
             align=right_align(), number_format='0.00"x"')
    set_cell(ws, "R55", "=M31*0.5",
             font_obj=font(size=9, bold=True, color=GREEN_OK),
             align=right_align(), number_format='0.0"x"')
    set_cell(ws, "S55", "=M32*1.5",
             font_obj=font(size=9, bold=True, color=GREEN_OK),
             align=right_align(), number_format='0.0"x"')

    set_cell(ws, "M56", "BASE (hypothèses centrales)",
             font_obj=font(size=9, color=NAVY_TEXT))
    set_cell(ws, "P56", "=N51",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(), number_format='0.0%')
    set_cell(ws, "Q56", "=N52",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(), number_format='0.00"x"')
    set_cell(ws, "R56", "=M31",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(), number_format='0.0"x"')
    set_cell(ws, "S56", "=M32",
             font_obj=font(size=9, bold=True, color=NAVY_TEXT),
             align=right_align(), number_format='0.0"x"')

    set_cell(ws, "M57", "BEAR (rev -15%, marge -200bps, exit -1x)",
             font_obj=font(size=9, color=RED_KO))
    set_cell(ws, "P57", "=IFERROR(((N48*0.85*(M36-1)-N49*1.3)/N45)^(1/'LBO MODEL'!C25)-1,0)",
             font_obj=font(size=9, bold=True, color=RED_KO),
             align=right_align(), number_format='0.0%')
    set_cell(ws, "Q57", "=IFERROR((N48*0.85*(M36-1)-N49*1.3)/N45,0)",
             font_obj=font(size=9, bold=True, color=RED_KO),
             align=right_align(), number_format='0.00"x"')
    set_cell(ws, "R57", "=M31*1.8",
             font_obj=font(size=9, bold=True, color=RED_KO),
             align=right_align(), number_format='0.0"x"')
    set_cell(ws, "S57", "=M32*0.5",
             font_obj=font(size=9, bold=True, color=RED_KO),
             align=right_align(), number_format='0.0"x"')

    # ═══════════════════════════════════════════════════════════════════
    # MASQUER L'ONGLET
    # ═══════════════════════════════════════════════════════════════════
    ws.sheet_state = "hidden"

    return ws


# ═════════════════════════════════════════════════════════════════════════════
# API PUBLIQUE — helper unique pour excel_writer.py
# ═════════════════════════════════════════════════════════════════════════════

def build_lbo_sheets(wb):
    """Construit les 2 feuilles LBO MODEL + _LBO_CALC dans un workbook openpyxl
    deja charge. A appeler avant wb.save().

    Pre-requis : le workbook doit contenir les feuilles INPUT et DCF du template
    FinSight (avec leurs structures usuelles).

    Idempotent : si les feuilles existent deja, elles sont remplacees.
    """
    build_lbo_visible(wb)
    build_lbo_calc(wb)
    return wb


# ═════════════════════════════════════════════════════════════════════════════
# LECTURE DES EXPORTS LBO (utilise par pptx_writer.py / pdf_writer.py)
# ═════════════════════════════════════════════════════════════════════════════

def read_lbo_data(xlsx_path) -> dict:
    """Lit les valeurs des cellules nommees LBO_* depuis un fichier Excel
    deja calcule (apres ouverture par Excel/LibreOffice qui resoud les
    formules en cached values).

    Args:
        xlsx_path: chemin vers le fichier .xlsx genere par excel_writer.py
                   (doit avoir ete ouvert au moins une fois par Excel pour
                   que les formules soient calculees, ou utiliser COM)

    Returns:
        dict avec les cles LBO_ELIGIBLE, LBO_MEGA_FLAG, LBO_IRR_BASE,
        LBO_MOIC_BASE, LBO_IRR_BULL, LBO_IRR_BEAR, LBO_LEVERAGE_EXIT,
        LBO_EQUITY_ENTRY, LBO_EQUITY_EXIT, plus eligible (bool) et
        mega_flag (str: standard/mega/theorique).

        Si la feuille LBO MODEL n'existe pas ou si une lecture echoue,
        retourne {"eligible": False, ...} avec valeurs None.
    """
    from pathlib import Path
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"eligible": False}

    path = Path(xlsx_path)
    if not path.exists():
        return {"eligible": False}

    try:
        wb = load_workbook(str(path), data_only=True)
        if "LBO MODEL" not in wb.sheetnames:
            return {"eligible": False}
        ws = wb["LBO MODEL"]

        # Lecture par references directes (cellules nommees peuvent ne pas
        # etre resolues par openpyxl en data_only mode)
        cell_map = {
            "LBO_ELIGIBLE":      "C62",
            "LBO_MEGA_FLAG":     "C63",
            "LBO_IRR_BASE":      "C64",
            "LBO_MOIC_BASE":     "C65",
            "LBO_IRR_BULL":      "C66",
            "LBO_IRR_BEAR":      "C67",
            "LBO_LEVERAGE_EXIT": "C68",
            "LBO_EQUITY_ENTRY":  "C69",
            "LBO_EQUITY_EXIT":   "C70",
        }
        out = {}
        for name, ref in cell_map.items():
            try:
                out[name] = ws[ref].value
            except Exception:
                out[name] = None

        # Champs derives pratiques
        out["eligible"] = (str(out.get("LBO_ELIGIBLE", "")).strip().upper() == "OUI")
        out["mega_flag"] = str(out.get("LBO_MEGA_FLAG", "standard") or "standard")
        out["irr_base"] = out.get("LBO_IRR_BASE")
        out["moic_base"] = out.get("LBO_MOIC_BASE")
        out["irr_bull"] = out.get("LBO_IRR_BULL")
        out["irr_bear"] = out.get("LBO_IRR_BEAR")
        out["leverage_exit"] = out.get("LBO_LEVERAGE_EXIT")
        out["equity_entry"] = out.get("LBO_EQUITY_ENTRY")
        out["equity_exit"] = out.get("LBO_EQUITY_EXIT")

        # Profil société (depuis LBO MODEL!C7/C8)
        try:
            out["company_name"] = ws["C7"].value
            out["ticker"] = ws["C8"].value
        except Exception:
            pass

        return out
    except Exception:
        return {"eligible": False}


# ═════════════════════════════════════════════════════════════════════════════
# MAIN (script standalone pour audit/iteration)
# ═════════════════════════════════════════════════════════════════════════════

def main():
    args = sys.argv[1:]
    if args:
        path = Path(args[0])
    else:
        path = Path("assets/_LBO_WORKING_AAPL.xlsx")

    if not path.exists():
        print(f"[ERR] Fichier introuvable : {path}")
        sys.exit(1)

    print(f"=" * 70)
    print(f"  BUILD LBO SHEETS — {path}")
    print(f"=" * 70)

    print("[1/4] Chargement workbook...")
    wb = load_workbook(str(path))
    print(f"      Sheets actuels : {wb.sheetnames}")

    print("[2/4] Construction LBO MODEL (visible)...")
    ws_visible = build_lbo_visible(wb)
    print(f"      OK ({ws_visible.max_row} lignes × {ws_visible.max_column} cols)")

    print("[3/4] Construction _LBO_CALC (masqué)...")
    ws_calc = build_lbo_calc(wb)
    print(f"      OK ({ws_calc.max_row} lignes × {ws_calc.max_column} cols)")

    print("[4/4] Sauvegarde...")
    wb.save(str(path))
    print(f"      OK -> {path}")
    print()
    print(f"Total sheets : {len(wb.sheetnames)}")
    print(f"Defined names : {list(wb.defined_names)}")
    print()
    print("=" * 70)
    print("  DONE")
    print("=" * 70)


if __name__ == "__main__":
    main()
