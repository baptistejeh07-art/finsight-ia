# -*- coding: utf-8 -*-
"""
outputs/excel_profile_router.py — Dispatcher XLSX société par profil sectoriel.

Sélectionne le template XLSX approprié en fonction du profil sectoriel détecté
(STANDARD / BANK / INSURANCE / REIT / UTILITY / OIL_GAS) et fournit le cell
mapping correspondant pour injecter les données du FinancialSnapshot.

ARCHITECTURE :

1. `detect_profile(sector, industry)` (déjà dans core.sector_profiles) retourne
   le profil métier à partir du secteur yfinance + industry.

2. `TEMPLATES[profile]` retourne le path vers le fichier xlsx de template.
   Si le template profil-spécifique n'existe pas, fallback transparent vers
   STANDARD (`assets/TEMPLATE.xlsx`). ZÉRO RÉGRESSION pour les corporates.

3. `CELL_MAPS[profile]` retourne un dict qui décrit où injecter chaque champ
   du snapshot dans le template. Structure :
       {
           "company_info": {"D5": "ticker", "D7": "company_name", ...},
           "is_rows":      {"revenue": 9, "cogs": 10, ...},
           "bs_asset_rows": {...},
           "bs_liab_rows":  {...},
           "cf_rows":       {...},
           "mkt_fields":    {"H95": "share_price", ...},
       }
   Pour STANDARD, c'est une copie des constantes existantes dans excel_writer.py.
   Pour les profils non-STANDARD, c'est vide tant que Baptiste n'a pas envoyé
   le template + la liste des coordonnées.

4. `FORMULA_CELLS_BY_PROFILE[profile]` est un SET de cell refs ("D120", "H108", ...)
   détectées automatiquement au CHARGEMENT du template : scan de toutes les
   cellules dont la valeur commence par "=". Le cache est construit une fois
   par profil au premier usage (lazy). Utilisé par `is_formula_cell_for_profile`
   pour que `_safe_write` ne les écrase jamais.

CONVENTION DE NOMMAGE (Baptiste Option B) :
    assets/TEMPLATE.xlsx                                  — STANDARD (existant)
    assets/FinSight_IA_Template_BANK_v1.xlsx              — à venir
    assets/FinSight_IA_Template_INSURANCE_v1.xlsx         — à venir
    assets/FinSight_IA_Template_REIT_v1.xlsx              — à venir
    assets/FinSight_IA_Template_UTILITY_v1.xlsx           — à venir
    assets/FinSight_IA_Template_OIL_GAS_v1.xlsx           — à venir

USAGE :
    from outputs.excel_profile_router import (
        get_template_for, get_cell_map, get_formula_cells_for,
    )

    profile = detect_profile(snapshot.company_info.sector,
                             getattr(snapshot.company_info, "industry", ""))
    tpl_path = get_template_for(profile)
    cell_map = get_cell_map(profile)
    formula_cells = get_formula_cells_for(profile, tpl_path)

    # Injection dans ExcelWriter.write() selon le cell_map
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)

_ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"


# ═════════════════════════════════════════════════════════════════════════════
# TEMPLATES — dispatcher profil → fichier
# ═════════════════════════════════════════════════════════════════════════════

# STANDARD est le template actuel (TEMPLATE.xlsx), les autres sont ajoutés
# quand Baptiste livre le fichier correspondant.
TEMPLATES: dict[str, str] = {
    "STANDARD":  "TEMPLATE.xlsx",
    "BANK":      "FinSight_IA_Template_BANK_v1.xlsx",
    "INSURANCE": "FinSight_IA_Template_INSURANCE_v1.xlsx",
    "REIT":      "FinSight_IA_Template_REIT_v1.xlsx",
    "UTILITY":   "FinSight_IA_Template_UTILITY_v1.xlsx",
    "OIL_GAS":   "FinSight_IA_Template_OIL_GAS_v1.xlsx",
}


def get_template_for(profile: str) -> Path:
    """Retourne le path absolu du template pour un profil donné.

    Si le template profil-spécifique n'existe pas dans `assets/`, fallback
    transparent vers STANDARD. Log un warning au premier fallback pour
    diagnostic mais ne casse PAS la génération.

    Args:
        profile : STANDARD | BANK | INSURANCE | REIT | UTILITY | OIL_GAS

    Returns:
        Path absolu vers le .xlsx à charger.
    """
    _tpl_name = TEMPLATES.get(profile) or TEMPLATES["STANDARD"]
    _path = _ASSETS_DIR / _tpl_name

    if _path.exists():
        log.debug(f"[excel_profile_router] profil={profile} → {_tpl_name}")
        return _path

    # Fallback STANDARD
    if profile != "STANDARD":
        log.warning(
            f"[excel_profile_router] Template profil '{profile}' introuvable "
            f"({_tpl_name}) — fallback STANDARD"
        )
    _fallback = _ASSETS_DIR / TEMPLATES["STANDARD"]
    if not _fallback.exists():
        raise FileNotFoundError(
            f"Template STANDARD introuvable : {_fallback}. "
            "Vérifier assets/TEMPLATE.xlsx"
        )
    return _fallback


def profile_template_exists(profile: str) -> bool:
    """True si un template profil-spécifique existe dans assets/ (ne compte
    PAS le fallback STANDARD)."""
    if profile == "STANDARD":
        return (_ASSETS_DIR / TEMPLATES["STANDARD"]).exists()
    _tpl_name = TEMPLATES.get(profile)
    if not _tpl_name:
        return False
    return (_ASSETS_DIR / _tpl_name).exists()


# ═════════════════════════════════════════════════════════════════════════════
# CELL MAPS — où injecter chaque champ du snapshot dans chaque template
# ═════════════════════════════════════════════════════════════════════════════
#
# Structure unifiée : pour chaque profil, on décrit les coordonnées des
# champs à injecter. STANDARD est peuplé avec les constantes existantes
# de excel_writer.py (source of truth = _IS_ROWS, _BS_*, _CF_ROWS, _MKT_FIELDS).
#
# Pour BANK / INSURANCE / etc. : dict vide pour l'instant. Sera rempli
# quand Baptiste livre le template + la liste des coordonnées.

# Cell refs (ligne, col, cellule entière pour market data) par profil.
# Les lignes sont des entiers (feuille INPUT), les colonnes sont déterminées
# dynamiquement par _build_year_col (D=N-4, E=N-3, ..., H=N).

_STANDARD_CELL_MAP: dict = {
    "company_info": {
        # Cellules fixes colonne D — voir config.excel_mapping.COMPANY_INFO
        # ("company_name" -> "D3", "ticker" -> "D4", etc.)
        # Ces valeurs sont déjà gérées dans excel_writer.py via _CI_CELLS.
        # On les réutilise telles quelles en les important.
        "_ref": "config.excel_mapping.COMPANY_INFO",
    },
    "is_rows": {
        "revenue":          9,
        "cogs":             10,
        "sga":              14,
        "rd":               15,
        "da":               16,
        "interest_expense": 20,
        "interest_income":  21,
        "tax_expense_real": 25,
        "dividends":        29,
    },
    "is_cost_fields": {
        # Champs IS à écrire en négatif (coûts attendus <0 par les formules Excel)
        "cogs", "sga", "rd", "da", "interest_expense",
        "tax_expense_real", "dividends",
    },
    "bs_asset_rows": {
        "cash":                 34,
        "accounts_receivable":  35,
        "inventories":          36,
        "other_current_assets": 37,
        "ppe_net":              41,
        "intangibles":          42,
        "other_lt_assets":      43,
    },
    "bs_liab_rows": {
        "accounts_payable":      47,
        "short_term_debt":       48,
        "income_tax_payable":    49,
        "other_current_liab":    50,
        "long_term_debt":        54,
        "common_equity_paid_in": 58,
        "retained_earnings_yf":  59,
    },
    "cf_rows": {
        "change_accounts_receivable": 71,
        "change_inventories":         72,
        "change_accounts_payable":    73,
        "other_wc_changes":           74,
        "capex":                      78,
        "other_investing":            79,
        "change_lt_debt":             83,
        "change_common_equity":       84,
        "dividends_paid":             85,
        "beginning_cash":             90,
    },
    "mkt_fields": {
        # cell_ref absolue (col+row) -> champ MarketData
        "share_price":          "H95",
        "shares_diluted":       "H96",
        "beta_levered":         "H99",
        "risk_free_rate":       "H100",
        "erp":                  "H101",
        "cost_of_debt_pretax":  "H103",
        "tax_rate":             "H104",
        "weight_equity":        "H106",
        "weight_debt":          "H107",
        # H108 WACC = formule Excel — NE PAS écraser
        "terminal_growth":      "H109",
        "days_in_period":       "H110",
    },
    # Row tax rate effectif par année (row 104 col E-H, calculé dynamiquement)
    "tax_rate_row": 104,
    # Sheet principale
    "input_sheet": "INPUT",
    # Sheet historique cours
    "stock_sheet": "STOCK_DATA",
}


# Placeholders — vides tant que Baptiste n'a pas livré les templates.
# Quand il envoie TEMPLATE_BANK avec la liste des coordonnées :
#   1. Copier _STANDARD_CELL_MAP comme base
#   2. Remplacer/adapter les keys qui diffèrent (ex: NII → ligne 9 au lieu
#      de revenue, provisions → ligne 14, etc.)
#   3. Supprimer les champs non pertinents (gross_profit pour une banque)
#   4. Ajouter les champs spécifiques (CET1, NPL, NIM, Cost/Income...)

_BANK_CELL_MAP: dict = {}
_INSURANCE_CELL_MAP: dict = {}
_REIT_CELL_MAP: dict = {
    "company_info": {
        "company_name":  "D136",
        "ticker":        "D137",
        "sector":        "D138",
        "base_year":     "D139",
        "currency":      "D140",
        "units":         "D141",
        "analysis_date": "D142",
    },
    "is_rows": {
        "revenue":          9,
        "cogs":             10,
        "sga":              14,
        "rd":               15,
        "da":               16,
        "interest_expense": 20,
        "interest_income":  21,
        "tax_expense_real": 25,
        "dividends":        29,
    },
    "is_cost_fields": {
        "cogs", "sga", "rd", "da", "interest_expense",
        "tax_expense_real", "dividends",
    },
    "bs_asset_rows": {
        "cash":                 34,
        "accounts_receivable":  35,
        "inventories":          36,
        "other_current_assets": 37,
        "ppe_net":              41,
        "intangibles":          42,
        "other_lt_assets":      43,
    },
    "bs_liab_rows": {
        "accounts_payable":      47,
        "short_term_debt":       48,
        "income_tax_payable":    49,
        "other_current_liab":    50,
        "long_term_debt":        54,
        "common_equity_paid_in": 58,
        "retained_earnings_yf":  59,
    },
    "cf_rows": {
        "change_accounts_receivable": 71,
        "change_inventories":         72,
        "change_accounts_payable":    73,
        "other_wc_changes":           74,
        "capex":                      78,
        "other_investing":            79,
        "change_lt_debt":             83,
        "change_common_equity":       84,
        "dividends_paid":             85,
        "beginning_cash":             90,
    },
    "mkt_fields": {
        "share_price":          "H117",
        "shares_diluted":       "H118",
        "beta_levered":         "H121",
        "risk_free_rate":       "H122",
        "erp":                  "H123",
        "cost_of_debt_pretax":  "H125",
        "tax_rate":             "H126",
        "weight_equity":        "H128",
        "weight_debt":          "H129",
        # H130 WACC = FORMULE — absent
        "terminal_growth":      "H131",
        "days_in_period":       "H132",
    },
    "tax_rate_row": 126,
    "input_sheet": "INPUT",
    "stock_sheet": "STOCK_DATA",
    # Row pour les labels d'année (backup helpers hors zone impression)
    "year_label_row": 5,
    # Pas de row 132 helper pour REIT (132 = days_in_period)
    "year_helper_row": None,
}
_UTILITY_CELL_MAP: dict = {
    "company_info": {
        "company_name":  "D128",
        "ticker":        "D129",
        "sector":        "D130",
        "base_year":     "D131",
        "currency":      "D132",
        "units":         "D133",
        "analysis_date": "D134",
    },
    "is_rows": {
        "revenue":          9,
        "cogs":             10,
        "sga":              14,
        "rd":               15,
        "da":               16,
        "interest_expense": 20,
        "interest_income":  21,
        "tax_expense_real": 25,
        "dividends":        29,
    },
    "is_cost_fields": {
        "cogs", "sga", "rd", "da", "interest_expense",
        "tax_expense_real", "dividends",
    },
    "bs_asset_rows": {
        "cash":                 34,
        "accounts_receivable":  35,
        "inventories":          36,
        "other_current_assets": 37,
        "ppe_net":              41,
        "intangibles":          42,
        "other_lt_assets":      43,
    },
    "bs_liab_rows": {
        "accounts_payable":      47,
        "short_term_debt":       48,
        "income_tax_payable":    49,
        "other_current_liab":    50,
        "long_term_debt":        54,
        "common_equity_paid_in": 58,
        "retained_earnings_yf":  59,
    },
    "cf_rows": {
        "change_accounts_receivable": 71,
        "change_inventories":         72,
        "change_accounts_payable":    73,
        "other_wc_changes":           74,
        "capex":                      78,
        "other_investing":            79,
        "change_lt_debt":             83,
        "change_common_equity":       84,
        "dividends_paid":             85,
        "beginning_cash":             90,
    },
    "mkt_fields": {
        "share_price":          "H109",
        "shares_diluted":       "H110",
        "beta_levered":         "H113",
        "risk_free_rate":       "H114",
        "erp":                  "H115",
        "cost_of_debt_pretax":  "H117",
        "tax_rate":             "H118",
        "weight_equity":        "H120",
        "weight_debt":          "H121",
        # H122 WACC = FORMULE — absent
        "terminal_growth":      "H123",
    },
    "tax_rate_row": 118,
    "input_sheet": "INPUT",
    "stock_sheet": "STOCK_DATA",
    "year_label_row": 5,
    # Pas de year_helper_row pour UTILITY (pas de row backup)
    "year_helper_row": None,
}

# OIL_GAS — template TTE.PA livré 2026-04-15 (Baptiste).
# Vérification du fichier : le layout INPUT est IDENTIQUE à STANDARD
# (tax row 25, weight_equity H106, weight_debt H107, terminal_growth H109).
# Le template diffère uniquement par ses feuilles sectorielles SOTP / DCF /
# COMPARABLES. On laisse le cell_map vide → fallback automatique STANDARD
# pour l'INPUT, ce qui permet de reconduire toute la plomberie existante
# sans risque. Les sheets SOTP/DCF/COMPARABLES restent non injectées par
# le writer — à traiter plus tard si automatisation souhaitée.
_OIL_GAS_CELL_MAP: dict = {}


CELL_MAPS: dict[str, dict] = {
    "STANDARD":  _STANDARD_CELL_MAP,
    "BANK":      _BANK_CELL_MAP,
    "INSURANCE": _INSURANCE_CELL_MAP,
    "REIT":      _REIT_CELL_MAP,
    "UTILITY":   _UTILITY_CELL_MAP,
    "OIL_GAS":   _OIL_GAS_CELL_MAP,
}


def get_cell_map(profile: str) -> dict:
    """Retourne le cell map pour un profil.

    Si le profil n'a pas de cell map défini (ex: BANK pas encore livré),
    retourne le cell map STANDARD. Cohérent avec get_template_for() qui
    fallback sur le template STANDARD.
    """
    _cm = CELL_MAPS.get(profile) or {}
    if not _cm or "is_rows" not in _cm:
        if profile != "STANDARD":
            log.debug(
                f"[excel_profile_router] Cell map profil '{profile}' non défini "
                f"— fallback STANDARD"
            )
        return _STANDARD_CELL_MAP
    return _cm


# ═════════════════════════════════════════════════════════════════════════════
# FORMULA CELLS — détection automatique au chargement du template
# ═════════════════════════════════════════════════════════════════════════════
#
# Cache lazy {profile: frozenset(cell_refs_avec_formules)} pour que chaque
# template soit scanné UNE SEULE FOIS. Le résultat sert à _safe_write pour
# ne jamais écraser une formule Excel (e.g. WACC = H108).

_FORMULA_CELLS_CACHE: dict[str, frozenset] = {}


def _scan_formula_cells(tpl_path: Path, sheet_name: str = "INPUT") -> frozenset:
    """Scan toutes les cellules d'une feuille et retourne le set des cell refs
    dont la valeur commence par "=" (formule Excel).

    Retourne frozenset pour immuabilité + hashability.
    """
    try:
        import openpyxl
    except ImportError:
        log.error("[excel_profile_router] openpyxl manquant — scan formules impossible")
        return frozenset()

    try:
        wb = openpyxl.load_workbook(str(tpl_path), data_only=False, keep_links=False)
    except Exception as e:
        log.error(f"[excel_profile_router] Load template échoué : {e}")
        return frozenset()

    if sheet_name not in wb.sheetnames:
        log.warning(
            f"[excel_profile_router] Sheet '{sheet_name}' absente dans {tpl_path.name} "
            f"— scan formules annulé"
        )
        return frozenset()

    ws = wb[sheet_name]
    formula_refs = set()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                formula_refs.add(cell.coordinate)

    log.info(
        f"[excel_profile_router] Scan formules {tpl_path.name} "
        f"sheet={sheet_name} → {len(formula_refs)} cellules protégées"
    )
    return frozenset(formula_refs)


def get_formula_cells_for(profile: str, tpl_path: Optional[Path] = None) -> frozenset:
    """Retourne le frozenset des cell refs contenant une formule pour le
    template du profil donné. Cache au premier usage.

    Args:
        profile  : STANDARD / BANK / ...
        tpl_path : override path (optionnel, sinon appelle get_template_for)

    Returns:
        frozenset({"D120", "E24", "H108", ...}) — cellules à NE PAS écraser.
    """
    if profile in _FORMULA_CELLS_CACHE:
        return _FORMULA_CELLS_CACHE[profile]

    _path = tpl_path or get_template_for(profile)
    _sheet = get_cell_map(profile).get("input_sheet", "INPUT")
    _cells = _scan_formula_cells(_path, _sheet)
    _FORMULA_CELLS_CACHE[profile] = _cells
    return _cells


def is_formula_cell_for_profile(profile: str, cell_ref: str) -> bool:
    """True si la cellule {cell_ref} contient une formule dans le template
    du {profile}. Utilisé par _safe_write pour la protection des formules.

    Exemple :
        if is_formula_cell_for_profile("BANK", "D120"):
            log.debug("Cellule D120 est une formule, skip write")
    """
    return cell_ref in get_formula_cells_for(profile)


def clear_formula_cache() -> None:
    """Clear le cache des formula cells. Utile pour les tests unitaires
    ou quand un template est mis à jour à chaud."""
    _FORMULA_CELLS_CACHE.clear()


# ═════════════════════════════════════════════════════════════════════════════
# API PROFILE DETECTION (wrapper pratique)
# ═════════════════════════════════════════════════════════════════════════════

def detect_profile_from_snapshot(snapshot) -> str:
    """Wrapper pratique : extrait sector + industry du snapshot et appelle
    core.sector_profiles.detect_profile().

    Args:
        snapshot : FinancialSnapshot

    Returns:
        Profil STANDARD | BANK | INSURANCE | REIT | UTILITY | OIL_GAS
    """
    try:
        from core.sector_profiles import detect_profile
    except ImportError:
        log.warning("[excel_profile_router] core.sector_profiles indisponible")
        return "STANDARD"

    try:
        ci = snapshot.company_info
        sector = getattr(ci, "sector", "") or ""
        industry = getattr(ci, "industry", "") or ""
        return detect_profile(sector, industry)
    except Exception as e:
        log.warning(f"[excel_profile_router] detect_profile_from_snapshot failed: {e}")
        return "STANDARD"
