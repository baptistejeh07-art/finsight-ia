"""
indice_comparison_writer.py — FinSight IA
Genere un fichier Excel comparatif pour deux indices boursiers.
Cree le fichier from scratch (openpyxl) — pas de template requis.

Usage:
    buf = IndiceComparisonWriter.generate_bytes(data)   # -> bytes
    path = IndiceComparisonWriter.generate(data, output_path)
"""
from __future__ import annotations

import io
import logging
from datetime import date
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _OXL = True
except ImportError:
    _OXL = False


# =============================================================================
# PALETTE (hex strings pour openpyxl, sans #)
# =============================================================================
_NAVY     = "1B3A6B"
_NAVY_L   = "2A5298"
_COL_A    = "2E5FA3"
_COL_B    = "1A7A4A"
_COL_A_L  = "EEF3FA"
_COL_B_L  = "EAF4EF"
_WHITE    = "FFFFFF"
_BLACK    = "1A1A1A"
_GREY_L   = "F5F7FA"
_GREY_M   = "E8ECF0"
_GREY_T   = "555555"
_BUY_G    = "1A7A4A"
_SELL_R   = "A82020"
_HOLD_A   = "B06000"

# Fills
_FILL_NAVY  = PatternFill("solid", fgColor=_NAVY)   if _OXL else None
_FILL_A     = PatternFill("solid", fgColor=_COL_A)  if _OXL else None
_FILL_B     = PatternFill("solid", fgColor=_COL_B)  if _OXL else None
_FILL_AL    = PatternFill("solid", fgColor=_COL_A_L) if _OXL else None
_FILL_BL    = PatternFill("solid", fgColor=_COL_B_L) if _OXL else None
_FILL_GREY  = PatternFill("solid", fgColor=_GREY_L)  if _OXL else None
_FILL_GREYM = PatternFill("solid", fgColor=_GREY_M)  if _OXL else None

# Fonts
_FONT_TITLE  = Font(name="Calibri", size=14, bold=True, color=_NAVY) if _OXL else None
_FONT_HDR    = Font(name="Calibri", size=9,  bold=True, color=_WHITE) if _OXL else None
_FONT_BODY   = Font(name="Calibri", size=9,  color=_BLACK) if _OXL else None
_FONT_GREY   = Font(name="Calibri", size=8,  color=_GREY_T) if _OXL else None
_FONT_A      = Font(name="Calibri", size=9,  bold=True, color=_COL_A) if _OXL else None
_FONT_B      = Font(name="Calibri", size=9,  bold=True, color=_COL_B) if _OXL else None
_FONT_BUY    = Font(name="Calibri", size=9,  bold=True, color=_BUY_G) if _OXL else None
_FONT_SELL   = Font(name="Calibri", size=9,  bold=True, color=_SELL_R) if _OXL else None
_FONT_HOLD   = Font(name="Calibri", size=9,  bold=True, color=_HOLD_A) if _OXL else None

# Borders
def _thin_border(top=True, bottom=True, left=True, right=True):
    s = Side(border_style="thin", color=_GREY_M)
    n = Side(border_style=None)
    return Border(
        top=s if top else n,
        bottom=s if bottom else n,
        left=s if left else n,
        right=s if right else n,
    )

_BORDER_ALL = _thin_border() if _OXL else None

# Alignments
_AL_LEFT   = Alignment(horizontal="left",   vertical="center") if _OXL else None
_AL_CENTER = Alignment(horizontal="center", vertical="center") if _OXL else None
_AL_RIGHT  = Alignment(horizontal="right",  vertical="center") if _OXL else None


# =============================================================================
# HELPERS
# =============================================================================

def _pct_str(v, signed=False) -> str:
    if v is None:
        return "\u2014"
    try:
        fv = float(v)
        if abs(fv) <= 2:
            fv = fv * 100
        s = f"{fv:+.2f}" if signed else f"{fv:.2f}"
        return s.replace(".", ",") + " %"
    except Exception:
        return "\u2014"


def _num_str(v, dp=2) -> str:
    if v is None:
        return "\u2014"
    try:
        return f"{float(v):.{dp}f}".replace(".", ",")
    except Exception:
        return "\u2014"


def _sig_font(signal: str):
    s = str(signal)
    if "Surp" in s:
        return _FONT_BUY
    if "Sous" in s:
        return _FONT_SELL
    return _FONT_HOLD


def _set(ws, row, col, value, font=None, fill=None, align=None, border=None, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if num_format:
        cell.number_format = num_format
    return cell


def _hdr(ws, row, col, text):
    """Cellule d'en-Tête navy."""
    _set(ws, row, col, text, font=_FONT_HDR, fill=_FILL_NAVY,
         align=_AL_CENTER, border=_BORDER_ALL)


def _hdr_a(ws, row, col, text):
    """En-Tête couleur indice A."""
    _set(ws, row, col, text, font=_FONT_HDR, fill=_FILL_A,
         align=_AL_CENTER, border=_BORDER_ALL)


def _hdr_b(ws, row, col, text):
    """En-Tête couleur indice B."""
    _set(ws, row, col, text, font=_FONT_HDR, fill=_FILL_B,
         align=_AL_CENTER, border=_BORDER_ALL)


def _label(ws, row, col, text):
    """Label gris gauche."""
    _set(ws, row, col, text, font=Font(name="Calibri", size=9, bold=True, color=_GREY_T),
         fill=_FILL_GREY, align=_AL_LEFT, border=_BORDER_ALL)


def _val(ws, row, col, text, alt_row=False):
    """Valeur corps."""
    fill = _FILL_GREYM if alt_row else None
    _set(ws, row, col, text, font=_FONT_BODY, fill=fill,
         align=_AL_CENTER, border=_BORDER_ALL)


def _col_w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def _row_h(ws, row, height):
    ws.row_dimensions[row].height = height


def _merge_title(ws, row, col1, col2, text, big=False):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1, value=text)
    cell.font = Font(name="Calibri", size=14 if big else 11,
                     bold=True, color=_NAVY)
    cell.alignment = _AL_LEFT
    _row_h(ws, row, 20 if big else 16)


def _merge_section(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1, value=text)
    cell.font = Font(name="Calibri", size=10, bold=True, color=_WHITE)
    cell.fill  = _FILL_NAVY
    cell.alignment = _AL_LEFT
    _row_h(ws, row, 18)


def _blank_row(ws, row):
    _row_h(ws, row, 6)


# =============================================================================
# SHEET BUILDERS
# =============================================================================

def _build_vue_ensemble(wb: Workbook, data: dict):
    ws = wb.create_sheet("VUE D ENSEMBLE")
    ws.sheet_view.showGridLines = False

    name_a  = data.get("name_a", "Indice A")
    name_b  = data.get("name_b", "Indice B")
    sig_a   = data.get("signal_a", "Neutre")
    sig_b   = data.get("signal_b", "Neutre")
    sc_a    = data.get("score_a", 50)
    sc_b    = data.get("score_b", 50)
    date_s  = data.get("date", str(date.today()))

    # Largeurs colonnes
    for col, w in [(1, 2), (2, 30), (3, 22), (4, 22), (5, 18), (6, 2)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # Titre principal
    _merge_title(ws, 1, 2, 5, f"{name_a}  vs  {name_b}  \u2014  Comparaison FinSight IA",
                 big=True)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)
    c = ws.cell(row=2, column=2, value=f"Analyse au {date_s}")
    c.font = _FONT_GREY
    c.alignment = _AL_LEFT
    _blank_row(ws, 3)

    # ── Signal global ─────────────────────────────────────────────────────────
    _merge_section(ws, 4, 2, 5, "Signal Global")
    _row_h(ws, 5, 16)
    # Noms
    _set(ws, 5, 3, name_a[:25], font=Font(name="Calibri", size=9, bold=True, color=_COL_A),
         fill=_FILL_AL, align=_AL_CENTER, border=_BORDER_ALL)
    _set(ws, 5, 4, name_b[:25], font=Font(name="Calibri", size=9, bold=True, color=_COL_B),
         fill=_FILL_BL, align=_AL_CENTER, border=_BORDER_ALL)
    _set(ws, 5, 2, "Indice", font=Font(name="Calibri", size=9, bold=True, color=_GREY_T),
         fill=_FILL_GREY, align=_AL_LEFT, border=_BORDER_ALL)
    _set(ws, 5, 5, "Commentaire", font=Font(name="Calibri", size=9, bold=True, color=_GREY_T),
         fill=_FILL_GREY, align=_AL_LEFT, border=_BORDER_ALL)

    signal_rows = [
        ("Signal",       sig_a,           sig_b,
         "Score et recommandation FinSight composite"),
        ("Score /100",   str(sc_a),       str(sc_b),
         "Score base 0-100, >60=Surpondérer, <40=Sous-pondérer"),
        ("Devise",       data.get("currency_a", "EUR"), data.get("currency_b", "USD"),
         "Devise de reference de l'indice"),
    ]
    for i, (lbl, va, vb, cmt) in enumerate(signal_rows):
        r = 6 + i
        _label(ws, r, 2, lbl)
        _set(ws, r, 3, va, font=_sig_font(va) if "Signal" in lbl or "Score" in lbl else _FONT_BODY,
             align=_AL_CENTER, border=_BORDER_ALL)
        _set(ws, r, 4, vb, font=_sig_font(vb) if "Signal" in lbl or "Score" in lbl else _FONT_BODY,
             align=_AL_CENTER, border=_BORDER_ALL)
        _set(ws, r, 5, cmt, font=_FONT_GREY, align=_AL_LEFT, border=_BORDER_ALL)

    _blank_row(ws, 9)

    # ── Performance ───────────────────────────────────────────────────────────
    _merge_section(ws, 10, 2, 5, "Performance")
    _hdr(ws, 11, 2, "Horizon")
    _hdr_a(ws, 11, 3, name_a[:22])
    _hdr_b(ws, 11, 4, name_b[:22])
    _hdr(ws, 11, 5, "Écart (pp)")

    perf_rows = [
        ("YTD",   data.get("perf_ytd_a"), data.get("perf_ytd_b")),
        ("1 an",  data.get("perf_1y_a"),  data.get("perf_1y_b")),
        ("3 ans", data.get("perf_3y_a"),  data.get("perf_3y_b")),
        ("5 ans", data.get("perf_5y_a"),  data.get("perf_5y_b")),
    ]
    for i, (horizon, va, vb) in enumerate(perf_rows):
        r = 12 + i
        alt = (i % 2 == 1)
        _label(ws, r, 2, horizon)
        _val(ws, r, 3, _pct_str(va, signed=True), alt)
        _val(ws, r, 4, _pct_str(vb, signed=True), alt)
        # Écart
        if va is not None and vb is not None:
            try:
                fa = float(va) * 100 if abs(float(va)) < 2 else float(va)
                fb = float(vb) * 100 if abs(float(vb)) < 2 else float(vb)
                ecart = f"{fa-fb:+.2f}".replace(".", ",") + " pp"
            except Exception:
                ecart = "\u2014"
        else:
            ecart = "\u2014"
        _val(ws, r, 5, ecart, alt)

    _blank_row(ws, 16)

    # ── Risque ────────────────────────────────────────────────────────────────
    _merge_section(ws, 17, 2, 5, "Risque")
    _hdr(ws, 18, 2, "Indicateur")
    _hdr_a(ws, 18, 3, name_a[:22])
    _hdr_b(ws, 18, 4, name_b[:22])
    _hdr(ws, 18, 5, "Note")

    risque_rows = [
        ("Volatilite annualisee",
         _pct_str(data.get("vol_1y_a")), _pct_str(data.get("vol_1y_b")),
         "Écart-type rendements quotidiens annualises"),
        ("Sharpe ratio (1 an)",
         _num_str(data.get("sharpe_1y_a"), 2), _num_str(data.get("sharpe_1y_b"), 2),
         "(Rendement - rf) / volatilite"),
        ("Max Drawdown",
         _pct_str(data.get("max_dd_a"), signed=True),
         _pct_str(data.get("max_dd_b"), signed=True),
         "Perte max du plus haut au plus bas"),
    ]
    for i, (lbl, va, vb, note) in enumerate(risque_rows):
        r = 19 + i
        _label(ws, r, 2, lbl)
        _val(ws, r, 3, va, i % 2 == 1)
        _val(ws, r, 4, vb, i % 2 == 1)
        _set(ws, r, 5, note, font=_FONT_GREY, align=_AL_LEFT, border=_BORDER_ALL)

    _blank_row(ws, 22)

    # ── Valorisation ──────────────────────────────────────────────────────────
    _merge_section(ws, 23, 2, 5, "Valorisation")
    _hdr(ws, 24, 2, "Indicateur")
    _hdr_a(ws, 24, 3, name_a[:22])
    _hdr_b(ws, 24, 4, name_b[:22])
    _hdr(ws, 24, 5, "Favorise")

    def _fav_low(va, vb):
        if not va or not vb:
            return "\u2014"
        return name_a[:18] if float(va) < float(vb) else name_b[:18]

    def _fav_high(va, vb):
        if not va or not vb:
            return "\u2014"
        fa = float(va) * 100 if abs(float(va)) < 1 else float(va)
        fb = float(vb) * 100 if abs(float(vb)) < 1 else float(vb)
        return name_a[:18] if fa > fb else name_b[:18]

    pe_a = data.get("pe_fwd_a")
    pe_b = data.get("pe_fwd_b")
    pb_a = data.get("pb_a")
    pb_b = data.get("pb_b")
    dy_a = data.get("div_yield_a")
    dy_b = data.get("div_yield_b")
    erp_a = data.get("erp_a", "\u2014")
    erp_b = data.get("erp_b", "\u2014")

    val_rows = [
        ("P/E Forward",
         _num_str(pe_a, 1) + "x" if pe_a else "\u2014",
         _num_str(pe_b, 1) + "x" if pe_b else "\u2014",
         _fav_low(pe_a, pe_b)),
        ("P/B (Book Value)",
         _num_str(pb_a, 1) + "x" if pb_a else "\u2014",
         _num_str(pb_b, 1) + "x" if pb_b else "\u2014",
         _fav_low(pb_a, pb_b)),
        ("Rendement dividende",
         _pct_str(dy_a), _pct_str(dy_b),
         _fav_high(dy_a, dy_b)),
        ("ERP (prime de risque)",
         str(erp_a), str(erp_b), "\u2014"),
    ]
    for i, (lbl, va, vb, fav) in enumerate(val_rows):
        r = 25 + i
        _label(ws, r, 2, lbl)
        _val(ws, r, 3, va, i % 2 == 1)
        _val(ws, r, 4, vb, i % 2 == 1)
        _val(ws, r, 5, fav, i % 2 == 1)

    _blank_row(ws, 29)

    # Footer
    ws.merge_cells(start_row=30, start_column=2, end_row=30, end_column=5)
    c = ws.cell(row=30, column=2,
                value=f"Source : yfinance  |  FinSight IA  |  {date_s}  |  Usage confidentiel")
    c.font = _FONT_GREY
    c.alignment = _AL_LEFT

    # Freeze top row
    ws.freeze_panes = "B5"


def _build_secteurs(wb: Workbook, data: dict):
    ws = wb.create_sheet("SECTEURS")
    ws.sheet_view.showGridLines = False

    name_a   = data.get("name_a", "Indice A")
    name_b   = data.get("name_b", "Indice B")
    sc_cmp   = data.get("sector_comparison", [])

    for col, w in [(1, 2), (2, 32), (3, 18), (4, 18), (5, 18), (6, 2)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    _merge_title(ws, 1, 2, 5, "Composition Sectorielle Comparative", big=True)
    _blank_row(ws, 2)
    _merge_section(ws, 3, 2, 5, "Poids par secteur GICS (%)")

    _hdr(ws, 4, 2, "Secteur")
    _hdr_a(ws, 4, 3, name_a[:22])
    _hdr_b(ws, 4, 4, name_b[:22])
    _hdr(ws, 4, 5, "Écart (pp)")

    if sc_cmp:
        for i, item in enumerate(sc_cmp[:15]):
            r = 5 + i
            sector = str(item[0] if len(item) > 0 else "")[:32]
            wa = float(item[1]) if len(item) > 1 and item[1] is not None else None
            wb_ = float(item[2]) if len(item) > 2 and item[2] is not None else None
            wa_s = f"{wa:.2f}".replace(".", ",") + " %" if wa is not None else "\u2014"
            wb_s = f"{wb_:.2f}".replace(".", ",") + " %" if wb_ is not None else "\u2014"
            if wa is not None and wb_ is not None:
                diff = wa - wb_
                diff_s = f"{diff:+.2f}".replace(".", ",") + " pp"
            else:
                diff_s = "\u2014"
            alt = (i % 2 == 1)
            _label(ws, r, 2, sector)
            _val(ws, r, 3, wa_s, alt)
            _val(ws, r, 4, wb_s, alt)
            _val(ws, r, 5, diff_s, alt)
    else:
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=5)
        ws.cell(row=5, column=2,
                value="Données de composition sectorielle non disponibles.").font = _FONT_GREY

    ws.freeze_panes = "B5"


def _build_constituants(wb: Workbook, data: dict):
    ws = wb.create_sheet("TOP 5 CONSTITUANTS")
    ws.sheet_view.showGridLines = False

    name_a  = data.get("name_a", "Indice A")
    name_b  = data.get("name_b", "Indice B")
    top5_a  = data.get("top5_a", [])
    top5_b  = data.get("top5_b", [])

    for col, w in [(1, 2), (2, 36), (3, 14), (4, 14), (5, 26), (6, 2)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    _merge_title(ws, 1, 2, 5, "Top 5 Constituants", big=True)
    _blank_row(ws, 2)

    for which, name, top5, fill_hdr in [
        ("A", name_a, top5_a, _FILL_A),
        ("B", name_b, top5_b, _FILL_B),
    ]:
        start = 3 if which == "A" else 11

        _merge_section(ws, start, 2, 5, f"Top 5  \u2014  {name}")
        _set(ws, start+1, 2, "Société",
             font=_FONT_HDR, fill=fill_hdr, align=_AL_LEFT, border=_BORDER_ALL)
        _hdr(ws, start+1, 3, "Ticker")
        _hdr(ws, start+1, 4, "Poids (%)")
        _hdr(ws, start+1, 5, "Secteur")
        for i, item in enumerate(top5[:5]):
            r = start + 2 + i
            company = str(item[0] if len(item) > 0 else "\u2014")[:40]
            ticker  = str(item[1] if len(item) > 1 else "")
            weight  = item[2] if len(item) > 2 else None
            sector  = str(item[3] if len(item) > 3 else "")[:28]
            wt_s    = f"{float(weight):.2f}".replace(".", ",") + " %" if weight is not None else "\u2014"
            alt = (i % 2 == 1)
            _label(ws, r, 2, company)
            _val(ws, r, 3, ticker, alt)
            _val(ws, r, 4, wt_s, alt)
            _set(ws, r, 5, sector, font=_FONT_BODY, align=_AL_LEFT, border=_BORDER_ALL)

        _blank_row(ws, start + 7)

    ws.freeze_panes = "B3"


# =============================================================================
# CLASSE PRINCIPALE
# =============================================================================

class IndiceComparisonWriter:

    @staticmethod
    def generate_bytes(data: dict) -> bytes:
        if not _OXL:
            raise ImportError("openpyxl requis pour IndiceComparisonWriter")

        wb = Workbook()
        # Supprimer la feuille vide par defaut
        if wb.active:
            del wb[wb.active.title]

        _build_vue_ensemble(wb, data)
        _build_secteurs(wb, data)
        _build_constituants(wb, data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate(data: dict, output_path: str) -> str:
        xlsx_bytes = IndiceComparisonWriter.generate_bytes(data)
        Path(output_path).write_bytes(xlsx_bytes)
        return output_path
