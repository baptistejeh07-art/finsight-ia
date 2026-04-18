"""Writer Excel dédié — analyse secteur Énergie/Oil.

Utilise le template `outputs/templates/SECTOR_ENERGIE_TEMPLATE.xlsx` fourni par
Baptiste comme référence. Le template contient :
  - DASHBOARD, VALUE, GROWTH, QUALITY, MOMENTUM : scoring multi-factoriel auto
  - ÉNERGIE : screening par société (formules COUNTIF/AVERAGEIF référençant DONNÉES BRUTES)
  - DONNÉES BRUTES : 88 rows (sociétés du secteur) × 41 colonnes (métriques + scores)

Notre rôle : remplir DONNÉES BRUTES. Toutes les autres feuilles se calculent
automatiquement via les formules du template.

Schema DONNÉES BRUTES (header row 2) :
    A=Ticker B=Société C=Secteur D=Cours E=MktCap F=EV
    G=RevLTM H=EBITDA_LTM I=EV/EBITDA J=EV/Rev K=P/E L=EPS
    M=MgBrute N=MgEBITDA O=MgNette P=CroRev Q=ROE R=ROA
    S=CurrentR T=ND/EBITDA U=AltmanZ V=BeneishM W=Mom52W
    X=ScoreVal Y=ScoreGr Z=ScoreQual AA=ScoreMom AB=ScoreGlobal
    AC=NextEarn AD=Signal AE=NetInc AF=Equity AG=Debt AH=Cash
    AI=FCF AJ=Capex AK=FCFYield AL=D/E AM=CashConv AN=Capex/Rev AO=NetDebt
"""
from __future__ import annotations
import logging
import shutil
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

log = logging.getLogger(__name__)

_TEMPLATE = Path(__file__).parent / "templates" / "SECTOR_ENERGIE_TEMPLATE.xlsx"
_DATA_SHEET = "DONNÉES BRUTES"
_DATA_START_ROW = 3
_DATA_MAX_ROW = 90  # ne pas dépasser sinon les formules d'agrégation cassent

# Colonnes (lettre Excel) ordonnées comme dans le template, pour mapping direct.
COLUMNS: list[tuple[str, str]] = [
    ("A", "ticker"),
    ("B", "company_name"),
    ("C", "sector"),
    ("D", "price"),
    ("E", "market_cap"),
    ("F", "enterprise_value"),
    ("G", "revenue_ltm"),
    ("H", "ebitda_ltm"),
    ("I", "ev_ebitda"),
    ("J", "ev_revenue"),
    ("K", "pe_ratio"),
    ("L", "eps"),
    ("M", "gross_margin"),
    ("N", "ebitda_margin"),
    ("O", "net_margin"),
    ("P", "revenue_growth"),
    ("Q", "roe"),
    ("R", "roa"),
    ("S", "current_ratio"),
    ("T", "net_debt_ebitda"),
    ("U", "altman_z"),
    ("V", "beneish_m"),
    ("W", "momentum_52w"),
    # Scores (X-AB) : laissés vides ici, calculés par les formules du template
    ("AC", "next_earnings"),
    ("AD", "signal"),
    ("AE", "net_income"),
    ("AF", "equity"),
    ("AG", "debt"),
    ("AH", "cash"),
    ("AI", "fcf"),
    ("AJ", "capex"),
    ("AK", "fcf_yield"),
    ("AL", "debt_equity"),
    ("AM", "cash_conversion"),
    ("AN", "capex_revenue"),
    ("AO", "net_debt"),
]


def write_energy_sector_xlsx(
    tickers_data: list[dict],
    output_path: str | Path,
    template_path: Optional[Path] = None,
) -> Path:
    """Génère le XLSX secteur énergie à partir d'une liste de dicts par société.

    Args:
        tickers_data: liste de dict, 1 par société. Clés attendues : voir COLUMNS.
        output_path: où sauver le fichier Excel généré.
        template_path: override du template (défaut = SECTOR_ENERGIE_TEMPLATE.xlsx).

    Returns:
        Path du fichier généré.

    Le template est copié puis modifié en place. Les feuilles autres que
    DONNÉES BRUTES (DASHBOARD/VALUE/GROWTH/QUALITY/MOMENTUM/ÉNERGIE) se mettent
    à jour automatiquement via les formules au prochain ouvrant Excel.
    """
    template = template_path or _TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"Template introuvable : {template}")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Copie le template (préserve formules + styling)
    shutil.copy(str(template), str(output_path))

    wb = load_workbook(str(output_path))
    if _DATA_SHEET not in wb.sheetnames:
        # Tolérance encodage : trouve la sheet par préfixe
        candidates = [s for s in wb.sheetnames if "DONN" in s.upper()]
        if not candidates:
            raise ValueError(f"Sheet '{_DATA_SHEET}' introuvable dans template")
        ws = wb[candidates[0]]
    else:
        ws = wb[_DATA_SHEET]

    # Vide les rows existantes (au cas où template contient des exemples)
    for row in range(_DATA_START_ROW, _DATA_MAX_ROW + 1):
        for col_letter, _key in COLUMNS:
            cell = ws[f"{col_letter}{row}"]
            # Ne pas écraser les formules (X-AB scores)
            if not (isinstance(cell.value, str) and str(cell.value).startswith("=")):
                cell.value = None

    # Remplit avec les nouvelles données
    n = min(len(tickers_data), _DATA_MAX_ROW - _DATA_START_ROW + 1)
    for i in range(n):
        row = _DATA_START_ROW + i
        d = tickers_data[i]
        for col_letter, key in COLUMNS:
            val = d.get(key)
            if val is not None:
                ws[f"{col_letter}{row}"] = val

    # Force le secteur "Énergie" partout (cohérence avec les COUNTIF du template)
    for i in range(n):
        row = _DATA_START_ROW + i
        if not ws[f"C{row}"].value:
            ws[f"C{row}"] = "Énergie"

    wb.save(str(output_path))
    log.info(
        f"[sector_energy_xlsx] {output_path.name} rempli avec {n} société(s) "
        f"(template DONNÉES BRUTES rows {_DATA_START_ROW}–{_DATA_START_ROW + n - 1})"
    )
    return output_path


def build_ticker_dict_from_yfinance(ticker: str, info: dict, ratios: Optional[dict] = None) -> dict:
    """Helper : convertit yfinance .info + ratios en dict format COLUMNS."""
    ratios = ratios or {}
    return {
        "ticker": ticker,
        "company_name": info.get("longName") or info.get("shortName") or ticker,
        "sector": "Énergie",
        "price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "market_cap": (info.get("marketCap") or 0) / 1e9 if info.get("marketCap") else None,
        "enterprise_value": (info.get("enterpriseValue") or 0) / 1e9 if info.get("enterpriseValue") else None,
        "revenue_ltm": (info.get("totalRevenue") or 0) / 1e9 if info.get("totalRevenue") else None,
        "ebitda_ltm": (info.get("ebitda") or 0) / 1e9 if info.get("ebitda") else None,
        "ev_ebitda": info.get("enterpriseToEbitda"),
        "ev_revenue": info.get("enterpriseToRevenue"),
        "pe_ratio": info.get("trailingPE"),
        "eps": info.get("trailingEps"),
        "gross_margin": info.get("grossMargins"),
        "ebitda_margin": info.get("ebitdaMargins"),
        "net_margin": info.get("profitMargins"),
        "revenue_growth": info.get("revenueGrowth"),
        "roe": info.get("returnOnEquity"),
        "roa": info.get("returnOnAssets"),
        "current_ratio": info.get("currentRatio"),
        "net_debt_ebitda": ratios.get("net_debt_ebitda"),
        "altman_z": ratios.get("altman_z"),
        "beneish_m": ratios.get("beneish_m"),
        "momentum_52w": info.get("52WeekChange"),
        "net_income": (info.get("netIncomeToCommon") or 0) / 1e9 if info.get("netIncomeToCommon") else None,
        "equity": (info.get("totalStockholderEquity") or 0) / 1e9 if info.get("totalStockholderEquity") else None,
        "debt": (info.get("totalDebt") or 0) / 1e9 if info.get("totalDebt") else None,
        "cash": (info.get("totalCash") or 0) / 1e9 if info.get("totalCash") else None,
        "fcf": (info.get("freeCashflow") or 0) / 1e9 if info.get("freeCashflow") else None,
    }
