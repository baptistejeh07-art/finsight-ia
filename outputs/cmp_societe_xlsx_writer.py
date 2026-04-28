"""
comparison_writer.py — FinSight IA
Injecte deux états pipeline dans TEMPLATE_COMPARISON.xlsx (feuille DATA_RAW).

Usage:
    path = CmpSocieteXlsxWriter().write(state_a, state_b)

state_a / state_b : dict renvoyé par le pipeline LangGraph (st.session_state.results)
  Champs utilisés : raw_data (FinancialSnapshot), ratios (RatiosResult),
                    synthesis (SynthesisResult), qa_python (QAResult)
"""
from __future__ import annotations

import io
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import yfinance as yf
from core.yfinance_cache import get_ticker

log = logging.getLogger(__name__)

# Chemin template source (OneDrive)
_TEMPLATE_SRC = Path("C:/Users/bapti/OneDrive/Perso/Excel Finsight/TEMPLATE_COMPARISON.xlsx")
# Chemin sortie
_OUT_DIR = Path(__file__).parent / "generated" / "cli_tests"

# Médiane sectorielle PE / EV-EBITDA — approx. pour benchmarks rapides
_SECTOR_PE: dict[str, float] = {
    "Technology": 30.0, "Information Technology": 30.0,
    "Health Care": 22.0, "Healthcare": 22.0,
    "Financials": 13.0, "Financial Services": 13.0,
    "Consumer Discretionary": 20.0, "Consumer Cyclical": 20.0,
    "Consumer Staples": 18.0, "Consumer Défensive": 18.0,
    "Communication Services": 18.0,
    "Industrials": 21.0,
    "Energy": 12.0,
    "Materials": 15.0, "Basic Materials": 15.0,
    "Real Estate": 30.0,
    "Utilities": 20.0,
}
_SECTOR_EVEBITDA: dict[str, float] = {
    "Technology": 18.0, "Information Technology": 18.0,
    "Health Care": 14.0, "Healthcare": 14.0,
    "Financials": 9.0,  "Financial Services": 9.0,
    "Consumer Discretionary": 12.0, "Consumer Cyclical": 12.0,
    "Consumer Staples": 12.0, "Consumer Défensive": 12.0,
    "Communication Services": 11.0,
    "Industrials": 13.0,
    "Energy": 6.0,
    "Materials": 10.0, "Basic Materials": 10.0,
    "Real Estate": 17.0,
    "Utilities": 11.0,
}


# ---------------------------------------------------------------------------
# Helpers d'extraction
# ---------------------------------------------------------------------------

def _safe(obj, *attrs, default=None):
    """Accès sécurisé sur chaîne d'attributs."""
    try:
        for a in attrs:
            if obj is None:
                return default
            obj = getattr(obj, a, None)
        return obj if obj is not None else default
    except Exception:
        return default


def _latest_yr(ratios):
    """Retourne le YearRatios le plus récent, ou None."""
    if ratios is None:
        return None
    try:
        latest = ratios.latest_year
        return ratios.years.get(latest)
    except Exception:
        return None


def _prev_yrs(ratios, n=2):
    """Retourne les n années précédentes (YearRatios), de la plus récente à la plus ancienne."""
    if ratios is None:
        return [None] * n
    try:
        labels = sorted(ratios.years.keys(), key=lambda y: int(y.split("_")[0]))
        # Toutes sauf la dernière
        prev_labels = labels[:-1][-n:]
        result = [ratios.years.get(l) for l in reversed(prev_labels)]
        while len(result) < n:
            result.append(None)
        return result
    except Exception:
        return [None] * n


def _ebitda_trend(yr_ltm, yr_prev1, yr_prev2) -> str:
    """Calcule la tendance EBITDA sur 3 ans."""
    try:
        m0 = getattr(yr_ltm,   "ebitda_margin", None)
        m1 = getattr(yr_prev1, "ebitda_margin", None) if yr_prev1 else None
        m2 = getattr(yr_prev2, "ebitda_margin", None) if yr_prev2 else None
        if m0 is None:
            return "N/A"
        if m1 is None:
            return "Stable"
        delta = m0 - m1
        if m2 is not None:
            delta2 = m1 - m2
            if delta > 0.01 and delta2 > 0:
                return "Croissante"
            if delta < -0.01 and delta2 < 0:
                return "Décroissante"
        if delta > 0.01:
            return "Croissante"
        if delta < -0.01:
            return "Décroissante"
        return "Stable"
    except Exception:
        return "N/A"


def _rev_cagr_3y(ratios) -> Optional[float]:
    """CAGR chiffre d'affaires 3 ans depuis les ratios. Fallback 2Y ou 1Y."""
    try:
        labels = sorted(ratios.years.keys(), key=lambda y: int(y.split("_")[0]))
        if len(labels) < 2:
            return None
        y0 = ratios.years[labels[-1]]
        r0 = _rev_from_yr(y0)
        if not r0:
            return None
        # Essayer 3Y, puis 2Y, puis 1Y
        for n_years in (3, 2, 1):
            idx = -(n_years + 1)
            if abs(idx) <= len(labels):
                yn = ratios.years[labels[idx]]
                rn = _rev_from_yr(yn)
                if rn and rn > 0:
                    return round((r0 / rn) ** (1 / n_years) - 1, 4)
        return None
    except Exception:
        return None


def _rev_from_yr(yr) -> Optional[float]:
    """Revenue estimé depuis YearRatios (GP / gross_margin)."""
    try:
        if yr is None:
            return None
        gp = getattr(yr, "gross_profit", None)
        gm = getattr(yr, "gross_margin", None)
        if gp and gm and gm > 0:
            return gp / gm
        return None
    except Exception:
        return None


def _eps_growth(ratios) -> Optional[float]:
    """EPS growth YoY depuis net_income/market_cap proxy."""
    try:
        labels = sorted(ratios.years.keys(), key=lambda y: int(y.split("_")[0]))
        if len(labels) < 2:
            return None
        y0 = ratios.years[labels[-1]]
        y1 = ratios.years[labels[-2]]
        ni0 = getattr(y0, "net_income", None)
        ni1 = getattr(y1, "net_income", None)
        if ni0 and ni1 and ni1 > 0:
            return round((ni0 - ni1) / abs(ni1), 4)
        return None
    except Exception:
        return None


def _finsight_score_full(snapshot, ratios) -> dict:
    """Score composite 0-100 + breakdown 4 axes pour la cmp société.

    Bug slide 18 audit 27/04 : avant, slide 18 affichait des proxies indépendants
    (value=25-PE*0.4, growth=12+CAGR, etc.) qui ne sommaient PAS au global score
    (NVDA proxies 0+25+25+14=64 alors que global=41). Maintenant on retourne le
    breakdown réel de `compute_score` v1.3 pour que slide 18 affiche les vrais
    sous-scores cohérents avec le total.

    Returns:
        dict {global, displayed, value, growth, quality, momentum, governance,
              weights} ou {} si compute échoue.
    """
    try:
        from core.finsight_score import compute_score
        yr = _latest_yr(ratios)
        if yr is not None:
            ratio_dict = {
                k: _safe(yr, k) for k in (
                    "pe_ratio", "ev_ebitda", "ev_revenue", "roe", "roic",
                    "ebitda_margin", "net_margin", "gross_margin",
                    "revenue_growth", "fcf_yield", "div_yield", "payout_ratio",
                    "altman_z", "piotroski_f", "beneish_m",
                    "net_debt_ebitda", "momentum_52w",
                    "shares_change_pct", "earnings_growth",
                )
            }
            if ratio_dict.get("momentum_52w") in (None, 0, 0.0):
                hist = getattr(snapshot, "stock_history", []) or []
                if len(hist) >= 13:
                    try:
                        p_now = float(hist[-1].price)
                        p_1y = float(hist[0].price)
                        if p_1y > 0:
                            ratio_dict["momentum_52w"] = (p_now / p_1y) - 1
                    except Exception:
                        pass
            mk = getattr(snapshot, "market", None)
            market_dict = {
                "share_price":    getattr(mk, "share_price", None) if mk else None,
                "beta_levered":   getattr(mk, "beta_levered", None) if mk else None,
                "dividend_yield": getattr(mk, "dividend_yield", None) if mk else None,
            }
            ci = getattr(snapshot, "company_info", None)
            sec = getattr(ci, "sector", None) if ci else None
            ind = getattr(ci, "industry", None) if ci else None
            return compute_score(ratio_dict, market_dict, sector=sec, industry=ind) or {}
    except Exception as e:
        log.debug(f"[cmp_societe_xlsx:_finsight_score_full] {e}")
    return {}


def _finsight_score(snapshot, ratios) -> Optional[int]:
    """Score composite 0-100 pour la comparaison société.

    Branché sur `core.finsight_score.compute_score` (v1.3) : la note renvoyée
    est le `displayed` recalibré (mean≈50, std élargie) pour que la cmp UI
    MC.PA vs OR.PA vs SU.PA ne clusterise plus à 48-50. Fallback silencieux
    sur l'ancienne heuristique value+growth+quality+momentum si la pipeline
    partielle ne permet pas le calcul v1.3 (ratios minimum manquants).
    """
    # ── Chemin principal : compute_score v1.3 ────────────────────────────
    try:
        from core.finsight_score import compute_score
        yr = _latest_yr(ratios)
        if yr is not None:
            ratio_dict = {
                k: _safe(yr, k) for k in (
                    "pe_ratio", "ev_ebitda", "ev_revenue", "roe", "roic",
                    "ebitda_margin", "net_margin", "gross_margin",
                    "revenue_growth", "fcf_yield", "div_yield", "payout_ratio",
                    "altman_z", "piotroski_f", "beneish_m",
                    "net_debt_ebitda", "momentum_52w",
                    "shares_change_pct", "earnings_growth",
                )
            }
            # Complément momentum_52w depuis stock_history si absent du dict.
            if ratio_dict.get("momentum_52w") in (None, 0, 0.0):
                hist = getattr(snapshot, "stock_history", []) or []
                if len(hist) >= 13:
                    try:
                        p_now = float(hist[-1].price)
                        p_1y  = float(hist[0].price)
                        if p_1y > 0:
                            ratio_dict["momentum_52w"] = (p_now / p_1y) - 1
                    except Exception:
                        pass
            mk = getattr(snapshot, "market", None)
            market_dict = {
                "share_price":    getattr(mk, "share_price", None)    if mk else None,
                "beta_levered":   getattr(mk, "beta_levered", None)   if mk else None,
                "dividend_yield": getattr(mk, "dividend_yield", None) if mk else None,
            }
            ci = getattr(snapshot, "company_info", None)
            sec = getattr(ci, "sector", None)   if ci else None
            ind = getattr(ci, "industry", None) if ci else None
            res = compute_score(ratio_dict, market_dict, sector=sec, industry=ind)
            # On expose `displayed` (score recalibré) qui propage directement
            # au XLSX/PDF/PPTX cmp et à l'UI Streamlit.
            return int(res.get("displayed") or res.get("global") or 0)
    except Exception:
        pass

    # ── Fallback legacy (pipeline partiel, ratios minimum manquants) ────
    try:
        yr = _latest_yr(ratios)
        pe  = _safe(yr, "pe_ratio") or 20.0
        rev = _rev_cagr_3y(ratios) or 0.0
        nm  = _safe(yr, "net_margin") or 0.0
        hist = getattr(snapshot, "stock_history", []) or []
        if len(hist) >= 13:
            try:
                p_now  = float(hist[-1].price)
                p_1y   = float(hist[0].price)
                mom52  = (p_now / p_1y - 1) * 100 if p_1y > 0 else 0.0
            except Exception:
                mom52 = 0.0
        else:
            mom52 = 0.0

        s_val = max(0.0, min(25.0, 25.0 - pe * 0.5))
        s_gro = max(0.0, min(25.0, 12.0 + rev * 100))
        s_qua = max(0.0, min(25.0, 10.0 + nm * 0.8))
        s_mom = max(0.0, min(25.0, 12.0 + mom52 * 0.15))
        return round(s_val + s_gro + s_qua + s_mom)
    except Exception:
        return None


def _momentum_score(perf_3m: Optional[float]) -> Optional[float]:
    """Score momentum /10 depuis performance 3 mois."""
    if perf_3m is None:
        return None
    # perf_3m est en décimal (ex. 0.112 = +11.2%)
    # Mapping : -20% → 0, 0% → 5, +20% → 10
    score = 5.0 + (perf_3m * 100) * 0.25
    return round(max(0.0, min(10.0, score)), 1)


# ---------------------------------------------------------------------------
# Fetch supplémentaire yfinance (données marché non stockées dans l'état)
# ---------------------------------------------------------------------------

def _fetch_supplements(ticker: str) -> dict:
    """
    Récupère via yfinance les métriques non présentes dans l'état pipeline :
    perf_1m/3m/1y, week52_high/low, avg_volume_30d, next_earnings_date,
    piotroski_score + composantes pio_*.
    """
    out: dict = {}
    try:
        stock = get_ticker(ticker)
        info  = stock.fast_info
        full  = stock.info or {}

        # Performance historique
        hist = stock.history(period="1y", interval="1mo", auto_adjust=True)
        if not hist.empty and len(hist) >= 2:
            try:
                price_now = float(hist["Close"].iloc[-1])
                # 1 mois
                if len(hist) >= 2:
                    p1m = float(hist["Close"].iloc[-2])
                    out["perf_1m"] = round((price_now / p1m - 1), 4) if p1m > 0 else None
                # 3 mois
                if len(hist) >= 4:
                    p3m = float(hist["Close"].iloc[-4])
                    out["perf_3m"] = round((price_now / p3m - 1), 4) if p3m > 0 else None
                # 1 an
                if len(hist) >= 13:
                    p1y = float(hist["Close"].iloc[-13])
                    out["perf_1y"] = round((price_now / p1y - 1), 4) if p1y > 0 else None
                elif len(hist) >= 2:
                    p1y = float(hist["Close"].iloc[0])
                    out["perf_1y"] = round((price_now / p1y - 1), 4) if p1y > 0 else None
            except Exception as e:
                log.debug(f"[comparison] perf calc error {ticker}: {e}")

        # Données marché
        _w52h = _g(info, "year_high") or _g(full, "fiftyTwoWeekHigh") or _g(full, "yearHigh")
        _w52l = _g(info, "year_low")  or _g(full, "fiftyTwoWeekLow")  or _g(full, "yearLow")
        out["week52_high"]   = round(float(_w52h), 2) if _w52h else None
        out["week52_low"]    = round(float(_w52l), 2) if _w52l else None
        out["avg_volume_30d"] = round((_g(full, "averageVolume30Day") or _g(full, "averageVolume") or 0) / 1e6, 1) or None
        # Share price (fallback chain)
        _sp = _g(info, "last_price") or _g(full, "currentPrice") or _g(full, "regularMarketPrice")
        if _sp:
            out["share_price"] = round(float(_sp), 2)
        # Dividend yield depuis yfinance.info — cascade 3 sources
        try:
            dy = _g(full, "trailingAnnualDividendYield") or _g(full, "dividendYield")
            if dy and dy > 0:
                # dividendYield est en % (ex: 3.25 = 3.25%), convertir en decimal si > 0.1
                if dy > 0.1:
                    dy = dy / 100.0
                out["dividend_yield"] = round(float(dy), 4)
            # Fallback : dividendRate / currentPrice
            if not out.get("dividend_yield"):
                dr = _g(full, "dividendRate")
                cp = _g(full, "currentPrice") or _g(info, "last_price")
                if dr and cp and float(cp) > 0:
                    out["dividend_yield"] = round(float(dr) / float(cp), 4)
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

        # Target price consensus (fallback DCF)
        try:
            tp = _g(full, "targetMeanPrice")
            if tp and float(tp) > 0:
                out["target_mean_price"] = round(float(tp), 2)
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

        # VaR 95% mensuelle (si historique disponible)
        try:
            hist_d = stock.history(period="1y", interval="1d", auto_adjust=True)
            if not hist_d.empty and len(hist_d) >= 21:
                import numpy as np
                returns = hist_d["Close"].pct_change().dropna()
                # VaR mensuelle : agréger en fenêtres 21j
                monthly_r = [float(sum(returns.iloc[i:i+21])) for i in range(0, len(returns) - 20, 21)]
                if monthly_r:
                    out["var_95_1m"] = round(float(np.percentile(monthly_r, 5)), 4)
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

        # Prochaine date de résultats
        try:
            cal = stock.calendar
            if isinstance(cal, dict):
                ed = cal.get("Earnings Date") or cal.get("earnings_date")
                if ed:
                    if hasattr(ed, "__iter__") and not isinstance(ed, str):
                        ed = list(ed)[0]
                    out["next_earnings_date"] = str(ed)[:10]
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

        # Piotroski F-Score + composantes
        try:
            from cli_analyze import _compute_piotroski
            pio = _compute_piotroski(stock)
            if pio is not None:
                out["piotroski_score"] = pio
            # Composantes détaillées (re-calculer)
            out.update(_compute_pio_components(stock))
        except Exception as e:
            log.debug(f"[comparison] piotroski error {ticker}: {e}")

        # Interest coverage depuis income statement (fallback si quant_node n'a pas calculé)
        try:
            import pandas as pd
            inc = getattr(stock, "income_stmt", None)
            if inc is None:
                inc = getattr(stock, "financials", None)
            if inc is not None and not getattr(inc, "empty", True) and len(inc.columns) >= 1:
                col = inc.columns[0]
                def _vi_ic(keys):
                    for k in keys:
                        if k in inc.index:
                            try:
                                v = float(inc.loc[k, col])
                                if pd.notna(v):
                                    return v
                            except Exception as _e:
                                log.debug(f"[cmp_societe_xlsx_writer:_vi_ic] exception skipped: {_e}")
                    return None
                ebit = _vi_ic(["EBIT", "Operating Income", "Ebit"])
                ie   = _vi_ic(["Interest Expense", "Interest Expense Non Operating",
                                "Net Interest Income"])
                if ebit is not None and ie is not None and abs(ie) > 1000:
                    out["interest_coverage"] = round(abs(float(ebit)) / abs(float(ie)), 1)
                elif ebit is not None and ie is None:
                    # Fallback : derive depuis EBIT - Pretax Income (ex. Apple dont
                    # l'interest expense net est absorbe par les revenus de trésorerie)
                    pretax = _vi_ic(["Pretax Income", "Income Before Tax"])
                    if pretax is not None:
                        net_non_op = ebit - pretax  # positif = charge nette
                        if net_non_op > 1000:
                            out["interest_coverage"] = round(ebit / net_non_op, 1)
                        elif net_non_op <= 0:
                            # Revenus d'intérêts > charges : couverture excellente
                            out["interest_coverage"] = 999.0
        except Exception as e:
            log.debug(f"[comparison] interest_coverage error {ticker}: {e}")

        # Volatilite annualisee 52 semaines + RSI 14j
        try:
            import numpy as np
            hist_v = stock.history(period="1y", interval="1d", auto_adjust=True)
            if not hist_v.empty and len(hist_v) >= 30:
                rets = hist_v["Close"].pct_change().dropna()
                vol = float(np.std(rets)) * np.sqrt(252)
                out["volatility_52w"] = round(vol, 4)
                # RSI 14j
                if len(rets) >= 14:
                    delta = rets.iloc[-14:]
                    gain = delta.clip(lower=0).mean()
                    loss = (-delta.clip(upper=0)).mean()
                    if loss > 0:
                        rs = gain / loss
                        out["rsi"] = round(100 - (100 / (1 + rs)), 1)
                    elif gain > 0:
                        out["rsi"] = 100.0
                    else:
                        out["rsi"] = 50.0
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

        # EV/Sales depuis yfinance.info
        try:
            evs = _g(full, "enterpriseToRevenue")
            if evs and float(evs) > 0:
                out["ev_sales"] = round(float(evs), 2)
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

        # Secteur (fallback si non stocke dans l'État pipeline)
        try:
            out["sector"] = full.get("sector", "") or ""
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

        # Forward estimates (consensus analystes)
        try:
            out["forward_eps"]          = full.get("forwardEps")
            out["forward_pe"]           = full.get("forwardPE")
            out["trailing_eps"]         = full.get("trailingEps")
            out["revenue_growth"]       = full.get("revenueGrowth")   # YoY fwd
            out["earnings_growth_est"]  = full.get("earningsGrowth")
        except Exception as _e:
            log.debug(f"[cmp_societe_xlsx_writer:_fetch_supplements] exception skipped: {_e}")

    except Exception as e:
        log.warning(f"[comparison] _fetch_supplements({ticker}) erreur : {e}")
    return out


def _g(obj, key, default=None):
    """Getter sécurisé pour dict ou objet."""
    try:
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)
    except Exception:
        return default


def _compute_pio_components(stock) -> dict:
    """
    Calcule les 9 composantes binaires du Piotroski F-Score.
    Retourne un dict {pio_roa_positive: 0/1, ...}.
    """
    out = {}
    try:
        import pandas as pd
        inc = getattr(stock, "income_stmt", None)
        if inc is None: inc = getattr(stock, "financials", None)
        bs  = getattr(stock, "balance_sheet", None)
        cf  = getattr(stock, "cash_flow", None)
        if cf is None: cf = getattr(stock, "cashflow", None)

        if (inc is None or getattr(inc, "empty", True) or
                bs  is None or getattr(bs,  "empty", True) or
                cf  is None or getattr(cf,  "empty", True)):
            return out
        if len(inc.columns) < 2 or len(bs.columns) < 2:
            return out

        def _v(df, keys, col):
            for k in keys:
                if k in df.index:
                    try:
                        v = float(df.loc[k, col])
                        if pd.notna(v):
                            return v
                    except (TypeError, ValueError):
                        pass
            return None

        ic0, ic1 = inc.columns[0], inc.columns[1]
        bc0 = bs.columns[0]
        bc1 = bs.columns[1] if len(bs.columns) > 1 else bc0

        ni0  = _v(inc, ["Net Income", "Net Income Common Stockholders"], ic0)
        ni1  = _v(inc, ["Net Income", "Net Income Common Stockholders"], ic1)
        gp0  = _v(inc, ["Gross Profit"], ic0)
        gp1  = _v(inc, ["Gross Profit"], ic1)
        rev0 = _v(inc, ["Total Revenue", "Revenue"], ic0)
        rev1 = _v(inc, ["Total Revenue", "Revenue"], ic1)
        ta0  = _v(bs, ["Total Assets"], bc0)
        ta1  = _v(bs, ["Total Assets"], bc1)
        tca0 = _v(bs, ["Current Assets", "Total Current Assets"], bc0)
        tcl0 = _v(bs, ["Current Liabilities", "Total Current Liabilities"], bc0)
        tca1 = _v(bs, ["Current Assets", "Total Current Assets"], bc1)
        tcl1 = _v(bs, ["Current Liabilities", "Total Current Liabilities"], bc1)
        ltd0 = _v(bs, ["Long Term Debt", "Long-Term Debt"], bc0)
        ltd1 = _v(bs, ["Long Term Debt", "Long-Term Debt"], bc1)
        shr0 = _v(bs, ["Common Stock", "Share Issued", "Ordinary Shares Number"], bc0)
        shr1 = _v(bs, ["Common Stock", "Share Issued", "Ordinary Shares Number"], bc1)
        ocf0 = _v(cf, ["Operating Cash Flow", "Total Cash From Operating Activities"], ic0)

        if ni0  is not None and ta0 and ta0 != 0:
            roa0 = ni0 / ta0
            out["pio_roa_positive"] = 1 if roa0 > 0 else 0
        if ocf0 is not None:
            out["pio_cfo_positive"] = 1 if ocf0 > 0 else 0
        if ni0 is not None and ni1 is not None and ta0 and ta1 and ta0 != 0 and ta1 != 0:
            out["pio_delta_roa"] = 1 if (ni0 / ta0) > (ni1 / ta1) else 0
        if ocf0 is not None and ni0 is not None and ta0 and ta0 != 0 and ni0 != 0:
            out["pio_accruals"] = 1 if (ocf0 / ta0) > (ni0 / ta0) else 0
        if ltd0 is not None and ltd1 is not None and ta0 and ta1 and ta0 != 0 and ta1 != 0:
            out["pio_delta_leverage"] = 1 if (ltd0 / ta0) <= (ltd1 / ta1) else 0
        if tca0 and tcl0 and tca1 and tcl1 and tcl0 != 0 and tcl1 != 0:
            cr0 = tca0 / tcl0
            cr1 = tca1 / tcl1
            out["pio_delta_liquidity"] = 1 if cr0 >= cr1 else 0
        if shr0 is not None and shr1 is not None:
            out["pio_no_dilution"] = 1 if shr0 <= shr1 else 0
        if gp0 and rev0 and gp1 and rev1 and rev0 != 0 and rev1 != 0:
            out["pio_delta_gross_margin"] = 1 if (gp0 / rev0) >= (gp1 / rev1) else 0
        if rev0 and ta0 and rev1 and ta1 and ta0 != 0 and ta1 != 0:
            out["pio_delta_asset_turnover"] = 1 if (rev0 / ta0) >= (rev1 / ta1) else 0

    except Exception as e:
        log.debug(f"[comparison] pio_components error: {e}")
    return out


# ---------------------------------------------------------------------------
# Extraction principale : état pipeline → dict 76 métriques
# ---------------------------------------------------------------------------

def extract_metrics(state: dict, supp: dict) -> dict:
    """
    Extrait les 76 métriques depuis un état pipeline + données supplémentaires yfinance.
    Retourne un dict {metric_name: value}.
    """
    # Compatibilité : graph.py stocke sous "raw_data", app.py sous "snapshot"
    snapshot   = state.get("raw_data") or state.get("snapshot")
    ratios     = state.get("ratios")
    synthesis  = state.get("synthesis")
    qa_python  = state.get("qa_python")
    sentiment  = state.get("sentiment")

    ci  = _safe(snapshot, "company_info")
    mkt = _safe(snapshot, "market")
    yr  = _latest_yr(ratios)
    yr1, yr2 = _prev_yrs(ratios, 2)

    # Monte Carlo depuis ratios.meta
    mc_meta = {}
    try:
        if ratios and ratios.meta:
            mc_meta = ratios.meta
    except Exception as _e:
        log.debug(f"[cmp_societe_xlsx_writer:extract_metrics] exception skipped: {_e}")

    # Prix courant (pipeline + fallback supplements)
    price = _safe(mkt, "share_price") or supp.get("share_price")

    # DCF targets — fallback targetMeanPrice depuis yfinance
    t_base = _safe(synthesis, "target_base")
    t_bear = _safe(synthesis, "target_bear")
    t_bull = _safe(synthesis, "target_bull")
    if t_base is None:
        t_base = supp.get("target_mean_price")
    if t_bear is None and t_base:
        try: t_bear = round(float(t_base) * 0.85, 2)
        except Exception: pass
    if t_bull is None and t_base:
        try: t_bull = round(float(t_base) * 1.15, 2)
        except Exception: pass

    # Finsight score (displayed v1.3) + breakdown 4 axes (slide 18 cmp)
    _fs_full = _finsight_score_full(snapshot, ratios)
    fs = int(_fs_full.get("displayed") or _fs_full.get("global") or 0) or _finsight_score(snapshot, ratios)
    fs_breakdown = {
        "value":      _fs_full.get("value"),
        "quality":    _fs_full.get("quality"),
        "momentum":   _fs_full.get("momentum"),
        "governance": _fs_full.get("governance"),
    } if _fs_full else {}
    fs_sector_percentile: Optional[int] = None
    try:
        from core.finsight_score import compute_sector_percentile
        yr_sp = _latest_yr(ratios)
        if yr_sp is not None:
            _rd_sp = {
                k: _safe(yr_sp, k) for k in (
                    "pe_ratio", "ev_ebitda", "roic", "net_margin",
                    "ebitda_margin", "revenue_growth", "fcf_yield",
                    "net_debt_ebitda", "momentum_52w",
                )
            }
            _ci = getattr(snapshot, "company_info", None)
            _sec = getattr(_ci, "sector", None) if _ci else None
            _sp = compute_sector_percentile(_rd_sp, _sec)
            if _sp:
                fs_sector_percentile = _sp.get("percentile")
    except Exception as _e_sp:
        log.debug(f"[cmp_societe_xlsx_writer:sector_percentile] skipped: {_e_sp}")
    perf_3m = supp.get("perf_3m")
    mom_sc  = _momentum_score(perf_3m)

    # Revenue CAGR
    rev_cagr = _rev_cagr_3y(ratios)

    # Cash conversion (FCF / Net Income latest year)
    cc = None
    try:
        fcf = _safe(yr, "fcf")
        ni  = _safe(yr, "net_income")
        if fcf is not None and ni and ni != 0:
            cc = round(fcf / abs(ni), 2)
    except Exception as _e:
        log.debug(f"[cmp_societe_xlsx_writer:extract_metrics] exception skipped: {_e}")

    # Sloan accruals depuis qa_python.meta
    sloan = None
    try:
        if qa_python and hasattr(qa_python, "meta"):
            sloan = qa_python.meta.get("sloan_ratio")
    except Exception as _e:
        log.debug(f"[cmp_societe_xlsx_writer:extract_metrics] exception skipped: {_e}")

    # Dividend yield (calculé depuis dividendes_paid / market_cap comme fallback)
    # Ne pas utiliser dividend_payout qui est le taux de distribution (NI), pas le yield
    div_yield = None
    try:
        dps = _safe(ci, "dividends_paid") if ci else None
        mc  = _safe(yr, "market_cap")
        if dps and mc and mc > 0:
            div_yield = round(abs(dps) / mc, 4)
    except Exception as _e:
        log.debug(f"[cmp_societe_xlsx_writer:extract_metrics] exception skipped: {_e}")

    # Market cap / EV en milliards
    mc_bn = None
    ev_bn = None
    try:
        mc_val = _safe(yr, "market_cap")
        ev_val = _safe(yr, "ev")
        if mc_val and mc_val > 0:
            mc_bn = round(mc_val / 1e3, 1)  # yfinance en M → Mds
        if ev_val and ev_val > 0:
            ev_bn = round(ev_val / 1e3, 1)
    except Exception as _e:
        log.debug(f"[cmp_societe_xlsx_writer:extract_metrics] exception skipped: {_e}")

    # Sector : priorité au snapshot, sinon supplément yfinance, sinon fallback
    # live get_ticker(ticker).info["sector"] en dernier recours (cover PPTX S1).
    sector = _safe(ci, "sector") or supp.get("sector") or ""
    if not sector:
        _tkr_for_sec = _safe(snapshot, "ticker") or state.get("ticker", "")
        if _tkr_for_sec:
            try:
                import yfinance as _yf_live
                _live_info = get_ticker(_tkr_for_sec).info or {}
                sector = _live_info.get("sector") or _live_info.get("sectorDisp") or ""
            except Exception:
                sector = ""

    m: dict = {
        # IDENTITE (rows 2-10)
        "company_name_a": _safe(ci, "company_name"),
        "company_name_b": _safe(ci, "company_name"),
        "ticker_a":       _safe(snapshot, "ticker"),
        "ticker_b":       _safe(snapshot, "ticker"),
        "sector_a":       sector,
        "sector_b":       sector,
        "currency_a":     _safe(ci, "currency"),
        "currency_b":     _safe(ci, "currency"),
        "analysis_date":  datetime.today(),

        # VALORISATION (rows 11-27)
        "share_price":      price,
        "dcf_base":         t_base,
        "dcf_bear":         t_bear,
        "dcf_bull":         t_bull,
        "dcf_upside_base":  round((t_base - price) / price, 4) if (t_base and price and price > 0) else None,
        "margin_of_safety": round((t_base - price) / t_base, 4) if (t_base and price and t_base > 0) else None,
        "pe_ratio":         _safe(yr, "pe_ratio"),
        "ev_ebitda":        _safe(yr, "ev_ebitda"),
        "price_to_book":    _safe(yr, "pb_ratio"),
        "fcf_yield":        _safe(yr, "fcf_yield"),
        "peg_ratio":        _safe(yr, "peg_ratio") or _compute_peg(yr, rev_cagr),
        "ev_sales":         _safe(yr, "ev_sales") or supp.get("ev_sales"),
        "wacc":             _safe(mkt, "wacc"),
        "terminal_growth":  _safe(mkt, "terminal_growth"),
        "entry_zone_ok":    1 if (t_base and price and t_base > price) else 0,
        "monte_carlo_p10":  mc_meta.get("dcf_mc_p10"),
        "monte_carlo_p50":  mc_meta.get("dcf_mc_p50"),
        "monte_carlo_p90":  mc_meta.get("dcf_mc_p90"),
        "gbm_sigma_annual": mc_meta.get("gbm_sigma_annual"),
        "gbm_mu_annual":    mc_meta.get("gbm_mu_annual"),

        # QUALITÉ FINANCIÈRE (rows 28-33)
        "piotroski_score": supp.get("piotroski_score"),
        "sloan_accruals":  sloan,
        "cash_conversion": cc,
        "beneish_mscore":  _safe(yr, "beneish_m"),
        "altman_z":        _safe(yr, "altman_z"),
        "altman_model":    _safe(yr, "altman_z_model"),

        # PROFITABILITE & CROISSANCE (rows 34-41)
        "roic":               _safe(yr, "roic"),
        "roe":                _safe(yr, "roe"),
        "ebitda_margin_ltm":  _safe(yr, "ebitda_margin"),
        "ebit_margin":        _safe(yr, "ebit_margin"),
        "net_margin_ltm":     _safe(yr, "net_margin"),
        "ebitda_margin_y1":   _safe(yr1, "ebitda_margin"),
        "ebitda_margin_y2":   _safe(yr2, "ebitda_margin"),
        "ebitda_margin_trend": _ebitda_trend(yr, yr1, yr2),
        "revenue_cagr_3y":    rev_cagr,
        "eps_growth":         _eps_growth(ratios),

        # LEVIER / RISQUE (rows 42-46)
        "net_debt_ebitda":    _safe(yr, "net_debt_ebitda"),
        "interest_coverage":  _safe(yr, "interest_coverage") or supp.get("interest_coverage"),
        "beta":               _safe(mkt, "beta_levered"),
        "rsi":                supp.get("rsi"),
        "var_95_1m":          supp.get("var_95_1m"),
        "volatility_52w":     supp.get("volatility_52w"),
        "duration_implicit":  _compute_duration(_safe(mkt, "wacc"), _safe(mkt, "terminal_growth")),

        # SCORES (rows 47-52) — remplis après les deux sociétés
        "finsight_score": fs,
        "finsight_sector_percentile": fs_sector_percentile,
        # Breakdown 4 axes (slide 18 cmp PPTX) — fix bug décomposition incohérente
        "finsight_score_value":      fs_breakdown.get("value"),
        "finsight_score_quality":    fs_breakdown.get("quality"),
        "finsight_score_momentum":   fs_breakdown.get("momentum"),
        "finsight_score_governance": fs_breakdown.get("governance"),
        "momentum_score": mom_sc,
        "recommendation": _safe(synthesis, "recommendation"),
        "conviction":     _safe(synthesis, "conviction"),
        "verdict_relative": None,  # calculé après comparaison
        "winner":           None,  # calculé après comparaison

        # LIQUIDITÉ (rows 53-55)
        "current_ratio":    _safe(yr, "current_ratio"),
        "quick_ratio":      _safe(yr, "quick_ratio"),
        "capex_to_revenue": _safe(yr, "capex_ratio"),

        # PERFORMANCE (rows 56-58)
        "perf_1m": supp.get("perf_1m"),
        "perf_3m": perf_3m,
        "perf_1y": supp.get("perf_1y"),

        # SECTEUR (rows 59-60)
        "sector_median_pe":       _SECTOR_PE.get(sector),
        "sector_median_ev_ebitda": _SECTOR_EVEBITDA.get(sector),

        # MARCHÉ (rows 61-67)
        "week52_high":       supp.get("week52_high"),
        "week52_low":        supp.get("week52_low"),
        "market_cap":        mc_bn,
        "enterprise_value":  ev_bn,
        "avg_volume_30d":    supp.get("avg_volume_30d"),
        "dividend_yield":    supp.get("dividend_yield") or div_yield,
        "next_earnings_date": supp.get("next_earnings_date"),
        "sentiment_label":   _safe(sentiment, "label"),

        # PIOTROSKI COMPOSANTES (rows 68-76)
        # Source primaire : supp (yfinance direct via _compute_pio_components).
        # Fallback : calcul depuis ratios.years[latest/prev] + snapshot.years
        # si yfinance a échoué (Streamlit Cloud rate limit).
        "pio_roa_positive":         supp.get("pio_roa_positive"),
        "pio_cfo_positive":         supp.get("pio_cfo_positive"),
        "pio_delta_roa":            supp.get("pio_delta_roa"),
        "pio_accruals":             supp.get("pio_accruals"),
        "pio_delta_leverage":       supp.get("pio_delta_leverage"),
        "pio_delta_liquidity":      supp.get("pio_delta_liquidity"),
        "pio_no_dilution":          supp.get("pio_no_dilution"),
        "pio_delta_gross_margin":   supp.get("pio_delta_gross_margin"),
        "pio_delta_asset_turnover": supp.get("pio_delta_asset_turnover"),

        # FORWARD ESTIMATES (consensus analystes)
        "forward_eps":        supp.get("forward_eps"),
        "forward_pe":         supp.get("forward_pe"),
        "trailing_eps":       supp.get("trailing_eps"),
        "revenue_growth_fwd": supp.get("revenue_growth"),
        "eps_growth_fwd":     supp.get("earnings_growth_est"),
    }

    # --- Piotroski F-Score : fallback depuis ratios.years si supp fail -------
    # Quand yfinance échoue (Streamlit Cloud rate limit), `supp` revient avec
    # tous les pio_* à None → P13 du PDF comparatif affiche "—" partout.
    # On recalcule depuis ratios.years[-1] et ratios.years[-2] + snapshot.years
    # (dispo via le pipeline LangGraph déjà exécuté).
    if m.get("pio_roa_positive") is None:
        try:
            yr_now  = yr
            yr_prev = yr1  # année N-1 (déjà extrait plus haut)
            if yr_now and yr_prev:
                _ni0 = _safe(yr_now,  "net_income")
                _ni1 = _safe(yr_prev, "net_income")
                _ta0 = _safe(yr_now,  "total_assets")
                _ta1 = _safe(yr_prev, "total_assets")
                _fcf0 = _safe(yr_now, "fcf")  # proxy CFO
                _gm0 = _safe(yr_now,  "gross_margin")
                _gm1 = _safe(yr_prev, "gross_margin")
                _cr0 = _safe(yr_now,  "current_ratio")
                _cr1 = _safe(yr_prev, "current_ratio")
                _td0 = _safe(yr_now,  "total_debt")
                _td1 = _safe(yr_prev, "total_debt")

                # Revenue depuis snapshot.years (pas dans YearRatios)
                _rev0 = _rev1 = None
                try:
                    _snap_yrs = getattr(snapshot, "years", None) or {}
                    _keys = sorted(_snap_yrs.keys(), reverse=True)
                    if _keys:
                        _rev0 = getattr(_snap_yrs[_keys[0]], "revenue", None)
                    if len(_keys) >= 2:
                        _rev1 = getattr(_snap_yrs[_keys[1]], "revenue", None)
                except Exception as _e:
                    log.debug(f"[cmp_societe_xlsx_writer:extract_metrics] exception skipped: {_e}")

                # 1. ROA > 0
                if _ni0 is not None and _ta0 and _ta0 > 0:
                    m["pio_roa_positive"] = 1 if (_ni0 / _ta0) > 0 else 0
                # 2. CFO > 0 (proxy via FCF puisque CFO pur non disponible dans YR)
                if _fcf0 is not None:
                    m["pio_cfo_positive"] = 1 if _fcf0 > 0 else 0
                # 3. Delta ROA > 0
                if (_ni0 is not None and _ni1 is not None and
                    _ta0 and _ta1 and _ta0 > 0 and _ta1 > 0):
                    m["pio_delta_roa"] = 1 if (_ni0/_ta0) > (_ni1/_ta1) else 0
                # 4. Accruals : FCF / TA > NI / TA  (i.e. FCF > NI)
                if _fcf0 is not None and _ni0 is not None:
                    m["pio_accruals"] = 1 if _fcf0 > _ni0 else 0
                # 5. Delta Levier : total_debt / TA baisse
                if (_td0 is not None and _td1 is not None and
                    _ta0 and _ta1 and _ta0 > 0 and _ta1 > 0):
                    m["pio_delta_leverage"] = 1 if (_td0/_ta0) <= (_td1/_ta1) else 0
                # 6. Delta Liquidité : current_ratio monte
                if _cr0 is not None and _cr1 is not None:
                    m["pio_delta_liquidity"] = 1 if _cr0 >= _cr1 else 0
                # 7. Pas de dilution : shares_outstanding stable (approx via market_cap/price)
                # Non calculable simplement depuis YR — laisse None
                # 8. Delta Gross Margin
                if _gm0 is not None and _gm1 is not None:
                    m["pio_delta_gross_margin"] = 1 if _gm0 >= _gm1 else 0
                # 9. Delta Asset Turnover : revenue / TA monte
                if (_rev0 and _rev1 and _ta0 and _ta1 and
                    _ta0 > 0 and _ta1 > 0):
                    m["pio_delta_asset_turnover"] = 1 if (_rev0/_ta0) >= (_rev1/_ta1) else 0

                # Recalcule piotroski_score depuis les composantes disponibles
                _pio_vals = [m.get(f"pio_{k}") for k in (
                    "roa_positive","cfo_positive","delta_roa","accruals",
                    "delta_leverage","delta_liquidity","no_dilution",
                    "delta_gross_margin","delta_asset_turnover"
                )]
                _pio_set = [v for v in _pio_vals if v is not None]
                if len(_pio_set) >= 5 and m.get("piotroski_score") is None:
                    # Règle de 3 pour normaliser sur 9 points
                    m["piotroski_score"] = round(sum(_pio_set) * 9 / len(_pio_set))
        except Exception as _e_pio_fb:
            log.debug(f"[comparison] piotroski fallback error: {_e_pio_fb}")

    # --- Clés derives pour PDF/PPTX (formatees ou calculées) ---
    # Prix formate (French decimal)
    try:
        if price and price > 0:
            cur = _safe(ci, "currency") or ""
            cur_sym = {"USD": "$", "EUR": "\u20ac", "GBP": "\u00a3", "CHF": "Fr"}.get(cur, cur)
            m["price_str"] = f"{price:,.2f}".replace(",", "\u202f").replace(".", ",") + f" {cur_sym}".strip()
        else:
            m["price_str"] = None
    except Exception:
        m["price_str"] = None

    # Upside % et cible formatee
    try:
        if t_base and price and price > 0:
            upside = (float(t_base) - float(price)) / float(price)
            sign = "+" if upside >= 0 else ""
            m["upside_str"] = f"{sign}{upside*100:.1f}".replace(".", ",") + " %"
            m["target_price"] = f"{float(t_base):,.0f}".replace(",", "\u202f")
        else:
            m["upside_str"]  = None
            m["target_price"] = None
    except Exception:
        m["upside_str"]  = None
        m["target_price"] = None

    # FCF absolu, Net Debt absolu, Cash approx, P/FCF
    # Bug B2 audit 27/04/2026 : yfinance retourne fcf/net_debt/total_debt en
    # MILLIONS USD, mais market_cap est en MILLIARDS. Le formatter _frm s'attend
    # à du milliards uniformément. On normalise tout à milliards à la source.
    try:
        fcf_abs  = _safe(yr, "fcf")
        nd_abs   = _safe(yr, "net_debt")
        td_abs   = _safe(yr, "total_debt")
        mc_raw   = _safe(yr, "market_cap")

        def _to_billions(v):
            """Convertit valeur en millions vers milliards. None → None."""
            if v is None:
                return None
            try:
                return float(v) / 1000.0
            except Exception:
                return None

        m["free_cash_flow"] = _to_billions(fcf_abs)
        m["net_debt"]       = _to_billions(nd_abs)
        cash_abs = None
        if td_abs is not None and nd_abs is not None:
            cash_abs = float(td_abs) - float(nd_abs)
        m["cash"] = _to_billions(cash_abs) if cash_abs is not None else None
        # P/FCF = market_cap / FCF — les deux DOIVENT être dans la même unité.
        # Bug B9 audit 27/04 : après normalisation FCF→milliards (B2 fix), le calc
        # mélangeait mc_raw (millions) avec FCF (milliards) → P/FCF=1005816x absurde.
        # Fix : utiliser fcf_abs original (millions) pour rester cohérent avec mc_raw.
        if mc_raw and fcf_abs and float(fcf_abs) > 0:
            m["p_fcf"] = round(float(mc_raw) / float(fcf_abs), 1)
        else:
            m["p_fcf"] = None
    except Exception:
        m.setdefault("free_cash_flow", None)
        m.setdefault("net_debt", None)
        m.setdefault("cash", None)
        m.setdefault("p_fcf", None)

    return m


def _compute_duration(wacc, tgr) -> Optional[float]:
    """Duration implicite = (1+g) / (WACC - g). Retourne en années."""
    try:
        if wacc is None or tgr is None:
            return None
        w = float(wacc)
        g = float(tgr)
        # WACC peut être en % (ex: 9.5) ou decimal (ex: 0.095)
        if w > 1: w = w / 100.0
        if g > 1: g = g / 100.0
        if w <= g or w <= 0:
            return None
        dur = (1 + g) / (w - g)
        if 0 < dur < 100:
            return round(dur, 1)
        return None
    except Exception:
        return None


def _compute_peg(yr, rev_cagr) -> Optional[float]:
    """PEG = PE / (revenue_cagr * 100)."""
    try:
        pe = _safe(yr, "pe_ratio")
        if pe and pe > 0 and rev_cagr and rev_cagr > 0:
            peg = pe / (rev_cagr * 100)
            if 0 < peg < 50:
                return round(peg, 2)
        return None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Writer principal
# ---------------------------------------------------------------------------

class CmpSocieteXlsxWriter:
    """Injecte deux états pipeline dans TEMPLATE_COMPARISON.xlsx."""

    def write(self, state_a: dict, state_b: dict, language: str = "fr", currency: str = "EUR") -> bytes:
        """
        Génère le fichier XLSX comparaison en mémoire.

        Retourne les bytes du fichier.
        Lève RuntimeError si le template est introuvable.
        """
        self._language = language
        self._currency = currency
        import openpyxl

        # 1. Trouver le template
        template_path = _find_template()
        if template_path is None:
            raise RuntimeError(
                "TEMPLATE_COMPARISON.xlsx introuvable. "
                "Vérifier OneDrive ou copier dans le répertoire projet."
            )

        # 2. Build context (dedupe refactor #191 — flow partagé avec pptx/pdf)
        #    with_synthesis=False : XLSX n'a pas besoin du verdict LLM, et on
        #    évite l'appel LLM si ce writer tourne seul. Si le cache session_state
        #    est déjà chaud (PPTX/PDF ont tourné avant), on bénéficie du winner LLM.
        from outputs.cmp_societe_common import build_cmp_context
        ctx = build_cmp_context(state_a, state_b, with_synthesis=False)
        tkr_a  = ctx.tkr_a
        tkr_b  = ctx.tkr_b
        supp_a = ctx.supp_a
        supp_b = ctx.supp_b
        m_a    = ctx.m_a
        m_b    = ctx.m_b
        log.info(f"[CmpSocieteXlsxWriter] context {tkr_a} / {tkr_b} (winner={ctx.winner})")

        # 3. Verdict_relative spécifique XLSX (utilise les NOMS affichables)
        name_a = ctx.name_a or tkr_a
        name_b = ctx.name_b or tkr_b
        if ctx.winner == tkr_a:
            verdict = f"{name_a} privilegie"
        elif ctx.winner == tkr_b:
            verdict = f"{name_b} privilegie"
        else:
            verdict = "Équivalent"
        m_a["verdict_relative"] = verdict
        m_b["verdict_relative"] = verdict

        # 5. Charger template
        wb = openpyxl.load_workbook(str(template_path))
        ws = wb["DATA_RAW"]

        # 6. Construire la liste ordonnée des métriques depuis le template (col A)
        row_map: dict[str, int] = {}  # metric_name → row_number
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=False):
            cell = row[0]
            if cell.value:
                row_map[str(cell.value).strip()] = cell.row

        # 7. Effacer les anciennes valeurs exemple (col C et E) avant injection
        for row_idx in row_map.values():
            ws.cell(row=row_idx, column=3).value = None
            ws.cell(row=row_idx, column=5).value = None

        # 8. Injecter valeurs — col C = société A, col E = société B
        for metric, row_idx in row_map.items():
            val_a = m_a.get(metric)
            val_b = m_b.get(metric)
            _write_cell(ws, row_idx, 3, val_a)  # col C
            _write_cell(ws, row_idx, 5, val_b)  # col E

        # 9b. Supprimer les doublons dans COMPARABLES (company A ou B dans pairs)
        try:
            ws_comp = wb["COMPARABLES"]
            tickers_to_exclude = {tkr_a.upper(), tkr_b.upper()} if tkr_a and tkr_b else set()
            rows_to_delete = []
            for row in ws_comp.iter_rows():
                ticker_cell = None
                for cell in row:
                    if cell.column == 3 and cell.value:  # col C = TICKER
                        ticker_cell = cell
                        break
                if ticker_cell and str(ticker_cell.value).upper() in tickers_to_exclude:
                    rows_to_delete.append(ticker_cell.row)
            # Supprimer de bas en haut pour ne pas decaler les indices
            for r in sorted(rows_to_delete, reverse=True):
                ws_comp.delete_rows(r)
        except Exception as _ce2:
            log.warning(f"[CmpSocieteXlsxWriter] comparables dedup: {_ce2}")

        # 9. Mettre à jour les noms de series dans TOUS les graphiques +
        #    remplacer les labels "Société A/B" des cellules DASHBOARD par les tickers.
        try:
            from openpyxl.chart.series import SeriesLabel
            from openpyxl.chart.data_source import StrRef

            # (a) Remplacer les labels textuels "Société A" / "Société B" dans DASHBOARD.
            ws_dash = wb["DASHBOARD"]
            name_a_display = m_a.get("company_name_a") or tkr_a
            name_b_display = m_b.get("company_name_b") or tkr_b
            _LABEL_TOKENS_A = ("Société A", "Société A", "Soci\u00e9t\u00e9 A", "Soc. A")
            _LABEL_TOKENS_B = ("Société B", "Société B", "Soci\u00e9t\u00e9 B", "Soc. B")
            for row in ws_dash.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    sv = str(cell.value)
                    if any(tok in sv for tok in _LABEL_TOKENS_A):
                        cell.value = name_a_display
                    elif any(tok in sv for tok in _LABEL_TOKENS_B):
                        cell.value = name_b_display

            # (b) Renommer les series dans TOUS les graphiques de TOUTES les feuilles.
            # Certaines series ont des titres hardcodes (ex: "Microsoft" / "Alphabet")
            # herites du template original et ne se mettent pas à jour via formules.
            for sheet_name in wb.sheetnames:
                ws_any = wb[sheet_name]
                if not hasattr(ws_any, "_charts") or not ws_any._charts:
                    continue
                for chart in ws_any._charts:
                    for i, ser in enumerate(chart.series):
                        new_name = tkr_a if i % 2 == 0 else tkr_b
                        try:
                            ser.tx = SeriesLabel(v=new_name)
                        except Exception as _e:
                            log.debug(
                                f"[CmpSocieteXlsxWriter] {sheet_name} chart series rename error: {_e}"
                            )
        except Exception as _ce:
            log.warning(f"[CmpSocieteXlsxWriter] chart rename global: {_ce}")

        # Sauvegarder en mémoire
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        log.info(f"[CmpSocieteXlsxWriter] fichier Généré ({buf.getbuffer().nbytes // 1024} ko)")
        return buf.read()


def _find_template() -> Optional[Path]:
    """Cherche TEMPLATE_COMPARISON.xlsx dans OneDrive puis dans le projet."""
    # Chemin relatif depuis ce fichier (outputs/ -> assets/)
    _project_root = Path(__file__).parent.parent
    candidates = [
        _TEMPLATE_SRC,
        _project_root / "assets" / "TEMPLATE_COMPARISON.xlsx",
        Path("C:/Users/bapti/finsight-ia/assets/TEMPLATE_COMPARISON.xlsx"),
        Path("C:/Users/bapti/finsight-ia/TEMPLATE_COMPARISON.xlsx"),
    ]
    for p in candidates:
        if p.exists():
            try:
                # Copie pour éviter le PermissionError si Excel est ouvert
                tmp = _project_root / "tmp_cmp_template.xlsx"
                shutil.copy2(str(p), str(tmp))
                return tmp
            except Exception:
                return p
    return None


def _write_cell(ws, row: int, col: int, value) -> None:
    """Écrit une valeur dans une cellule (skip None)."""
    if value is None:
        return
    try:
        ws.cell(row=row, column=col, value=value)
    except Exception as e:
        log.debug(f"[CmpSocieteXlsxWriter] cell({row},{col})={value!r} erreur: {e}")
