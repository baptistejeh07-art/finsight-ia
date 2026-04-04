# =============================================================================
# FinSight IA -- Indice Excel Writer
# outputs/indice_excel_writer.py
#
# Injecte les donnees d'analyse d'indice dans TEMPLATE_INDICE.xlsx.
#
# REGLES ABSOLUES :
#   - NE PAS TOUCHER les cellules formule (VALUE cols O-P, QUALITY cols O-R, etc.)
#   - SEULE feuille DONNEES BRUTES est 100% injection Python
#   - Feuilles par secteur (TECHNOLOGY, CONSUMER DISC., etc.) = formules auto
#   - Guard systematique : if cell.value startswith("=") -> skip
# =============================================================================

from __future__ import annotations

import logging
import shutil
from datetime import date
from pathlib import Path

log = logging.getLogger(__name__)

# Template source
_TEMPLATE = Path(__file__).parent.parent / "assets" / "TEMPLATE_INDICE.xlsx"

# Sentinelle valeur manquante
_NA = "\u2014"   # em-dash

# Barres visuelles (U+2588 plein, U+2591 clair)
_FULL  = "\u2588"
_EMPTY = "\u2591"

# Mapping secteur yfinance canonical -> nom affiche dans template
# Couvre les deux nomenclatures yfinance (ancienne et moderne)
_SECT_DISP = {
    # Noms anciens yfinance (GICS)
    "Technology":             "Technology",
    "Health Care":            "Healthcare",
    "Financials":             "Financials",
    "Consumer Discretionary": "Consumer Disc.",
    "Communication Services": "Comm. Services",
    "Industrials":            "Industrials",
    "Consumer Staples":       "Consumer Defensive",
    "Energy":                 "Energy",
    "Materials":              "Materials",
    "Real Estate":            "Real Estate",
    "Utilities":              "Utilities",
    # Noms modernes yfinance (retournes pour tickers EU)
    "Financial Services":     "Financials",
    "Consumer Cyclical":      "Consumer Disc.",
    "Consumer Defensive":     "Consumer Defensive",
    "Basic Materials":        "Materials",
    "Healthcare":             "Healthcare",
    "Comm. Services":         "Comm. Services",
}

# Ordre des secteurs (pour PAR SECTEUR, SECTOR OVERVIEW col M)
# Valeurs uniques preservant l'ordre (dict preserving insertion order Python 3.7+)
_SECTOR_ORDER = list(dict.fromkeys(_SECT_DISP.values()))

# Orientation signal depuis score
def _orientation(score: int) -> str:
    if score >= 65: return "\u2191 Positif"
    if score < 45:  return "\u2193 Negatif"
    return "\u2192 Neutre"

def _bar(score: int, n: int = 10) -> str:
    filled = max(0, min(n, round(score / 10)))
    return _FULL * filled + _EMPTY * (n - filled)

# ---------------------------------------------------------------------------
# Helpers format
# ---------------------------------------------------------------------------

def _v(val, fallback=_NA):
    """Retourne val si non-None sinon fallback."""
    return val if val is not None else fallback

def _pct(v, decimals=1, sign=False):
    """Decimal -> string '%'. v=0.711 -> '71.1%'."""
    if v is None: return _NA
    try:
        s = f"{float(v)*100:.{decimals}f}%"
        if sign and float(v) >= 0: s = "+" + s
        return s
    except Exception:
        return _NA

def _pct_raw(v, decimals=1, sign=True):
    """Deja en % -> string. v=5.5 -> '+5.5%'."""
    if v is None: return _NA
    try:
        s = f"{float(v):.{decimals}f}%"
        if sign and float(v) >= 0: s = "+" + s
        return s
    except Exception:
        return _NA

def _num(v, decimals=2):
    if v is None: return _NA
    try: return round(float(v), decimals)
    except Exception: return _NA

def _x(v, decimals=1):
    if v is None: return _NA
    try: return f"{float(v):.{decimals}f}x"
    except Exception: return _NA

def _mds(v, decimals=1, prefix=""):
    """En Mds (valeur deja en Mds)."""
    if v is None: return _NA
    try: return f"{prefix}{float(v):.{decimals}f}"
    except Exception: return _NA

def _price_str(v):
    if v is None: return _NA
    try: return round(float(v), 2)
    except Exception: return _NA

def _score_str(score: int | None) -> str:
    if score is None: return _NA
    return f"{score}/100"

# ---------------------------------------------------------------------------
# Scoring 4D par percentile ranking
# ---------------------------------------------------------------------------

def _percentile_scores(values: list, higher_is_better: bool = True) -> list[int]:
    """Rank une liste de valeurs -> scores 0-100. None -> 0."""
    valid = [(i, float(v)) for i, v in enumerate(values) if v is not None]
    n = len(valid)
    result = [0] * len(values)
    if n == 0:
        return result
    sorted_valid = sorted(valid, key=lambda x: x[1], reverse=higher_is_better)
    for rank, (i, _) in enumerate(sorted_valid):
        result[i] = round(100 - (rank / max(1, n - 1)) * 100) if n > 1 else 50
    return result

def _compute_4d_scores(tickers: list[dict]) -> list[dict]:
    """
    Calcule VALUE, GROWTH, QUALITY, MOMENTUM, GLOBAL pour chaque ticker.
    Retourne la liste enrichie avec champs score_value/growth/quality/momentum/global.
    """
    n = len(tickers)

    # ---- VALUE : faible EV/EBITDA, faible P/E, forte mg brute, bon Altman Z, FCF Yield positif
    s_ev  = _percentile_scores([t.get("ev_ebitda")    for t in tickers], higher_is_better=False)
    s_pe  = _percentile_scores([t.get("pe_trailing")  for t in tickers], higher_is_better=False)
    s_gm  = _percentile_scores([t.get("gross_margins") for t in tickers], higher_is_better=True)
    s_az  = _percentile_scores([t.get("altman_z")     for t in tickers], higher_is_better=True)
    s_fcf = _percentile_scores([t.get("fcf_yield")    for t in tickers], higher_is_better=True)
    val   = [round(s_ev[i]*0.30 + s_pe[i]*0.20 + s_gm[i]*0.20 + s_az[i]*0.15 + s_fcf[i]*0.15)
             for i in range(n)]

    # ---- GROWTH : forte croissance revenus, forte mg EBITDA, bon ROE
    s_rg  = _percentile_scores([t.get("rev_growth")      for t in tickers], higher_is_better=True)
    s_em  = _percentile_scores([t.get("ebitda_margins")  for t in tickers], higher_is_better=True)
    s_roe = _percentile_scores([t.get("roe")             for t in tickers], higher_is_better=True)
    s_eg  = _percentile_scores([t.get("earnings_growth") for t in tickers], higher_is_better=True)
    gr    = [round(s_rg[i]*0.35 + s_em[i]*0.30 + s_roe[i]*0.20 + s_eg[i]*0.15)
             for i in range(n)]

    # ---- QUALITY : Altman Z eleve, faible ND/EBITDA, bon current ratio, forte mg nette
    s_az2 = _percentile_scores([t.get("altman_z")     for t in tickers], higher_is_better=True)
    s_nd  = _percentile_scores([t.get("nd_ebitda")    for t in tickers], higher_is_better=False)
    s_cr  = _percentile_scores([t.get("current_ratio") for t in tickers], higher_is_better=True)
    s_pm  = _percentile_scores([t.get("profit_margins") for t in tickers], higher_is_better=True)
    ql    = [round(s_az2[i]*0.30 + s_nd[i]*0.25 + s_em[i]*0.20 + s_cr[i]*0.15 + s_pm[i]*0.10)
             for i in range(n)]

    # ---- MOMENTUM : 52W momentum
    s_mo  = _percentile_scores([t.get("mom_52w") for t in tickers], higher_is_better=True)
    mo    = [s_mo[i] for i in range(n)]

    # ---- GLOBAL composite
    gl    = [round(val[i]*0.25 + gr[i]*0.30 + ql[i]*0.25 + mo[i]*0.20) for i in range(n)]

    enriched = []
    for i, t in enumerate(tickers):
        t2 = dict(t)
        t2["score_value"]    = val[i]
        t2["score_growth"]   = gr[i]
        t2["score_quality"]  = ql[i]
        t2["score_momentum"] = mo[i]
        t2["score_global"]   = gl[i]
        enriched.append(t2)
    return enriched

# ---------------------------------------------------------------------------
# Agregats par secteur
# ---------------------------------------------------------------------------

def _sector_aggregates(tickers: list[dict]) -> dict:
    """Retourne un dict sector_name -> {nb, score, top_ticker, mg_ebitda, rev_growth, altman_z}."""
    from collections import defaultdict
    import statistics

    by_sec = defaultdict(list)
    for t in tickers:
        _s_raw = (t.get("sector") or "").strip()
        sec = _SECT_DISP.get(_s_raw, _s_raw or "Autre")
        if sec and sec.strip() and sec != "Autre":
            by_sec[sec].append(t)

    agg = {}
    for sec, members in by_sec.items():
        scores = [m["score_global"] for m in members]
        avg_sc = round(statistics.mean(scores))
        top = max(members, key=lambda m: m["score_global"])

        def _med(field, pct=False):
            vals = [m.get(field) for m in members if m.get(field) is not None]
            if not vals: return None
            v = statistics.median(vals)
            return v * 100 if pct else v

        agg[sec] = {
            "nb":         len(members),
            "score":      avg_sc,
            "top_ticker": top.get("ticker", ""),
            "top_name":   (top.get("name") or top.get("ticker") or "")[:20],
            "mg_ebitda":  _med("ebitda_margins"),  # decimal
            "rev_growth": _med("rev_growth"),       # decimal
            "altman_z":   _med("altman_z"),
            "signal":     (_orientation(avg_sc)
                           .replace("\u2191 ", "").replace("\u2192 ", "").replace("\u2193 ", "")),
        }
    return agg


# ---------------------------------------------------------------------------
# Ecriture cellule (guard formule)
# ---------------------------------------------------------------------------

def _write(ws, row: int, col: int | str, value):
    """Ecrit dans une cellule sauf si elle contient deja une formule ou ArrayFormula."""
    if isinstance(col, str):
        cell = ws[f"{col}{row}"]
    else:
        cell = ws.cell(row=row, column=col)
    try:
        from openpyxl.worksheet.formula import ArrayFormula
        if isinstance(cell.value, ArrayFormula):
            return  # guard formule matricielle
    except ImportError:
        pass
    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
        return  # guard formule normale
    cell.value = value


# ---------------------------------------------------------------------------
# Remplissage feuille DONNEES BRUTES
# ---------------------------------------------------------------------------

def _fill_donnees_brutes(ws, tickers: list[dict], universe: str) -> None:
    """
    Colonnes A-AD, lignes 3-42 (max 40 tickers).
    Ligne 1: titre, Ligne 2: headers (pre-existants, pas touches).
    """
    # Titre ligne 1
    _write(ws, 1, "A", f"FinSight IA  \u00b7  {universe}  \u00b7  Donnees brutes")

    # Effacer les lignes 3-42 pour eviter donnees stales du template
    try:
        from openpyxl.worksheet.formula import ArrayFormula as _AF
    except ImportError:
        _AF = None
    for clr_row in range(3, 43):
        for clr_col in range(1, 31):
            cell = ws.cell(row=clr_row, column=clr_col)
            if _AF and isinstance(cell.value, _AF):
                continue
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None

    for offset, t in enumerate(tickers[:40]):
        row = 3 + offset
        sec_disp = _SECT_DISP.get(t.get("sector") or "", (t.get("sector") or "Autre"))

        # A: Ticker
        _write(ws, row, 1,  t.get("ticker", ""))
        # B: Societe
        _write(ws, row, 2,  t.get("name",   ""))
        # C: Secteur
        _write(ws, row, 3,  sec_disp)
        # D: Cours
        _write(ws, row, 4,  _price_str(t.get("price")))
        # E: Mkt Cap (Mds)
        _write(ws, row, 5,  _num(t.get("mkt_cap"), 1))
        # F: EV (Mds)
        _write(ws, row, 6,  _num(t.get("ev"), 1))
        # G: Rev LTM (Mds)
        _write(ws, row, 7,  _mds(t.get("rev_ltm")))
        # H: EBITDA LTM (Mds)
        _write(ws, row, 8,  _mds(t.get("ebitda_ltm")))
        # I: EV/EBITDA
        _write(ws, row, 9,  _num(t.get("ev_ebitda"), 1))
        # J: EV/Rev
        _write(ws, row, 10, _num(t.get("ev_revenue"), 1))
        # K: P/E
        _write(ws, row, 11, _num(t.get("pe_trailing"), 1))
        # L: EPS
        _write(ws, row, 12, _num(t.get("eps"), 2))
        # M: Mg Brute %
        _write(ws, row, 13, _pct(t.get("gross_margins")))
        # N: Mg EBITDA %
        _write(ws, row, 14, _pct(t.get("ebitda_margins")))
        # O: Mg Nette %
        _write(ws, row, 15, _pct(t.get("profit_margins")))
        # P: Cro Rev %
        _write(ws, row, 16, _pct(t.get("rev_growth"), sign=True))
        # Q: ROE %
        _write(ws, row, 17, _pct(t.get("roe"), sign=True))
        # R: ROA %
        _write(ws, row, 18, _pct(t.get("roa"), sign=True))
        # S: Current Ratio
        _write(ws, row, 19, _num(t.get("current_ratio"), 2))
        # T: ND/EBITDA
        _nd = t.get("nd_ebitda")
        _write(ws, row, 20, _x(_nd) if _nd is not None else _NA)
        # U: Altman Z
        _write(ws, row, 21, _num(t.get("altman_z"), 2))
        # V: Beneish M
        _write(ws, row, 22, _num(t.get("beneish_m"), 2))
        # W: Mom 52W %
        _mom = t.get("mom_52w")
        _write(ws, row, 23, _pct_raw(_mom, sign=True) if _mom is not None else _NA)
        # X: Score Value
        _write(ws, row, 24, t.get("score_value"))
        # Y: Score Growth
        _write(ws, row, 25, t.get("score_growth"))
        # Z: Score Quality
        _write(ws, row, 26, t.get("score_quality"))
        # AA: Score Momentum
        _write(ws, row, 27, t.get("score_momentum"))
        # AB: Score Global
        _write(ws, row, 28, t.get("score_global"))
        # AC: Next Earnings
        _write(ws, row, 29, t.get("next_earnings") or _NA)
        # AD: Signal — calcule depuis score_global recompute (evite signal stale de l'input)
        _sg = t.get("score_global")
        _signal = ("Surpond\u00e9rer" if (_sg or 0) >= 60
                   else ("Sous-pond\u00e9rer" if (_sg or 0) < 40 else "Neutre"))
        _write(ws, row, 30, _signal)


# ---------------------------------------------------------------------------
# Remplissage feuilles VALUE / GROWTH / QUALITY / MOMENTUM
# ---------------------------------------------------------------------------

def _fill_ranking_sheet(ws, tickers_sorted: list[dict], score_field: str,
                        cols: list[tuple], universe: str, label: str) -> None:
    """
    Remplit une feuille de classement (VALUE, GROWTH, QUALITY, MOMENTUM).
    tickers_sorted : deja tries par score_field desc.
    cols : liste de (col_letter, field_name_or_callable).
    Donnees en lignes 4-22 (max 19 tickers).
    Colonnes formule (O, P, Q, R) jamais touchees grace au guard.
    """
    # Ligne 1: titre
    _write(ws, 1, "A", f"FinSight IA  \u00b7  Analyse {label}  \u00b7  {universe}")
    _write(ws, 2, "A", "Usage confidentiel  \u00b7  Calculs Python/yfinance")

    # Limite a 15 lignes — formatage conditionnel du template couvre lignes 4-18 uniquement
    MAX_ROWS = 15
    tickers_capped = tickers_sorted[:MAX_ROWS]

    for offset, t in enumerate(tickers_capped):
        row = 4 + offset
        sec_disp = _SECT_DISP.get(t.get("sector") or "", (t.get("sector") or "Autre"))

        # Rang
        _write(ws, row, "A", offset + 1)
        # Ticker
        _write(ws, row, "B", t.get("ticker", ""))
        # Societe
        _write(ws, row, "C", t.get("name", ""))
        # Secteur
        _write(ws, row, "D", sec_disp)
        # Score (format "75/100")
        _write(ws, row, "E", _score_str(t.get(score_field)))

        # Colonnes metriques (col lettre -> fonction)
        for col_letter, getter in cols:
            val = getter(t) if callable(getter) else t.get(getter)
            _write(ws, row, col_letter, val)

    # Effacer les lignes vides au-dela du nombre de tickers (jusqu'a ligne 19 max)
    for row in range(4 + len(tickers_capped), 4 + MAX_ROWS):
        for col_letter in [c for c, _ in cols] + ["A", "B", "C", "D", "E"]:
            _write(ws, row, col_letter, None)


# ---------------------------------------------------------------------------
# Remplissage feuille PAR SECTEUR
# ---------------------------------------------------------------------------

def _fill_par_secteur(ws, sector_agg: dict, universe: str, nb_total: int,
                      today_str: str) -> None:
    """
    Remplit la feuille PAR SECTEUR.

    Layout template (fixe, ne pas modifier) :
      Lignes 1-6  : header zone (titre, sous-titre, SCORECARD title)
      Ligne 7     : headers SCORECARD (statique template)
      Lignes 8-17 : donnees SCORECARD (max 10 secteurs)
      Ligne 18    : total (fixe)
      Ligne 19    : 'COMPOSITION DETAIL' titre (STATIQUE template — ne pas toucher)
      Ligne 20    : separateur (STATIQUE template — ne pas toucher)
      Ligne 21    : headers COMPOSITION DETAIL (STATIQUE template — ne pas toucher)
      Lignes 22+  : donnees COMPOSITION DETAIL (injection Python)

    Col L-M  : aux tables pour graphiques (score par secteur, lignes 1-12)
    Col H-I  : aux tables Mg EBITDA (lignes 20+), Nb Soc (lignes apres)
    """
    import statistics

    # Tri secteurs par score desc
    sectors_sorted = sorted(sector_agg.items(), key=lambda x: x[1]["score"], reverse=True)
    # Scorecard : max 10 secteurs (template formate lignes 8-17)
    scorecard_sectors = sectors_sorted[:10]

    # ---- Header ----
    _write(ws, 1, "B", f"FinSight IA  \u00b7  {universe} \u2014 Vue Macro Sectorielle")
    _write(ws, 1, "H", today_str)

    # ---- Table auxiliaire col L-M (lignes 1-12) : score par secteur pour graphique ----
    _write(ws, 1, "L", "Secteur")
    _write(ws, 1, "M", "Score")
    for i, (sec, agg) in enumerate(sectors_sorted[:11]):
        _write(ws, 2 + i, "L", sec)
        _write(ws, 2 + i, "M", agg["score"])
    # Effacer lignes non utilisees
    for clr_i in range(len(sectors_sorted), 11):
        _write(ws, 2 + clr_i, "L", None)
        _write(ws, 2 + clr_i, "M", None)

    # ---- SCORECARD SECTORIELLE ----
    _write(ws, 5, "B", "SCORECARD SECTORIELLE  \u00b7  Score composite, metriques medianes, signal d'investissement \u2014 " + universe)

    # Effacer d'abord toute la zone de donnees scorecard (lignes 8-17)
    for clr in range(8, 18):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            cell = ws[f"{col}{clr}"]
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    for offset, (sec, agg) in enumerate(scorecard_sectors):
        row = 8 + offset
        _write(ws, row, "B", sec)
        _write(ws, row, "C", agg["nb"])
        _write(ws, row, "D", agg["score"])
        _write(ws, row, "E", _bar(agg["score"]))
        _write(ws, row, "F", agg["top_name"])
        mg = agg.get("mg_ebitda")
        _write(ws, row, "G", round(mg, 3) if (mg is not None and mg > 0.005) else _NA)
        rg = agg.get("rev_growth")
        _write(ws, row, "H", round(rg, 3) if rg is not None else _NA)
        az = agg.get("altman_z")
        _write(ws, row, "I", round(az, 2) if az is not None else _NA)
        _write(ws, row, "J", _orientation(agg["score"]).replace("\u2191 ", "").replace("\u2192 ", "").replace("\u2193 ", ""))

    # Ligne total (toujours en ligne 18)
    tot_score = round(sum(a["score"] for a in sector_agg.values()) / max(1, len(sector_agg)))
    mg_vals   = [a["mg_ebitda"] for a in sector_agg.values() if a.get("mg_ebitda") and a["mg_ebitda"] > 0.005]
    rg_vals   = [a["rev_growth"] for a in sector_agg.values() if a.get("rev_growth") is not None]
    tot_mg    = round(statistics.median(mg_vals), 3) if mg_vals else _NA
    tot_rg    = round(statistics.median(rg_vals), 3) if rg_vals else _NA

    _write(ws, 18, "B", f"{universe} \u2014 Total")
    _write(ws, 18, "C", nb_total)
    _write(ws, 18, "D", tot_score)
    _write(ws, 18, "F", "Mediane")
    _write(ws, 18, "G", tot_mg)
    _write(ws, 18, "H", tot_rg)

    # Lignes 19-21 : STATIQUES dans le template (titre + headers COMPOSITION DETAIL)
    # NE PAS TOUCHER ces lignes

    # ---- COMPOSITION DETAIL (lignes 22+) ----
    # Effacer d'abord la zone de donnees existante (lignes 22-35)
    for clr in range(22, 36):
        for col in ["B", "C", "D", "E", "F"]:
            cell = ws[f"{col}{clr}"]
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    for offset, (sec, agg) in enumerate(sectors_sorted):
        crow = 22 + offset
        _write(ws, crow, "B", sec)
        _write(ws, crow, "C", agg["nb"])
        _write(ws, crow, "D", _score_str(agg["score"]))
        _write(ws, crow, "E", _bar(agg["score"]))
        _write(ws, crow, "F", _orientation(agg["score"]))

    # ---- Tables auxiliaires col H-I (lignes 20+) pour graphiques charts ----
    # Mg EBITDA (lignes 20-31)
    mg_sorted = sorted(sector_agg.items(),
                       key=lambda x: x[1]["mg_ebitda"] or 0, reverse=True)
    _write(ws, 20, "H", "Secteur")
    _write(ws, 20, "I", "Mg EBITDA %")
    for j, (sec, agg) in enumerate(mg_sorted):
        mg = agg.get("mg_ebitda")
        _write(ws, 21 + j, "H", sec)
        _write(ws, 21 + j, "I", round(mg * 100, 1) if (mg and mg > 0.005) else _NA)

    # Nb societes par secteur (lignes apres mg_ebitda + 2)
    nb_sorted = sorted(sector_agg.items(), key=lambda x: x[1]["nb"], reverse=True)
    nb_aux_r  = 21 + len(mg_sorted) + 2
    _write(ws, nb_aux_r,     "H", "Secteur")
    _write(ws, nb_aux_r,     "I", "Nb Soc.")
    for j, (sec, agg) in enumerate(nb_sorted):
        _write(ws, nb_aux_r + 1 + j, "H", sec)
        _write(ws, nb_aux_r + 1 + j, "I", agg["nb"])


# ---------------------------------------------------------------------------
# Remplissage SECTOR OVERVIEW col M (noms secteurs pour formules)
# ---------------------------------------------------------------------------

def _fill_sector_overview(ws, sector_agg: dict) -> None:
    """Injecte les noms de secteurs en col M (M1:M21) pour alimenter les formules."""
    # Les formules de SECTOR OVERVIEW referentent 'SECTOR OVERVIEW'!M$1:M$21
    # On ecrit les secteurs presents tries alphabetiquement
    sectors_alpha = sorted(k for k in sector_agg.keys() if k and k.strip() and k != "Autre")
    for i, sec in enumerate(sectors_alpha[:21]):
        _write(ws, 1 + i, "M", sec)
    # Effacer les lignes vides au-dela
    for i in range(len(sectors_alpha), 21):
        cell = ws.cell(row=1 + i, column=13)
        if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
            cell.value = None


# ---------------------------------------------------------------------------
# Remplissage DASHBOARD (date + titre uniquement — le reste = formules)
# ---------------------------------------------------------------------------

def _fill_dashboard(ws, universe: str, today_str: str) -> None:
    """
    DASHBOARD : injecte la date (V3) et le titre (A1).
    D3, P3, J3 et les lignes 8-12 sont des formules -> skip automatique via guard.
    """
    _write(ws, 1, "A", f"FinSight IA  \u00b7  Tableau de Bord \u2014 Screening {universe}")
    _write(ws, 3, "V", today_str)


# ---------------------------------------------------------------------------
# Classe principale
# ---------------------------------------------------------------------------

class IndiceExcelWriter:

    @staticmethod
    def generate(data: dict, output_path: str,
                 template_path: str | None = None) -> None:
        """
        Genere l'Excel indice en injectant les donnees dans le template.

        Args:
            data         : dict retourne par _fetch_real_indice_data / _make_test_indice_data
            output_path  : chemin de sortie .xlsx
            template_path: optionnel, chemin vers template (defaut: assets/TEMPLATE_INDICE.xlsx)
        """
        try:
            import openpyxl
        except ImportError:
            log.error("[IndiceExcelWriter] openpyxl non installe")
            return

        tpl = Path(template_path) if template_path else _TEMPLATE
        if not tpl.exists():
            log.error("[IndiceExcelWriter] Template introuvable: %s", tpl)
            return

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        # Copie du template
        shutil.copy2(tpl, out)
        log.info("[IndiceExcelWriter] Template copie -> %s", out)

        # Donnees tickers
        tickers_raw: list[dict] = data.get("tickers_raw", [])
        universe: str = data.get("universe", data.get("indice", "Indice"))

        if not tickers_raw:
            log.warning("[IndiceExcelWriter] tickers_raw vide — Excel non rempli")
            return

        # Calcul des 4 scores par percentile
        tickers = _compute_4d_scores(tickers_raw)

        # Agregats par secteur
        sector_agg = _sector_aggregates(tickers)

        # Date du jour
        today_str = date.today().strftime("%d/%m/%Y")

        # Tris pour chaque dimension
        by_value    = sorted(tickers, key=lambda t: t.get("score_value", 0),    reverse=True)
        by_growth   = sorted(tickers, key=lambda t: t.get("score_growth", 0),   reverse=True)
        by_quality  = sorted(tickers, key=lambda t: t.get("score_quality", 0),  reverse=True)
        by_momentum = sorted(tickers, key=lambda t: t.get("score_momentum", 0), reverse=True)
        by_global   = sorted(tickers, key=lambda t: t.get("score_global", 0),   reverse=True)

        # Ouvrir le workbook copie
        wb = openpyxl.load_workbook(str(out))

        # ---- DONNEES BRUTES ----
        db_name = next((s for s in wb.sheetnames
                        if "brutes" in s.lower() or "brut" in s.lower()), None)
        if db_name:
            _fill_donnees_brutes(wb[db_name], by_global, universe)
        else:
            log.warning("[IndiceExcelWriter] Feuille DONNEES BRUTES introuvable")

        # ---- VALUE ----
        def _v_ev(t):   return _num(t.get("ev_ebitda"), 1)
        def _v_evr(t):  return _num(t.get("ev_revenue"), 1)
        def _v_pe(t):   return _num(t.get("pe_trailing"), 1)
        def _v_gm(t):   return _pct(t.get("gross_margins"))
        def _v_em(t):   return _pct(t.get("ebitda_margins"))
        def _v_az(t):   return _num(t.get("altman_z"), 2)
        def _v_fcf(t):  return _pct_raw(t.get("fcf_yield"), sign=True) if t.get("fcf_yield") is not None else _NA
        def _v_dcf(t):  return _NA   # Decote DCF: non calculee en V1

        value_cols = [
            ("F", _v_ev), ("G", _v_evr), ("H", _v_pe),
            ("I", _v_gm), ("J", _v_em),  ("K", _v_az),
            ("L", _v_fcf), ("M", _v_dcf),
        ]
        if "VALUE" in wb.sheetnames:
            _fill_ranking_sheet(wb["VALUE"], by_value, "score_value",
                                value_cols, universe, "VALUE")

        # ---- GROWTH ----
        def _g_rg(t):  return _pct(t.get("rev_growth"), sign=True)
        def _g_em(t):  return _pct(t.get("ebitda_margins"))
        def _g_nm(t):  return _pct(t.get("profit_margins"))
        def _g_roe(t): return _pct(t.get("roe"), sign=True)
        def _g_eg(t):  return _pct(t.get("earnings_growth"), sign=True)
        def _g_rev(t):
            v = t.get("analyst_revision")
            if v is None: return _NA
            return f"+{int(v)}" if v > 0 else str(int(v))
        def _g_sig(t): return t.get("signal", "Neutre")

        growth_cols = [
            ("F", _g_rg), ("G", _g_em), ("H", _g_nm),
            ("I", _g_roe), ("J", _g_eg), ("K", _g_rev), ("L", _g_sig),
        ]
        if "GROWTH" in wb.sheetnames:
            _fill_ranking_sheet(wb["GROWTH"], by_growth, "score_growth",
                                growth_cols, universe, "GROWTH")

        # ---- QUALITY ----
        def _q_az(t):   return _num(t.get("altman_z"), 2)
        def _q_sigz(t):
            az = t.get("altman_z")
            if az is None: return _NA
            return "Sain" if az > 2.9 else ("Zone grise" if az > 1.23 else "Detresse")
        def _q_bm(t):   return _num(t.get("beneish_m"), 2)
        def _q_sigm(t):
            bm = t.get("beneish_m")
            if bm is None: return _NA
            return "Risque" if bm > -1.78 else "Ok"
        def _q_em(t):   return _pct(t.get("ebitda_margins"))
        def _q_roe(t):  return _pct(t.get("roe"), sign=True)
        def _q_ic(t):   return _num(t.get("interest_coverage"), 1)
        def _q_cr(t):   return _num(t.get("current_ratio"), 2)

        quality_cols = [
            ("F", _q_az), ("G", _q_sigz), ("H", _q_bm), ("I", _q_sigm),
            ("J", _q_em), ("K", _q_roe),  ("L", _q_ic), ("M", _q_cr),
        ]
        if "QUALITY" in wb.sheetnames:
            _fill_ranking_sheet(wb["QUALITY"], by_quality, "score_quality",
                                quality_cols, universe, "QUALITY")

        # ---- MOMENTUM ----
        def _m_mom(t):  return _pct(t.get("mom_52w") / 100, sign=True) if t.get("mom_52w") is not None else _NA
        def _m_prix(t): return _price_str(t.get("price"))
        def _m_mkt(t):  return _num(t.get("mkt_cap"), 1)
        def _m_glb(t):  return t.get("score_global")
        def _m_earn(t): return t.get("next_earnings") or _NA
        def _m_sig(t):  return t.get("signal", "Neutre")

        momentum_cols = [
            ("F", _m_mom), ("G", _m_prix), ("H", _m_mkt),
            ("I", _m_glb), ("J", _m_earn), ("K", _m_sig),
        ]
        if "MOMENTUM" in wb.sheetnames:
            _fill_ranking_sheet(wb["MOMENTUM"], by_momentum, "score_momentum",
                                momentum_cols, universe, "MOMENTUM")

        # ---- PAR SECTEUR ----
        if "PAR SECTEUR" in wb.sheetnames:
            _fill_par_secteur(wb["PAR SECTEUR"], sector_agg, universe,
                              len(tickers), today_str)

        # ---- SECTOR OVERVIEW col M ----
        if "SECTOR OVERVIEW" in wb.sheetnames:
            _fill_sector_overview(wb["SECTOR OVERVIEW"], sector_agg)

        # ---- DASHBOARD ----
        if "DASHBOARD" in wb.sheetnames:
            _fill_dashboard(wb["DASHBOARD"], universe, today_str)

        # ---- Print areas (sans fitToPage : evite la compression illegible) ----
        # PAR SECTEUR : A1:K35 englobe SCORECARD + COMPOSITION DETAIL
        if "PAR SECTEUR" in wb.sheetnames:
            wb["PAR SECTEUR"].print_area = "A1:K35"
        # DASHBOARD : A1:N30 exclut les cols aux (O-X)
        if "DASHBOARD" in wb.sheetnames:
            wb["DASHBOARD"].print_area = "A1:N30"
        # SECTOR OVERVIEW : A1:L50 (formules + noms secteurs col M exclus)
        if "SECTOR OVERVIEW" in wb.sheetnames:
            wb["SECTOR OVERVIEW"].print_area = "A1:L50"

        # ---- Date stale : remplace toute date DD/MM/YYYY dans les 5 premieres lignes ----
        # Couvre B1 des feuilles sectorielles, headers SECTOR OVERVIEW, PAR SECTEUR, etc.
        # Remplace toute date stale DD/MM/YYYY dans les 5 premieres lignes (50 cols)
        # Couvre aussi les formules string ("=...date...") pour SECTOR OVERVIEW headers
        import re as _re
        _date_pat = _re.compile(r'\d{2}/\d{2}/\d{4}')
        for _sname in wb.sheetnames:
            _ws2 = wb[_sname]
            for _r in range(1, 6):
                for _c in range(1, 51):
                    _cell = _ws2.cell(row=_r, column=_c)
                    if (
                        _cell.value
                        and isinstance(_cell.value, str)
                        and _date_pat.search(_cell.value)
                    ):
                        _cell.value = _date_pat.sub(today_str, _cell.value)

        # Sauvegarder
        wb.save(str(out))
        log.info("[IndiceExcelWriter] Excel indice sauvegarde: %s (%d Ko)",
                 out.name, out.stat().st_size // 1024)
