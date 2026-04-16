# =============================================================================
# FinSight IA — Excel Writer v2
# outputs/excel_writer.py
#
# Injecte FinancialSnapshot dans TEMPLATE.xlsx (openpyxl).
# Respecte le contrat config/excel_mapping.py v2 :
#   - NE PAS écraser les cellules formule (is_formula_cell strict)
#   - Colonnes : D=le plus ancien, alignement gauche (H peut être vide si <5 ans)
#   - Aucune projection injectée (formules Excel dans feuille dédiée)
#   - Market data : colonne H uniquement (données point-in-time)
#   - Sheet principale : "INPUT"
#   - Sheet historique : "STOCK_DATA"
#
# Usage :
#   writer = ExcelWriter()
#   path   = writer.write(snapshot, synthesis, ratios)
# =============================================================================

from __future__ import annotations

import logging
import math
import os
import shutil
from datetime import date
from pathlib import Path
from typing import Optional

from config.excel_mapping import is_formula_cell, COMPANY_INFO as _CI_CELLS

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Chemins
# ---------------------------------------------------------------------------

_DEFAULT_TEMPLATE = Path(__file__).parent.parent / "assets" / "TEMPLATE.xlsx"
_TEMPLATE_ENV = os.getenv("TEMPLATE_PATH", "")
# Toujours résoudre en absolu : si relatif, ancrer sur la racine du projet
# NOTE : _TEMPLATE est le DEFAULT (STANDARD). Le vrai template utilisé est
# résolu dynamiquement par excel_profile_router.get_template_for(profile)
# dans la méthode write() — ça permet d'utiliser un template BANK/INSURANCE/
# REIT/UTILITY/OIL_GAS spécifique si disponible dans assets/, sinon fallback
# STANDARD transparent.
_TEMPLATE = (
    Path(_TEMPLATE_ENV).resolve()
    if _TEMPLATE_ENV
    else _DEFAULT_TEMPLATE
)
_OUTPUT_DIR = Path(__file__).parent / "generated"

# Mois anglais locale-indépendants pour STOCK_DATA
_MONTH_EN = ["Jan","Feb","Mar","Apr","May","Jun",
             "Jul","Aug","Sep","Oct","Nov","Dec"]

# ---------------------------------------------------------------------------
# Mapping année → colonne Excel (5 exercices, du plus ancien au plus récent)
# D=N-4, E=N-3, F=N-2, G=N-1, H=N
# ---------------------------------------------------------------------------

def _build_year_col(snapshot) -> dict:
    """
    Construit le mapping annee_label -> colonne Excel.
    Alignement a droite : H = annee la plus recente, G = N-1, F = N-2, etc.
    - 5 ans : D,E,F,G,H tous remplis
    - 4 ans : E,F,G,H remplis, D vide
    - 3 ans : F,G,H remplis, D,E vides
    Exclut les annees sans revenue pour eviter des colonnes creuses.
    Avec alignement droite, H est TOUJOURS la colonne LTM (la plus recente).
    """
    cols = ["D", "E", "F", "G", "H"]
    labels = sorted(snapshot.years.keys(), key=lambda y: int(y.split("_")[0]))
    labels = labels[-5:]  # max 5, les plus recents
    # Exclure les années sans Données suffisantes (revenue + au moins un agrégat IS)
    # pour éviter des colonnes creuses dans RATIOS/DCF qui afficheraient N/A.
    labels_with_rev = [
        l for l in labels
        if getattr(snapshot.years.get(l), "revenue", None) is not None
        and (
            getattr(snapshot.years.get(l), "gross_profit_yf", None) is not None
            or getattr(snapshot.years.get(l), "ebit_yf", None) is not None
            or getattr(snapshot.years.get(l), "net_income_yf", None) is not None
        )
    ]
    if labels_with_rev:
        labels = labels_with_rev[-5:]
    n = len(labels)
    # Alignement droite : plus récent -> H (index 4), precedent -> G (index 3), etc.
    mapping = {}
    for i, label in enumerate(reversed(labels)):
        col = cols[4 - i]  # 0->H, 1->G, 2->F, 3->E, 4->D
        mapping[label] = col
    return mapping

# ---------------------------------------------------------------------------
# Champs IS à injecter en négatif (coûts attendus <0 par les formules Excel)
# ---------------------------------------------------------------------------

_IS_COST_FIELDS = {"cogs", "sga", "rd", "da", "interest_expense",
                   "tax_expense_real", "dividends"}

# FinancialYear field → ligne Excel
_IS_ROWS = {
    "revenue":          9,
    "cogs":             10,
    "sga":              14,
    "rd":               15,
    "da":               16,
    "interest_expense": 20,
    "interest_income":  21,
    "tax_expense_real": 25,
    "dividends":        29,
}

_BS_ASSET_ROWS = {
    "cash":                 34,
    "accounts_receivable":  35,
    "inventories":          36,
    "other_current_assets": 37,
    "ppe_net":              41,
    "intangibles":          42,
    "other_lt_assets":      43,
}

_BS_LIAB_ROWS = {
    "accounts_payable":      47,
    "short_term_debt":       48,
    "income_tax_payable":    49,
    "other_current_liab":    50,
    "long_term_debt":        54,
    # D58 : Paid-In Capital = Capital Stock + APIC (calculé dans yfinance_source)
    "common_equity_paid_in": 58,
    # D59 : Retained Earnings (bilan BS)
    "retained_earnings_yf":  59,
}

_CF_ROWS = {
    "change_accounts_receivable": 71,
    "change_inventories":         72,
    "change_accounts_payable":    73,
    "other_wc_changes":           74,
    "capex":                      78,
    "other_investing":            79,
    "change_lt_debt":             83,
    "change_common_equity":       84,
    "dividends_paid":             85,
    "beginning_cash":             90,
}

# Market data — colonne H uniquement (données point-in-time, D-G = tirets dans le template)
# H108 WACC = FORMULE Excel — absent de la liste
_MKT_FIELDS = {
    "share_price":          "H95",
    "shares_diluted":       "H96",
    "beta_levered":         "H99",
    "risk_free_rate":       "H100",
    "erp":                  "H101",
    "cost_of_debt_pretax":  "H103",
    "tax_rate":             "H104",
    "weight_equity":        "H106",
    "weight_debt":          "H107",
    # H108 WACC = FORMULE — absent de la liste
    "terminal_growth":      "H109",
    "days_in_period":       "H110",
}


# ---------------------------------------------------------------------------
# ExcelWriter
# ---------------------------------------------------------------------------

class ExcelWriter:
    """
    Injecte un FinancialSnapshot dans TEMPLATE.xlsx.
    Conserve toutes les formules Excel intactes (vérification stricte).
    """

    def write(
        self,
        snapshot,
        synthesis=None,
        ratios=None,
        output_path: Optional[Path] = None,
        comparables=None,   # list[PeerData] ou None
    ) -> Path:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl requis : pip install openpyxl")

        # ── ROUTAGE PROFIL SECTORIEL (nouveau #180) ──────────────────────
        # Détecte le profil (STANDARD/BANK/INSURANCE/REIT/UTILITY/OIL_GAS)
        # et sélectionne le template approprié. Fallback transparent vers
        # STANDARD si le template profil-spécifique n'existe pas dans assets/.
        #
        # Zéro régression : tant qu'aucun template autre que TEMPLATE.xlsx
        # n'est présent, toutes les analyses utilisent STANDARD (comportement
        # actuel inchangé).
        _active_template = _TEMPLATE
        _profile = "STANDARD"
        _cell_map: dict = {}  # rempli par le router selon _profile
        try:
            from outputs.excel_profile_router import (
                detect_profile_from_snapshot, get_template_for,
                profile_template_exists, get_cell_map,
            )
            _profile = detect_profile_from_snapshot(snapshot)
            if profile_template_exists(_profile):
                _active_template = get_template_for(_profile)
                log.info(
                    f"[ExcelWriter] Profil sectoriel : {_profile} → "
                    f"template spécifique {_active_template.name}"
                )
            else:
                if _profile != "STANDARD":
                    log.info(
                        f"[ExcelWriter] Profil sectoriel : {_profile} "
                        f"(template spécifique absent → fallback STANDARD)"
                    )
            _cell_map = get_cell_map(_profile)
        except Exception as _e_router:
            log.warning(
                f"[ExcelWriter] Profile router failed ({_e_router}) — "
                f"fallback TEMPLATE.xlsx"
            )
            _active_template = _TEMPLATE
            _profile = "STANDARD"
            _cell_map = {}

        # Sélectionne les cell_maps actifs — fallback sur les constantes module
        # (STANDARD) si le router n'a rien retourné. Ainsi le comportement reste
        # identique pour toutes les analyses STANDARD, et OIL_GAS/etc. utilisent
        # leur propre layout dès qu'un template spécifique existe.
        _is_rows_active        = _cell_map.get("is_rows")        or _IS_ROWS
        _is_cost_fields_active = _cell_map.get("is_cost_fields") or _IS_COST_FIELDS
        _bs_asset_rows_active  = _cell_map.get("bs_asset_rows")  or _BS_ASSET_ROWS
        _bs_liab_rows_active   = _cell_map.get("bs_liab_rows")   or _BS_LIAB_ROWS
        _cf_rows_active        = _cell_map.get("cf_rows")        or _CF_ROWS
        _mkt_fields_active     = _cell_map.get("mkt_fields")     or _MKT_FIELDS
        _tax_rate_row_active   = _cell_map.get("tax_rate_row", 104)

        if not _active_template.exists():
            raise FileNotFoundError(f"Template introuvable : {_active_template}")

        ticker = snapshot.ticker.replace(".", "_")
        today  = date.today().isoformat()

        if output_path is None:
            _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_path = _OUTPUT_DIR / f"{ticker}_{today}.xlsx"

        shutil.copy2(_active_template, output_path)
        wb = openpyxl.load_workbook(
            str(output_path),
            keep_links=False,
            keep_vba=False,
            data_only=False,
        )
        ws = wb["INPUT"]

        # ------------------------------------------------------------------
        # 1. Company info (cellules fixes colonne D)
        # ------------------------------------------------------------------
        # ------------------------------------------------------------------
        # 2. Données financières — 5 exercices historiques (H=LTM → D)
        #    Alignement droite : H = LTM, G = LTM-1, ...
        #    Colonnes inactives (ex: D,E si 3 ans) : rien écrit — None implicite
        # ------------------------------------------------------------------
        year_col_map = _build_year_col(snapshot)
        written = 0

        ci = snapshot.company_info

        # D117 Base Year = toujours l'année LTM (alignement droite -> H = toujours la plus récente)
        # Avec right-align, H est toujours le label le plus récent dans year_col_map
        most_recent_lbl = max(year_col_map.keys(), key=lambda y: int(y.split("_")[0])) if year_col_map else None
        ltm_year = int(most_recent_lbl.split("_")[0]) if most_recent_lbl else ci.base_year

        # Ecrire les etiquettes d'année directement dans D5-H5 et dans D132-H132.
        # D5-H5 : remplace les formules du template par des valeurs statiques
        #   (ex: "2022", "2023", "2025 (LTM)") — alignement droite garanti.
        # D132-H132 : cellules helper hors zone impression (backup / formules tier).
        # Note : on ecrit directement sur ws[] sans passer par _safe_write
        #   car les cellules D5-H5 contiennent des formules template (startswith "=")
        #   que la garde dynamique bloquerait. Ces formules sont intentionnellement
        #   remplacees par des valeurs statiques calculées côté Python.
        # Construire d'abord le dict col -> valeur année
        _year_vals: dict[str, object] = {}
        for label, col in year_col_map.items():
            year_int = int(label.split("_")[0])
            is_ltm = (col == 'H')
            _year_vals[col] = f"{year_int} (LTM)" if is_ltm else year_int
        # Ecrire D5-H5 : "" pour les colonnes inactives (évite le 0
        # que retournerait =INPUT!Dx dans RATIOS/DCF/SCÉNARIOS quand la cellule est None)
        for col in ['D', 'E', 'F', 'G', 'H']:
            ws[f'{col}5']   = _year_vals.get(col, "")
        # D132-H132 helper row — REIT n'en a pas (row 132 = days_in_period)
        _year_helper_row = _cell_map.get("year_helper_row", 132)
        if _year_helper_row is not None:
            for col in ['D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{_year_helper_row}'] = _year_vals.get(col, None)

        # Company info — cellules fixes colonne D
        # REIT/profils non-STANDARD peuvent avoir des positions différentes
        _ci_map_override = _cell_map.get("company_info")
        _ci_map_active = _ci_map_override if _ci_map_override else {
            "company_name":  _CI_CELLS["company_name"],
            "ticker":        _CI_CELLS["ticker"],
            "sector":        _CI_CELLS["sector"],
            "base_year":     _CI_CELLS["base_year"],
            "currency":      _CI_CELLS["currency"],
            "units":         _CI_CELLS["units"],
            "analysis_date": _CI_CELLS["analysis_date"],
        }
        _write_cells(ws, {
            _ci_map_active["company_name"]:  ci.company_name,
            _ci_map_active["ticker"]:        ci.ticker,
            _ci_map_active["sector"]:        ci.sector or "",
            _ci_map_active["base_year"]:     ltm_year,
            _ci_map_active["currency"]:      ci.currency,
            _ci_map_active["units"]:         ci.units,
            _ci_map_active["analysis_date"]: ci.analysis_date or today,
        })

        for year_str, fy in snapshot.years.items():
            col = year_col_map.get(year_str)
            if col is None:
                log.warning(f"[ExcelWriter] Année '{year_str}' sans colonne — ignore")
                continue

            for field, row in _is_rows_active.items():
                val = getattr(fy, field, None)
                # Fallback dividends IS : si absent, utiliser dividends_paid du CF
                # (sociétés européennes ne reportent pas toujours les dividendes en IS)
                if val is None and field == "dividends":
                    dp = getattr(fy, "dividends_paid", None)
                    if dp is not None:
                        val = abs(dp)  # positif avant negation par _IS_COST_FIELDS
                if val is not None and field in _is_cost_fields_active:
                    val = -abs(val)
                # Fallback 0 pour interest_expense/income : certaines sociétés
                # (ex: Apple FY2024+) ne les declarent plus separement.
                # Le template F128 requiert ces cellules non-vides pour
                # afficher les headers d'années — 0 = "non declare separement".
                if val is None and field in ("interest_expense", "interest_income"):
                    if getattr(fy, "revenue", None) is not None:
                        val = 0
                written += _safe_write(ws, col, row, val)

            for field, row in _bs_asset_rows_active.items():
                written += _safe_write(ws, col, row, getattr(fy, field, None))

            for field, row in _bs_liab_rows_active.items():
                written += _safe_write(ws, col, row, getattr(fy, field, None))

            for field, row in _cf_rows_active.items():
                written += _safe_write(ws, col, row, getattr(fy, field, None))

            # Tax rate effectif par année -> row tax_rate_row (col E-H selon année)
            # EBT = net_income_yf + |tax_expense_real| (approximation IS)
            _tax_abs = abs(fy.tax_expense_real or 0.0)
            _ni      = fy.net_income_yf or 0.0
            _ebt_yr  = _ni + _tax_abs
            if _ebt_yr > 0 and _tax_abs > 0:
                _tr_yr = round(min(0.40, max(0.05, _tax_abs / _ebt_yr)), 4)
                _safe_write(ws, col, _tax_rate_row_active, _tr_yr)
                written += 1

        # ------------------------------------------------------------------
        # 2b. Mise à jour formules E24:H24 : utiliser taux effectif (row 104)
        #     au lieu du taux fixe 21% — bypass safe_write (formule->formule)
        # ------------------------------------------------------------------
        for _c in ["E", "F", "G", "H"]:
            _cell = ws[f"{_c}24"]
            if _cell.value and isinstance(_cell.value, str) and "0.21" in _cell.value:
                _cell.value = f"=-ROUND({_c}22*{_c}104,0)"
                log.debug(f"[ExcelWriter] UPDATE {_c}24 -> =-ROUND({_c}22*{_c}104,0)")

        # ------------------------------------------------------------------
        # 3. Market data — colonne H uniquement (données point-in-time)
        # ------------------------------------------------------------------
        mkt = snapshot.market
        for field, cell_ref in _mkt_fields_active.items():
            val = getattr(mkt, field, None)
            if val is None:
                continue
            # Filtre NaN/inf
            if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                log.warning(f"[ExcelWriter] SKIP {cell_ref} ({field}) — valeur NaN/inf ignoree")
                continue
            # Garde 1 statique
            if is_formula_cell(cell_ref):
                log.error(f"[ExcelWriter] BLOQUE (statique) {cell_ref} — cellule formule market")
                continue
            # Garde 2 dynamique
            current = ws[cell_ref].value
            if isinstance(current, str) and current.startswith("="):
                log.error(f"[ExcelWriter] BLOQUE (dynamique) {cell_ref} — formule non declaree : {current[:60]}")
                continue
            log.debug(f"[ExcelWriter] WRITE {cell_ref} = {val!r}")
            ws[cell_ref] = val
            written += 1

        # ------------------------------------------------------------------
        # 4. Historique cours — onglet STOCK_DATA
        #    Toujours ecrire B3:C15 (13 lignes) pour effacer residus template
        # ------------------------------------------------------------------
        if "STOCK_DATA" in wb.sheetnames:
            ws_stock = wb["STOCK_DATA"]
            history = snapshot.stock_history or []
            for i in range(13):
                row = 3 + i
                if i < len(history):
                    pt = history[i]
                    # Format MMM-YY locale-independant (ex: "Mar-26")
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(pt.month, "%b-%y")
                        month_str = f"{_MONTH_EN[dt.month - 1]}-{dt.strftime('%y')}"
                    except Exception:
                        month_str = pt.month  # fallback : valeur brute
                    ws_stock[f"B{row}"] = month_str
                    ws_stock[f"C{row}"] = pt.price
                else:
                    ws_stock[f"B{row}"] = None
                    ws_stock[f"C{row}"] = None

        # ------------------------------------------------------------------
        # 5. Comparables — onglet COMPARABLES (peers lignes 9-13)
        # ------------------------------------------------------------------
        if comparables and "COMPARABLES" in wb.sheetnames:
            written += _write_comparables(wb["COMPARABLES"], comparables)

        # ------------------------------------------------------------------
        # 6. Modèle LBO — feuilles LBO MODEL (visible) + _LBO_CALC (masquée)
        # ------------------------------------------------------------------
        try:
            _lbo_ref = Path(__file__).parent.parent / "assets" / "TSLA_LBO_REFERENCE.xlsx"
            if _lbo_ref.exists():
                _copy_lbo_from_template(wb, str(_lbo_ref))
                # Repositionner LBO MODEL juste après DCF (référence Baptiste)
                _sheet_names = wb.sheetnames
                _dcf_idx = _sheet_names.index("DCF") if "DCF" in _sheet_names else None
                if _dcf_idx is not None:
                    for _lbo_name in ("LBO MODEL", "_LBO_CALC"):
                        if _lbo_name in _sheet_names:
                            _cur_idx = wb.sheetnames.index(_lbo_name)
                            _target = _dcf_idx + 1
                            wb.move_sheet(_lbo_name, offset=_target - _cur_idx)
                log.info("[ExcelWriter] LBO sheets copiées depuis template référence (après DCF)")
            else:
                log.warning("[ExcelWriter] Template LBO introuvable: %s", _lbo_ref)
        except Exception as _lbo_ex:
            log.warning(f"[ExcelWriter] LBO sheets failed: {_lbo_ex}")

        wb.calculation.fullCalcOnLoad = True
        wb.save(str(output_path))
        log.info(f"[ExcelWriter] '{ticker}' — {written} cellules ecrites → {output_path.name}")
        return output_path


def _copy_lbo_from_template(wb_dst, template_path: str):
    """Copie les feuilles LBO MODEL et _LBO_CALC depuis le template de référence.
    Copie cellule par cellule : valeurs, formules, styles, merged cells, dimensions."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from copy import copy

    wb_src = load_workbook(template_path)

    for sheet_name in ("LBO MODEL", "_LBO_CALC"):
        if sheet_name not in wb_src.sheetnames:
            continue
        # Supprimer si existe déjà
        if sheet_name in wb_dst.sheetnames:
            del wb_dst[sheet_name]

        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst.create_sheet(sheet_name)

        # Copier les dimensions de colonnes
        for col_letter, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col_letter].width = dim.width
            ws_dst.column_dimensions[col_letter].hidden = dim.hidden

        # Copier les dimensions de lignes
        for row_num, dim in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[row_num].height = dim.height
            ws_dst.row_dimensions[row_num].hidden = dim.hidden

        # Copier cellule par cellule (valeur/formule + style)
        for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row,
                                     min_col=1, max_col=ws_src.max_column):
            for cell in row:
                dst_cell = ws_dst.cell(row=cell.row, column=cell.column)
                dst_cell.value = cell.value  # préserve les formules (str commençant par =)
                if cell.has_style:
                    dst_cell.font = copy(cell.font)
                    dst_cell.fill = copy(cell.fill)
                    dst_cell.border = copy(cell.border)
                    dst_cell.alignment = copy(cell.alignment)
                    dst_cell.number_format = cell.number_format
                    dst_cell.protection = copy(cell.protection)

        # Copier les merged cells
        for merged_range in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merged_range))

        # Masquer _LBO_CALC
        if sheet_name == "_LBO_CALC":
            ws_dst.sheet_state = "hidden"

    wb_src.close()
    log.info("[ExcelWriter] LBO sheets copiées depuis %s", template_path)


# ---------------------------------------------------------------------------
# Helpers internes
# ---------------------------------------------------------------------------

def _safe_write(ws, col: str, row: int, value) -> int:
    """
    Ecrit une valeur avec double-garde anti-formule :
      1. Garde statique  : FORMULA_CELLS du mapping (O(1))
      2. Garde dynamique : inspecte la valeur courante de la cellule dans le template
    Filtre NaN/inf pour eviter la corruption XML Excel.
    Log audit systematique avant chaque ecriture effective.
    """
    if value is None:
        return 0
    # Filtre NaN et inf : openpyxl ecrit "nan" ou "inf" comme string
    # ce qui Généré des erreurs de formule dans sheet1.xml
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        log.warning(f"[ExcelWriter] SKIP {col}{row} — valeur NaN/inf ignoree")
        return 0
    cell_ref = f"{col}{row}"

    # Garde 1 — liste statique
    if is_formula_cell(cell_ref):
        log.error(f"[ExcelWriter] BLOQUE (statique) {cell_ref} — cellule formule")
        return 0

    # Garde 2 — valeur dynamique dans le classeur chargé
    current = ws[cell_ref].value
    if isinstance(current, str) and current.startswith("="):
        log.error(f"[ExcelWriter] BLOQUE (dynamique) {cell_ref} — formule non declaree : {current[:60]}")
        return 0

    # Log audit
    log.debug(f"[ExcelWriter] WRITE {cell_ref} = {value!r}")
    ws[cell_ref] = value
    return 1


def _write_cells(ws, mapping: dict) -> None:
    """Ecrit un dict {cell_ref: value} — COMPANY_INFO uniquement.
    Applique garde statique + garde dynamique formule."""
    for cell_ref, value in mapping.items():
        if value is None:
            continue
        # Filtre NaN float
        if isinstance(value, float) and math.isnan(value):
            log.warning(f"[ExcelWriter] SKIP {cell_ref} — valeur NaN ignoree")
            continue
        # Garde 1 statique
        if is_formula_cell(cell_ref):
            log.error(f"[ExcelWriter] BLOQUE (statique) {cell_ref} — cellule formule COMPANY_INFO")
            continue
        # Garde 2 dynamique
        current = ws[cell_ref].value
        if isinstance(current, str) and current.startswith("="):
            log.error(f"[ExcelWriter] BLOQUE (dynamique) {cell_ref} — formule non declaree : {current[:60]}")
            continue
        log.debug(f"[ExcelWriter] WRITE {cell_ref} = {value!r}")
        ws[cell_ref] = value


def _write_comparables(ws_comp, peers: list) -> int:
    """
    Injecte les 5 peers dans l'onglet COMPARABLES (lignes 9-13).
    Colonnes injectees : B (nom), C (ticker), D (price), E (EV $M),
                         F (Revenue $M), G (EBITDA $M), J (EBITDA growth), K (EPS)
    Colonnes formules intouchables : H (EV/EBITDA), I (EV/Rev), L (P/E)
    Lignes 15-18 (Mean/Med/Max/Min) et 20+ (Target + Implied) : intouchables.
    """
    # Cellules formule dans COMPARABLES — ne jamais ecraser
    _COMP_FORMULA_COLS = {"H", "I", "L"}   # par ligne peer
    _COMP_FORMULA_ROWS = set(range(15, 40)) # Mean/Med/Max/Min + Target + Implied

    written = 0
    for i, peer in enumerate(peers[:5]):
        row = 9 + i

        # Garde : ne jamais ecrire dans les lignes formule
        if row in _COMP_FORMULA_ROWS:
            continue

        def _w(col: str, val) -> None:
            nonlocal written
            if val is None:
                return
            if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                return
            if col in _COMP_FORMULA_COLS:
                log.error(f"[ExcelWriter] COMPARABLES : tentative col formule {col}{row}")
                return
            current = ws_comp[f"{col}{row}"].value
            if isinstance(current, str) and current.startswith("="):
                log.error(f"[ExcelWriter] COMPARABLES : formule non declaree {col}{row}")
                return
            log.debug(f"[ExcelWriter] COMPARABLES {col}{row} = {val!r}")
            ws_comp[f"{col}{row}"] = val
            written += 1

        _w("B", getattr(peer, "name", None) or peer.ticker)
        _w("C", peer.ticker)
        _w("D", getattr(peer, "share_price", None))
        _w("E", getattr(peer, "ev", None))
        _w("F", getattr(peer, "revenue_ltm", None))
        _w("G", getattr(peer, "ebitda_ltm", None))
        _w("J", getattr(peer, "ebitda_growth_ntm", None))
        _w("K", getattr(peer, "eps_ltm", None))

    log.info(f"[ExcelWriter] COMPARABLES — {written} cellules ecrites ({len(peers[:5])} peers)")
    return written


def _safe_write_cell(ws, cell_ref: str, value) -> None:
    """Ecriture securisee d'une cellule avec garde formule (statique + dynamique)."""
    if value is None:
        return
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        log.warning(f"[ExcelWriter] SKIP {cell_ref} — valeur NaN/inf ignoree")
        return
    # Garde statique
    if is_formula_cell(cell_ref):
        log.error(f"[ExcelWriter] BLOQUE (statique) {cell_ref} — cellule formule synthesis")
        return
    # Garde dynamique
    current = ws[cell_ref].value
    if isinstance(current, str) and current.startswith("="):
        log.error(f"[ExcelWriter] BLOQUE (dynamique) {cell_ref} — formule : {current[:60]}")
        return
    ws[cell_ref] = value


