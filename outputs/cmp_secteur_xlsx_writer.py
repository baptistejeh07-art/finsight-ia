"""
cmp_secteur_xlsx_writer.py — FinSight IA
Writer XLSX comparatif sectoriel (secteur A vs secteur B).

Charge le template assets/CMP_SECTEUR_TEMPLATE.xlsx (designe par Baptiste,
6 feuilles : SYNTHÈSE, AGRÉGATS, TOP_TECH, TOP_HC, DATA_TECH, DATA_HC) puis :
1. Remplit DATA_TECH avec les tickers du secteur A
2. Remplit DATA_HC avec les tickers du secteur B
3. Remplace les labels hardcodés "TECHNOLOGIE"/"HEALTHCARE"/"Tech"/"HC"
   par les noms de secteurs réels dans toutes les feuilles formule-driven
4. Étend les ranges de formules DATA_HC!$A$4:$A$9 → DATA_HC!$A$4:$A$200
   pour que le template supporte n'importe quel nombre de tickers (pas
   juste les 6 de l'exemple Baptiste)
5. Met à jour le titre "COMPARATIF SECTORIEL — {A} vs {B}" et les
   compteurs "— N valeurs"

Toutes les autres feuilles (SYNTHÈSE, AGRÉGATS, TOP_TECH, TOP_HC) sont
formule-driven (MEDIAN/INDEX/LARGE/AVERAGE/RANK.EQ/QUARTILE...) et se
mettent à jour automatiquement à l'ouverture dans Excel (Ctrl+Alt+F9 si
besoin).

Layout DATA_TECH / DATA_HC :
    A3:H3 = headers (TICKER, SOCIÉTÉ, SCORE FINSIGHT, P/E, EV/EBITDA,
            MARGE EBITDA, ROE, CROISS. REV.)
    DATA_TECH I3 = RANG_UNIQUE (col supplémentaire pour formule rank)
    A4..  = données ticker par ligne
    C1    = titre "DATA TECH/HC — Données brutes"
    J1/I1 = nom du secteur (référence croisée, voir note ci-dessous)

Note sur les sheet names : DATA_TECH et DATA_HC sont des noms FIXES, pas
des indicateurs du contenu. Le writer met les tickers du sector_a dans
DATA_TECH et ceux du sector_b dans DATA_HC quoi qu'il arrive.

Usage :
    from outputs.cmp_secteur_xlsx_writer import generate_cmp_secteur_xlsx
    xlsx_bytes = generate_cmp_secteur_xlsx(
        tickers_a, sector_a, universe_a,
        tickers_b, sector_b, universe_b,
    )
"""
from __future__ import annotations

import io
import logging
import re
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# Labels — replacement maps
# ═══════════════════════════════════════════════════════════════════════════

# Les labels hardcodés dans le template de référence Baptiste. Ils sont
# remplacés au moment de la génération par les noms de secteurs réels.
# La table est ordonnée : les patterns les PLUS LONGS en premier pour éviter
# que "Tech" soit remplacé avant "Tech vs HC" (sinon on casse l'ordre).
_LABEL_REPLACEMENTS_ORDER = [
    # 1. Patterns composés (doivent être traités AVANT les mots seuls)
    "Tech vs HC",
    "TECH Médiane", "TECH Moyenne", "TECH Min", "TECH Max",
    "HC Médiane", "HC Moyenne", "HC Min", "HC Max",
    '"✓ Tech"', '"✓ HC"',
    '"Prime Tech"', '"Prime HC"',
    '"Tech"', '"HC"',
    "Technologie (pondérée)",
    "Healthcare (pondérée)",
    # 2. Mots pleins
    "TECHNOLOGIE",
    "HEALTHCARE",
    "Technologie",
    "Healthcare",
]


def _build_replacements(sector_a: str, sector_b: str) -> dict:
    """Construit la table de remplacement labels → labels adaptés au cas.
    L'ordre du dict est préservé (Python 3.7+).

    Utilise le nom complet du secteur partout (pas d'abréviation) pour
    éviter les troncatures moches type 'Techn' / 'Sant'. Excel ajuste
    automatiquement la largeur des colonnes."""
    a_upper = sector_a.upper()
    b_upper = sector_b.upper()
    a_cap   = sector_a
    b_cap   = sector_b

    return {
        # Composés (priorité haute — doivent matcher avant les mots seuls)
        "Tech vs HC":        f"{a_cap} vs {b_cap}",
        "TECH Médiane":      f"{a_upper} Médiane",
        "TECH Moyenne":      f"{a_upper} Moyenne",
        "TECH Min":          f"{a_upper} Min",
        "TECH Max":          f"{a_upper} Max",
        "HC Médiane":        f"{b_upper} Médiane",
        "HC Moyenne":        f"{b_upper} Moyenne",
        "HC Min":            f"{b_upper} Min",
        "HC Max":            f"{b_upper} Max",
        # Formules IF(...) contenant des chaînes quotées
        '"✓ Tech"':          f'"✓ {a_cap}"',
        '"✓ HC"':            f'"✓ {b_cap}"',
        '"Prime Tech"':      f'"Prime {a_cap}"',
        '"Prime HC"':        f'"Prime {b_cap}"',
        '"Tech"':            f'"{a_cap}"',
        '"HC"':              f'"{b_cap}"',
        "Technologie (pondérée)": f"{a_cap} (pondérée)",
        "Healthcare (pondérée)":  f"{b_cap} (pondérée)",
        # Mots pleins
        "TECHNOLOGIE":       a_upper,
        "HEALTHCARE":        b_upper,
        "Technologie":       a_cap,
        "Healthcare":        b_cap,
    }


# ═══════════════════════════════════════════════════════════════════════════
# Helpers de conversion (pareil que le tab DATA_TECH)
# ═══════════════════════════════════════════════════════════════════════════

def _safe_float(v) -> Optional[float]:
    """Convertit en float, retourne None si invalide ou NaN."""
    if v is None:
        return None
    try:
        fv = float(v)
        if fv != fv:  # NaN
            return None
        return fv
    except Exception:
        return None


def _to_pct_100(v) -> Optional[float]:
    """Convertit un ratio (0.27 ou 27) en pourcentage 27.0 (multiplicateur 100
    si valeur absolue <= 2). Retourne None si invalide."""
    fv = _safe_float(v)
    if fv is None:
        return None
    return round(fv * 100, 2) if abs(fv) <= 2.0 else round(fv, 2)


def _to_mult(v) -> Optional[float]:
    """Convertit un multiple (P/E, EV/EBITDA) en float arrondi a 2 decimales.
    Filtre les valeurs aberrantes (> 999 ou < 0)."""
    fv = _safe_float(v)
    if fv is None or fv > 999 or fv <= 0:
        return None
    return round(fv, 2)


def _ticker_row_values(ticker: dict) -> list:
    """Construit la ligne A-H pour un ticker (sans col I = formule rank).

    Conversion strict aux unites attendues par le template :
        A: ticker (str)
        B: nom de societe (str, max 60 chars)
        C: score 0-100 (int)
        D: P/E (float, multiple brut, vide si negatif)
        E: EV/EBITDA (float, multiple brut, vide si negatif)
        F: Marge EBITDA en % (ex: 27.5 = 27,5%)
        G: ROE en %
        H: Croissance Revenus en %
    """
    return [
        ticker.get("ticker", "") or "",
        (ticker.get("company") or ticker.get("ticker") or "")[:60],
        int(round(_safe_float(ticker.get("score_global")) or 0)),
        _to_mult(ticker.get("pe_ratio") or ticker.get("pe")),
        _to_mult(ticker.get("ev_ebitda")),
        _to_pct_100(ticker.get("ebitda_margin")),
        _to_pct_100(ticker.get("roe")),
        _to_pct_100(ticker.get("revenue_growth")),
    ]


# ═══════════════════════════════════════════════════════════════════════════
# Injection
# ═══════════════════════════════════════════════════════════════════════════

def _clear_data_sheet(ws, max_clear_row: int, max_col: int) -> None:
    """Efface les Données existantes (rows 4..max_clear_row, cols 1..max_col).
    Preserve les headers (row 3) et la cellule J1 ou I1 (nom secteur)."""
    for r in range(4, max_clear_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None


def _rewrite_cells(ws, replacements: dict, range_patches: list[tuple[str, str]]) -> None:
    """Parcourt toutes les cellules d'une feuille et applique les
    remplacements (labels + ranges de formules).

    Args:
        ws            : worksheet openpyxl
        replacements  : dict {pattern: remplacement} appliqué aux VALEURS et aux
                        chaînes à l'intérieur des formules
        range_patches : liste de (old_ref, new_ref) pour étendre les ranges
                        type [('DATA_HC!$A$4:$A$9', 'DATA_HC!$A$4:$A$200'), ...]
    """
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            new_v = v

            # 1. Range patches (formules seulement)
            if new_v.startswith("="):
                for old, new in range_patches:
                    if old in new_v:
                        new_v = new_v.replace(old, new)

            # 2. Label replacements (valeurs + chaînes dans formules)
            for pattern in _LABEL_REPLACEMENTS_ORDER:
                if pattern in new_v:
                    new_v = new_v.replace(pattern, replacements[pattern])

            if new_v != v:
                cell.value = new_v


def _build_range_patches() -> list[tuple[str, str]]:
    """Liste des patches à appliquer aux formules pour étendre les ranges :
    - DATA_HC!*!4:*!9 (6 rows) → !4:!200 (197 rows)
    - DATA_TECH!*!4:*!91 (88 rows) → !4:!200 (197 rows)

    Format explicite pour toutes les colonnes A-J (au cas où le template
    en utiliserait certaines que je n'ai pas vues dans le dump)."""
    patches = []
    for col in "ABCDEFGHIJ":
        # DATA_HC : cap actuel $9 → $200
        patches.append((f"DATA_HC!${col}$4:${col}$9",   f"DATA_HC!${col}$4:${col}$200"))
        patches.append((f"DATA_HC!{col}4:{col}9",       f"DATA_HC!{col}4:{col}200"))
        # DATA_TECH : cap actuel $91 → $200 (pour supporter > 88 tickers)
        patches.append((f"DATA_TECH!${col}$4:${col}$91", f"DATA_TECH!${col}$4:${col}$200"))
        patches.append((f"DATA_TECH!{col}4:{col}91",     f"DATA_TECH!{col}4:{col}200"))
    return patches


def _inject_data_tech(ws, tickers: list, sector_name: str) -> None:
    """Injecte les Données du secteur 1 dans DATA_TECH.
    - rows 4..(4+N-1) cols A-H : valeurs ticker
    - col I : formule RANK.EQ + COUNTIF (anti-egalite)
    - J1 : nom du secteur
    """
    # Titre cellule J1
    ws["J1"] = sector_name

    # Clear ancien contenu (jusqu'a la row 91 = max template, 9 cols)
    _clear_data_sheet(ws, max_clear_row=91, max_col=9)

    # Injection
    sorted_tk = sorted(
        tickers,
        key=lambda t: _safe_float(t.get("score_global")) or 0,
        reverse=True,
    )
    for i, tk in enumerate(sorted_tk[:88]):  # cap a 88 lignes (template)
        r = 4 + i
        vals = _ticker_row_values(tk)
        for ci, v in enumerate(vals, start=1):
            ws.cell(row=r, column=ci, value=v)
        # Col I (9) : formule RANK.EQ pour le scoring unique
        # =RANK.EQ(C{r},$C$4:$C$9999,0)+COUNTIF($C$5:C{r},C{r})-1
        ws.cell(
            row=r,
            column=9,
            value=f"=RANK.EQ(C{r},$C$4:$C$9999,0)+COUNTIF($C$5:C{r},C{r})-1",
        )

    log.info("[cmp_secteur_xlsx] DATA_TECH injectee : %d lignes (%s)",
             len(sorted_tk[:88]), sector_name)


def _inject_data_hc(ws, tickers: list, sector_name: str) -> None:
    """Injecte les Données du secteur 2 dans DATA_HC.
    - rows 4..(4+N-1) cols A-H : valeurs ticker (pas de col I)
    - I1 : nom du secteur
    """
    ws["I1"] = sector_name

    # Clear (jusqu'a row 91 = sécurité, 8 cols)
    _clear_data_sheet(ws, max_clear_row=91, max_col=8)

    sorted_tk = sorted(
        tickers,
        key=lambda t: _safe_float(t.get("score_global")) or 0,
        reverse=True,
    )
    for i, tk in enumerate(sorted_tk[:88]):
        r = 4 + i
        vals = _ticker_row_values(tk)
        for ci, v in enumerate(vals, start=1):
            ws.cell(row=r, column=ci, value=v)

    log.info("[cmp_secteur_xlsx] DATA_HC injectee : %d lignes (%s)",
             len(sorted_tk[:88]), sector_name)


# ═══════════════════════════════════════════════════════════════════════════
# API publique
# ═══════════════════════════════════════════════════════════════════════════

def generate_cmp_secteur_xlsx(
    tickers_a: list,
    sector_a: str,
    universe_a: str,
    tickers_b: list,
    sector_b: str,
    universe_b: str,
    language: str = "fr",
    currency: str = "EUR",
) -> bytes:
    """
    Genere un fichier XLSX comparatif sectoriel en memoire.

    Charge assets/CMP_SECTEUR_TEMPLATE.xlsx puis injecte uniquement
    DATA_TECH et DATA_HC. Toutes les autres feuilles (SYNTHÈSE, AGRÉGATS,
    TOP_TECH, TOP_HC) sont formule-driven et se mettent a jour
    automatiquement a l'ouverture dans Excel (Ctrl+Alt+F9 si nécessaire).

    Args:
        tickers_a    : liste de dicts ticker du secteur 1 (Tech/etc)
        sector_a     : nom de secteur 1 (FR ou EN, ira en J1 de DATA_TECH)
        universe_a   : univers (informatif, non utilise par le template)
        tickers_b    : liste de dicts ticker du secteur 2 (HC/etc)
        sector_b     : nom de secteur 2 (FR ou EN, ira en I1 de DATA_HC)
        universe_b   : univers (informatif)

    Returns:
        bytes du fichier xlsx genere.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as _ie:
        raise RuntimeError(f"openpyxl manquant : {_ie}")

    # Résolution du template — cherche dans assets/ relatif au repo root
    _root = Path(__file__).resolve().parent.parent
    tpl_path = _root / "assets" / "CMP_SECTEUR_TEMPLATE.xlsx"

    if not tpl_path.exists():
        log.warning(
            "[cmp_secteur_xlsx] Template introuvable a %s — fallback minimal",
            tpl_path,
        )
        return _generate_fallback_xlsx(
            tickers_a, sector_a, tickers_b, sector_b
        )

    log.info("[cmp_secteur_xlsx] Chargement template : %s", tpl_path.name)
    wb = load_workbook(str(tpl_path), keep_links=False, data_only=False)

    # Vérification structure attendue
    _expected_sheets = {"SYNTHÈSE", "AGRÉGATS", "TOP_TECH", "TOP_HC",
                        "DATA_TECH", "DATA_HC"}
    _missing = _expected_sheets - set(wb.sheetnames)
    if _missing:
        log.error(
            "[cmp_secteur_xlsx] Template invalide — sheets manquantes : %s "
            "(trouvé : %s)",
            _missing, wb.sheetnames,
        )
        return _generate_fallback_xlsx(
            tickers_a, sector_a, tickers_b, sector_b
        )

    # ── Step 1 : Labels dynamiques + extension ranges ──────────────────
    # Parcours de toutes les feuilles formule-driven pour (a) remplacer
    # "TECHNOLOGIE"/"HEALTHCARE"/... par les noms de secteurs réels et
    # (b) étendre DATA_HC!$A$4:$A$9 → $A$4:$A$200 pour supporter > 6 tickers.
    replacements = _build_replacements(sector_a, sector_b)
    range_patches = _build_range_patches()
    for sheet_name in ("SYNTHÈSE", "AGRÉGATS", "TOP_TECH", "TOP_HC",
                       "DATA_TECH", "DATA_HC"):
        _rewrite_cells(wb[sheet_name], replacements, range_patches)

    # ── Step 2 : Titres dynamiques spécifiques ──────────────────────────
    _ws_syn = wb["SYNTHÈSE"]
    _ws_agg = wb["AGRÉGATS"]

    # SYNTHÈSE C1 : "COMPARATIF SECTORIEL — {A} vs {B}"
    _cur_c1 = _ws_syn["C1"].value or ""
    if "COMPARATIF SECTORIEL" in _cur_c1:
        _ws_syn["C1"] = f"COMPARATIF SECTORIEL — {sector_a} vs {sector_b}"

    # AGRÉGATS A4 : "TECHNOLOGIE — 88 valeurs" → "{A} — {N_a} valeurs"
    _a4 = _ws_agg["A4"].value or ""
    if "valeurs" in str(_a4):
        _ws_agg["A4"] = f"{sector_a.upper()} — {len(tickers_a)} valeurs"
    # AGRÉGATS A13 : "HEALTHCARE — 6 valeurs" → "{B} — {N_b} valeurs"
    _a13 = _ws_agg["A13"].value or ""
    if "valeurs" in str(_a13):
        _ws_agg["A13"] = f"{sector_b.upper()} — {len(tickers_b)} valeurs"

    # ── Step 3 : Injection des données tickers ────────────────────────
    _inject_data_tech(wb["DATA_TECH"], tickers_a, sector_a)
    _inject_data_hc(wb["DATA_HC"], tickers_b, sector_b)

    log.info(
        "[cmp_secteur_xlsx] Template rempli : %s (%d tickers) vs %s (%d tickers)",
        sector_a, len(tickers_a), sector_b, len(tickers_b),
    )

    # Sauvegarde en memoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# Fallback minimaliste si le template est introuvable
# ═══════════════════════════════════════════════════════════════════════════

def _generate_fallback_xlsx(tickers_a, sector_a, tickers_b, sector_b) -> bytes:
    """Fallback minimal si le template n'est pas dispo (erreur de deploiement).
    Cree juste 2 feuilles brutes pour ne pas casser le download stream."""
    try:
        from openpyxl import Workbook
    except ImportError as _ie:
        raise RuntimeError(f"openpyxl manquant : {_ie}")

    wb = Workbook()
    wb.active.title = "DATA_TECH"
    ws_t = wb.active
    ws_t["J1"] = sector_a
    ws_t["A3"] = "TICKER"
    ws_t["B3"] = "SOCI\u00c9T\u00c9"
    ws_t["C3"] = "SCORE FINSIGHT"
    ws_t["D3"] = "P/E"
    ws_t["E3"] = "EV/EBITDA"
    ws_t["F3"] = "MARGE EBITDA"
    ws_t["G3"] = "ROE"
    ws_t["H3"] = "CROISS. REV."
    for i, tk in enumerate(sorted(tickers_a, key=lambda t: _safe_float(t.get("score_global")) or 0, reverse=True)[:88]):
        for ci, v in enumerate(_ticker_row_values(tk), start=1):
            ws_t.cell(row=4 + i, column=ci, value=v)

    ws_h = wb.create_sheet("DATA_HC")
    ws_h["I1"] = sector_b
    for ci, h in enumerate(["TICKER", "SOCI\u00c9T\u00c9", "SCORE FINSIGHT", "P/E",
                            "EV/EBITDA", "MARGE EBITDA", "ROE", "CROISS. REV."], start=1):
        ws_h.cell(row=3, column=ci, value=h)
    for i, tk in enumerate(sorted(tickers_b, key=lambda t: _safe_float(t.get("score_global")) or 0, reverse=True)[:88]):
        for ci, v in enumerate(_ticker_row_values(tk), start=1):
            ws_h.cell(row=4 + i, column=ci, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
