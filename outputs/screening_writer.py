# =============================================================================
# FinSight IA — Screening Writer
# outputs/screening_writer.py
#
# Charge le template assets/FinSight_IA_Screening_CAC40_v3.xlsx et injecte
# les donnees aux positions exactes sans modifier le formatage existant.
# Fallback from-scratch si le template est absent.
# Zero LLM. Pure openpyxl.
# =============================================================================

from __future__ import annotations

import logging
from collections import defaultdict
from datetime import date
from pathlib import Path
from statistics import median
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Palette couleurs (IB style) — utilisee pour le fallback from-scratch
# ---------------------------------------------------------------------------
_NAVY   = "1B2A4A"
_DARK   = "2C3E50"
_LGRAY  = "F4F6F7"
_MGRAY  = "BDC3C7"
_WHITE  = "FFFFFF"
_GREEN  = "1a7a52"
_RED    = "C0392B"
_GOLD   = "B8922A"
_BLACK  = "111111"

_FILL_NAVY  = PatternFill("solid", fgColor=_NAVY)
_FILL_DARK  = PatternFill("solid", fgColor=_DARK)
_FILL_LGRAY = PatternFill("solid", fgColor=_LGRAY)
_FILL_MGRAY = PatternFill("solid", fgColor="D5D8DC")
_FILL_GOLD  = PatternFill("solid", fgColor="F0D080")

_FONT_HDR   = Font(name="Calibri", bold=True, color=_WHITE, size=10)
_FONT_SHDR  = Font(name="Calibri", bold=True, color=_WHITE, size=9)
_FONT_TITLE = Font(name="Calibri", bold=True, color=_WHITE, size=11)
_FONT_BOLD  = Font(name="Calibri", bold=True, color=_BLACK, size=9)
_FONT_BODY  = Font(name="Calibri", color=_BLACK, size=9)
_FONT_SMALL = Font(name="Calibri", color="5D6D7E", size=8, italic=True)
_FONT_POS   = Font(name="Calibri", color=_GREEN, size=9, bold=True)
_FONT_NEG   = Font(name="Calibri", color=_RED,   size=9, bold=True)

_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_L = Alignment(horizontal="left",   vertical="center")
_ALIGN_R = Alignment(horizontal="right",  vertical="center")

_THIN = Side(style="thin", color=_MGRAY)
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CCY = {"USD": "$", "EUR": "\u20ac", "GBP": "\u00a3", "CHF": "CHF ", "JPY": "\u00a5"}

# ---------------------------------------------------------------------------
# Helpers formatage
# ---------------------------------------------------------------------------

def _sym(ccy: Optional[str]) -> str:
    return _CCY.get(ccy or "USD", "$")


def _na(v: Any, fallback: str = "\u2014") -> str:
    return fallback if v is None else str(v)


def _fmt_price(v: Any, ccy: str = "USD") -> str:
    if v is None:
        return "\u2014"
    try:
        return f"{_sym(ccy)}{float(v):,.2f}"
    except (ValueError, TypeError):
        return str(v)


def _fmt_mds(v: Any, ccy: str = "USD") -> str:
    if v is None:
        return "\u2014"
    try:
        return f"{_sym(ccy)}{float(v):,.1f}"
    except (ValueError, TypeError):
        return str(v)


def _fmt_mult(v: Any) -> str:
    if v is None:
        return "\u2014"
    try:
        return f"{float(v):.1f}x"
    except (ValueError, TypeError):
        return str(v)


def _fmt_pct(v: Any, sign: bool = True) -> str:
    if v is None:
        return "\u2014"
    try:
        f = float(v)
        prefix = "+" if sign and f >= 0 else ""
        return f"{prefix}{f:.1f}%"
    except (ValueError, TypeError):
        return str(v)


def _fmt_score(v: Any) -> str:
    if v is None:
        return "N/A"
    try:
        return f"{int(float(v))}/100"
    except (ValueError, TypeError):
        return str(v)


def _fmt_z(v: Any) -> str:
    if v is None:
        return "\u2014"
    try:
        return f"{float(v):.2f}"
    except (ValueError, TypeError):
        return str(v)


def _fmt_cov(v: Any) -> str:
    if v is None:
        return "\u2014"
    if str(v).lower() in ("n/a", "none", "inf"):
        return "n/a"
    try:
        return f"{float(v):.0f}"
    except (ValueError, TypeError):
        return str(v)


def _sector_short(s: str) -> str:
    MAP = {
        "Consumer Discretionary": "Consumer Disc.",
        "Consumer Staples":       "Consumer Staples",
        "Information Technology": "Technology",
        "Communication Services": "Comm. Services",
        "Health Care":            "Healthcare",
        "Healthcare":             "Healthcare",
    }
    if not s:
        return ""
    return MAP.get(s, s)


def _signal(score: Optional[float]) -> str:
    if score is None:
        return "Neutre"
    s = float(score)
    if s >= 75:
        return "Surpond\u00e9rer"
    if s >= 55:
        return "Neutre"
    return "Sous-pond\u00e9rer"


def _dominant_ccy(data: list[dict]) -> str:
    from collections import Counter
    c = Counter(t.get("currency", "USD") or "USD" for t in data)
    return c.most_common(1)[0][0] if c else "USD"


# ---------------------------------------------------------------------------
# Helpers openpyxl — mode from-scratch (fallback)
# ---------------------------------------------------------------------------

def _w(ws, row: int, col: int, value: Any,
       font=None, fill=None, align=None, border=None, num_fmt: str = None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell


def _hdr(ws, row: int, col: int, value: str, span: int = 1,
         fill=None, font=None):
    cell = _w(ws, row, col, value,
              font=font or _FONT_HDR,
              fill=fill or _FILL_NAVY,
              align=_ALIGN_C)
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1,
        )
    return cell


def _title(ws, row: int, value: str, ncols: int, date_str: str = ""):
    txt = f"{value}  \u2022  {date_str}" if date_str else value
    _hdr(ws, row, 1, txt, span=ncols,
         font=_FONT_TITLE, fill=_FILL_NAVY)
    ws.row_dimensions[row].height = 22


def _section(ws, row: int, value: str, ncols: int):
    _hdr(ws, row, 1, f"  {value}", span=ncols,
         font=_FONT_SHDR, fill=_FILL_DARK)
    ws.row_dimensions[row].height = 16


def _set_col_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, row: int, col: int):
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


# ---------------------------------------------------------------------------
# Mode template — helpers d'injection (preservent le formatage existant)
# ---------------------------------------------------------------------------

def _v(ws, row: int, col: int, value: Any) -> None:
    """
    Injecte uniquement la valeur — preserve le formatage du template.
    Si la cellule cible est une MergedCell (non top-left), ecrit dans
    le coin superieur gauche de la plage fusionnee correspondante.
    """
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if (mr.min_row <= row <= mr.max_row and
                    mr.min_col <= col <= mr.max_col):
                cell = ws.cell(row=mr.min_row, column=mr.min_col)
                break
        else:
            return  # cellule fusionnee sans range trouvee — ignorer
    cell.value = value if value is not None else "\u2014"


def _med_vals(lst: list, key: str) -> Optional[float]:
    vals = [t.get(key) for t in lst if t.get(key) is not None]
    try:
        return median([float(v) for v in vals]) if vals else None
    except Exception:
        return None


def _signal_altman(z: Any) -> str:
    if z is None:
        return "\u2014"
    try:
        fz = float(z)
        if fz > 2.99:
            return "Sain"
        if fz > 1.81:
            return "Zone grise"
        return "Risque"
    except Exception:
        return "\u2014"


def _signal_beneish(m: Any) -> str:
    if m is None:
        return "\u2014"
    try:
        return "Manipulation" if float(m) > -1.78 else "Neutre"
    except Exception:
        return "\u2014"


# ---------------------------------------------------------------------------
# Mapping secteur Python → nom de feuille Excel
# ---------------------------------------------------------------------------

_SECTOR_SHEET_MAP = {
    "Technology":      "TECHNOLOGY",
    "Consumer Disc.":  "CONSUMER DISC.",
    "Financials":      "FINANCIALS",
    "Industrials":     "INDUSTRIALS",
    "Healthcare":      "HEALTHCARE",
    "Energy":          "ENERGY",
    "Materials":       "MATERIALS",
    "Real Estate":     "REAL ESTATE",
}

# Ligne ou injecter les tickers (en-tete du tableau ratios)
# Structure template : ligne merged (label) juste au-dessus, puis ticker row, puis 12 lignes valeurs
_RATIO_TICKER_ROW = {
    "TECHNOLOGY":     19,   # A18=merged label, A19='Ratio'/tickers, A20..A31=valeurs
    "CONSUMER DISC.": 20,   # A19=merged label, A20='Ratio'/tickers, A21..A32=valeurs
    "FINANCIALS":     19,   # identique TECHNOLOGY
    "INDUSTRIALS":    18,   # A17=merged label, A18='Ratio'/tickers, A19..A30=valeurs
    "HEALTHCARE":     18,
    "ENERGY":         18,
    "MATERIALS":      18,
    "REAL ESTATE":    18,
}

# 12 ratios a injecter dans l'ordre (label, cle dict, fonction de formatage)
_RATIO_ROWS = [
    ("EV/EBITDA",       "ev_ebitda",       _fmt_mult),
    ("P/E",             "pe",              _fmt_mult),
    ("Marge Brute",     "gross_margin",    lambda v: _fmt_pct(v, sign=False)),
    ("Marge EBITDA",    "ebitda_margin",   lambda v: _fmt_pct(v, sign=False)),
    ("Marge Nette",     "net_margin",      lambda v: _fmt_pct(v, sign=False)),
    ("Croissance Rev.", "revenue_growth",  _fmt_pct),
    ("ROE",             "roe",             _fmt_pct),
    ("Current Ratio",   "current_ratio",   _fmt_z),
    ("Net Debt/EBITDA", "net_debt_ebitda", _fmt_mult),
    ("Altman Z-Score",  "altman_z",        _fmt_z),
    ("Beneish M-Score", "beneish_m",       _fmt_z),
    ("Momentum 52W",    "momentum_52w",    _fmt_pct),
]


# ===========================================================================
# FONCTIONS D'INJECTION (mode template)
# ===========================================================================

def _inject_donnees_brutes(wb, data: list[dict]) -> None:
    """
    DONNEES BRUTES — A3, 40 lignes max, 30 colonnes.
    Ticker, Societe, Secteur, Cours, Mkt Cap, EV, Rev LTM, EBITDA LTM,
    EV/EBITDA, EV/Rev, P/E, EPS, Mg Brute, Mg EBITDA, Mg Nette, Cro Rev,
    ROE, ROA, Current R, ND/EBITDA, Altman Z, Beneish M, Mom 52W,
    Score Val, Score Gr, Score Qual, Score Mom, Score Global, Next Earn, Signal.
    """
    ws_name = "DONN\u00c9ES BRUTES"
    if ws_name not in wb.sheetnames:
        log.warning("[inject] feuille '%s' absente du template", ws_name)
        return
    ws = wb[ws_name]
    ccy = _dominant_ccy(data)

    for i, t in enumerate(data[:40]):
        r = 3 + i
        tccy = t.get("currency", ccy) or ccy
        sg = t.get("score_global")
        row_vals = [
            t.get("ticker", "") or "\u2014",
            t.get("company", "") or "\u2014",
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_price(t.get("price"), tccy),
            _fmt_mds(t.get("market_cap"), tccy),
            _fmt_mds(t.get("ev"), tccy),
            _fmt_mds(t.get("revenue_ltm"), tccy),
            _fmt_mds(t.get("ebitda_ltm"), tccy),
            _fmt_mult(t.get("ev_ebitda")),
            _fmt_mult(t.get("ev_revenue")),
            _fmt_mult(t.get("pe")),
            _fmt_price(t.get("eps"), tccy),
            _fmt_pct(t.get("gross_margin"),  sign=False),
            _fmt_pct(t.get("ebitda_margin"), sign=False),
            _fmt_pct(t.get("net_margin"),    sign=False),
            _fmt_pct(t.get("revenue_growth")),
            _fmt_pct(t.get("roe")),
            _fmt_pct(t.get("roa")),
            _fmt_z(t.get("current_ratio")),
            _fmt_mult(t.get("net_debt_ebitda")),
            _fmt_z(t.get("altman_z")),
            _fmt_z(t.get("beneish_m")),
            _fmt_pct(t.get("momentum_52w")),
            t.get("score_value"),
            t.get("score_growth"),
            t.get("score_quality"),
            t.get("score_momentum"),
            sg,
            _na(t.get("next_earnings")),
            _signal(sg),
        ]
        for j, v in enumerate(row_vals, 1):
            _v(ws, r, j, v)

    log.debug("[inject] DONNEES BRUTES : %d lignes", min(len(data), 40))


def _inject_dashboard(wb, data: list[dict], universe_name: str,
                      date_str: str) -> None:
    """
    DASHBOARD :
      A4 = nb societes  G4 = nb secteurs  M4 = meilleur ticker/score  S4 = date
      B11:F15 = Top5 Value    H11:L15 = Top5 Growth
      N11:R15 = Top5 Quality  T11:X15 = Top5 Momentum
      B19:R28 = Top10 Global (17 colonnes)
    """
    ws_name = "DASHBOARD"
    if ws_name not in wb.sheetnames:
        log.warning("[inject] feuille '%s' absente du template", ws_name)
        return
    ws = wb[ws_name]

    N = len(data)
    sectors_set = set(
        _sector_short(t.get("sector") or "")
        for t in data if t.get("sector")
    )
    best = max(data, key=lambda x: x.get("score_global") or 0)
    best_sg = best.get("score_global")

    # KPIs ligne 7 (ligne 6 = labels dans le template)
    _v(ws, 7, 1,  N)                                                # A7
    _v(ws, 7, 7,  len(sectors_set))                                 # G7
    _v(ws, 7, 13, f"{best['ticker']} {int(best_sg or 0)}/100")      # M7
    _v(ws, 7, 19, date_str)                                         # S7

    # Top 5 par categorie — colonnes de depart : B=2, H=8, N=14, T=20
    top5_cats = [
        ("score_value",    "ev_ebitda",      _fmt_mult,  2),
        ("score_growth",   "revenue_growth", _fmt_pct,   8),
        ("score_quality",  "altman_z",       _fmt_z,     14),
        ("score_momentum", "momentum_52w",   _fmt_pct,   20),
    ]
    for score_key, metric_key, fmt_fn, start_col in top5_cats:
        top5 = sorted(
            [t for t in data if t.get(score_key) is not None],
            key=lambda x, sk=score_key: x[sk], reverse=True
        )[:5]
        for i, t in enumerate(top5):
            r = 11 + i
            _v(ws, r, start_col,     i + 1)
            _v(ws, r, start_col + 1, t.get("ticker", ""))
            _v(ws, r, start_col + 2, t.get("company", ""))
            _v(ws, r, start_col + 3, _fmt_score(t.get(score_key)))
            _v(ws, r, start_col + 4, fmt_fn(t.get(metric_key)))

    # Top 10 Global — B19:R28 (17 colonnes, depart col B=2)
    ccy = _dominant_ccy(data)
    top10 = sorted(
        [t for t in data if t.get("score_global") is not None],
        key=lambda x: x["score_global"], reverse=True
    )[:10]
    for i, t in enumerate(top10):
        r = 19 + i
        tccy = t.get("currency", ccy) or ccy
        sg = t.get("score_global")
        vals = [
            i + 1,
            t.get("ticker", ""),
            t.get("company", ""),
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_score(sg),
            _fmt_score(t.get("score_value")),
            _fmt_score(t.get("score_growth")),
            _fmt_score(t.get("score_quality")),
            _fmt_score(t.get("score_momentum")),
            _fmt_price(t.get("price"), tccy),
            _fmt_mds(t.get("market_cap"), tccy),
            _fmt_mult(t.get("ev_ebitda")),
            _fmt_pct(t.get("ebitda_margin"), sign=False),
            _fmt_pct(t.get("revenue_growth")),
            _fmt_pct(t.get("momentum_52w")),
            _na(t.get("next_earnings")),
            _signal(sg),
        ]
        for j, v in enumerate(vals, 2):  # col B = 2
            _v(ws, r, j, v)

    log.debug("[inject] DASHBOARD : Top10 injecte, KPIs mis a jour")


def _inject_value(wb, data: list[dict]) -> None:
    """
    VALUE — A5:M19 (15 lignes, 13 cols), medianes F21:M21.
    Cols : Rang, Ticker, Societe, Secteur, Score, EV/EBITDA, EV/Revenue, P/E,
           Mg Brute, Mg EBITDA, Altman Z, FCF Yield (—), Decote DCF (—).
    """
    ws_name = "VALUE"
    if ws_name not in wb.sheetnames:
        log.warning("[inject] feuille '%s' absente du template", ws_name)
        return
    ws = wb[ws_name]

    sorted_data = sorted(
        [t for t in data if t.get("score_value") is not None],
        key=lambda x: x["score_value"], reverse=True
    )[:15]

    for i, t in enumerate(sorted_data):
        r = 5 + i
        vals = [
            i + 1,
            t.get("ticker", ""),
            t.get("company", ""),
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_score(t.get("score_value")),
            _fmt_mult(t.get("ev_ebitda")),
            _fmt_mult(t.get("ev_revenue")),
            _fmt_mult(t.get("pe")),
            _fmt_pct(t.get("gross_margin"),  sign=False),
            _fmt_pct(t.get("ebitda_margin"), sign=False),
            _fmt_z(t.get("altman_z")),
            "\u2014",   # FCF Yield — non calcule
            "\u2014",   # Decote DCF — non calculee
        ]
        for j, v in enumerate(vals, 1):
            _v(ws, r, j, v)

    # Medianes ligne 22 (row 21 = label merge, row 22 = cellules vides)
    # Cols B a M (2..13) — on injecte les medianes des colonnes numeriques
    med_defs = [
        (2,  "ev_ebitda",    _fmt_mult),
        (3,  "ev_revenue",   _fmt_mult),
        (4,  "pe",           _fmt_mult),
        (5,  "gross_margin", lambda v: _fmt_pct(v, sign=False)),
        (6,  "ebitda_margin",lambda v: _fmt_pct(v, sign=False)),
        (7,  "altman_z",     _fmt_z),
    ]
    for col, key, fmt_fn in med_defs:
        med = _med_vals(sorted_data, key)
        _v(ws, 22, col, fmt_fn(med) if med is not None else "\u2014")
    _v(ws, 22, 8,  "\u2014")   # FCF Yield
    _v(ws, 22, 9,  "\u2014")   # Decote DCF

    log.debug("[inject] VALUE : %d lignes", len(sorted_data))


def _inject_growth(wb, data: list[dict]) -> None:
    """
    GROWTH — A5:L19 (15 lignes, 12 cols).
    Cols : Rang, Ticker, Societe, Secteur, Score, Cro Rev YoY, Mg EBITDA,
           Mg Nette, ROE, Cro EPS (—), Revisions Analystes (—), Signal.
    """
    ws_name = "GROWTH"
    if ws_name not in wb.sheetnames:
        log.warning("[inject] feuille '%s' absente du template", ws_name)
        return
    ws = wb[ws_name]

    sorted_data = sorted(
        [t for t in data if t.get("score_growth") is not None],
        key=lambda x: x["score_growth"], reverse=True
    )[:15]

    for i, t in enumerate(sorted_data):
        r = 5 + i
        vals = [
            i + 1,
            t.get("ticker", ""),
            t.get("company", ""),
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_score(t.get("score_growth")),
            _fmt_pct(t.get("revenue_growth")),
            _fmt_pct(t.get("ebitda_margin"), sign=False),
            _fmt_pct(t.get("net_margin"),    sign=False),
            _fmt_pct(t.get("roe")),
            "\u2014",   # Cro EPS — non calcule
            "\u2014",   # Revisions Analystes — non calcule
            _signal(t.get("score_global")),
        ]
        for j, v in enumerate(vals, 1):
            _v(ws, r, j, v)

    log.debug("[inject] GROWTH : %d lignes", len(sorted_data))


def _inject_quality(wb, data: list[dict]) -> None:
    """
    QUALITY — A5:M19 (15 lignes, 13 cols), medianes F21:M21.
    Cols : Rang, Ticker, Societe, Secteur, Score, Altman Z, Signal Z,
           Beneish M, Signal M, Mg EBITDA, ROE, Interest Coverage, Current Ratio.
    """
    ws_name = "QUALITY"
    if ws_name not in wb.sheetnames:
        log.warning("[inject] feuille '%s' absente du template", ws_name)
        return
    ws = wb[ws_name]

    sorted_data = sorted(
        [t for t in data if t.get("score_quality") is not None],
        key=lambda x: x["score_quality"], reverse=True
    )[:15]

    for i, t in enumerate(sorted_data):
        r = 5 + i
        az = t.get("altman_z")
        bm = t.get("beneish_m")
        vals = [
            i + 1,
            t.get("ticker", ""),
            t.get("company", ""),
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_score(t.get("score_quality")),
            _fmt_z(az),
            _signal_altman(az),
            _fmt_z(bm),
            _signal_beneish(bm),
            _fmt_pct(t.get("ebitda_margin"), sign=False),
            _fmt_pct(t.get("roe")),
            _fmt_cov(t.get("interest_coverage")),
            _fmt_z(t.get("current_ratio")),
        ]
        for j, v in enumerate(vals, 1):
            _v(ws, r, j, v)

    # Medianes ligne 22 (row 21 = label merge)
    # Cols B a M (2..13) : Altman Z(2), Signal Z(3), Beneish M(4), Signal M(5),
    # Mg EBITDA(6), ROE(7), Int Cov(8), Current Ratio(9)
    med_defs = [
        (2,  "altman_z",      _fmt_z),
        (4,  "beneish_m",     _fmt_z),
        (6,  "ebitda_margin", lambda v: _fmt_pct(v, sign=False)),
        (7,  "roe",           _fmt_pct),
        (9,  "current_ratio", _fmt_z),
    ]
    for col, key, fmt_fn in med_defs:
        med = _med_vals(sorted_data, key)
        _v(ws, 22, col, fmt_fn(med) if med is not None else "\u2014")
    _v(ws, 22, 3,  "\u2014")   # Signal Z — textuel
    _v(ws, 22, 5,  "\u2014")   # Signal M — textuel
    _v(ws, 22, 8,  "\u2014")   # Interest Coverage

    log.debug("[inject] QUALITY : %d lignes", len(sorted_data))


def _inject_momentum(wb, data: list[dict]) -> None:
    """
    MOMENTUM — A5:K19 (15 lignes, 11 cols), medianes F21:K21.
    Cols : Rang, Ticker, Societe, Secteur, Score, Momentum 52W, Cours,
           Mkt Cap, Score Global, Next Earnings, Signal.
    """
    ws_name = "MOMENTUM"
    if ws_name not in wb.sheetnames:
        log.warning("[inject] feuille '%s' absente du template", ws_name)
        return
    ws = wb[ws_name]
    ccy = _dominant_ccy(data)

    sorted_data = sorted(
        [t for t in data if t.get("score_momentum") is not None],
        key=lambda x: x["score_momentum"], reverse=True
    )[:15]

    for i, t in enumerate(sorted_data):
        r = 5 + i
        tccy = t.get("currency", ccy) or ccy
        vals = [
            i + 1,
            t.get("ticker", ""),
            t.get("company", ""),
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_score(t.get("score_momentum")),
            _fmt_pct(t.get("momentum_52w")),
            _fmt_price(t.get("price"), tccy),
            _fmt_mds(t.get("market_cap"), tccy),
            _fmt_score(t.get("score_global")),
            _na(t.get("next_earnings")),
            _signal(t.get("score_global")),
        ]
        for j, v in enumerate(vals, 1):
            _v(ws, r, j, v)

    # Medianes ligne 22 (row 21 = label merge)
    med_mom = _med_vals(sorted_data, "momentum_52w")
    _v(ws, 22, 2, _fmt_pct(med_mom) if med_mom is not None else "\u2014")
    for col in range(3, 8):
        _v(ws, 22, col, "\u2014")

    log.debug("[inject] MOMENTUM : %d lignes", len(sorted_data))


def _inject_par_secteur(wb, data: list[dict]) -> None:
    """
    PAR SECTEUR :
      A5:J12  = 8 secteurs (10 colonnes)
      A16:F23 = composition sectorielle (6 colonnes)
      A24     = ligne TOTAL
    """
    ws_name = "PAR SECTEUR"
    if ws_name not in wb.sheetnames:
        log.warning("[inject] feuille '%s' absente du template", ws_name)
        return
    ws = wb[ws_name]

    sectors: dict[str, list] = defaultdict(list)
    for t in data:
        sec = _sector_short(t.get("sector") or "") or "Autres"
        sectors[sec].append(t)

    total_mc = sum(t.get("market_cap") or 0 for t in data)

    sorted_secs = sorted(
        sectors.items(),
        key=lambda x: -sum((t.get("market_cap") or 0) for t in x[1])
    )

    # A5:J12 — analyse par secteur (max 8 lignes)
    for i, (sec, tlist) in enumerate(sorted_secs[:8]):
        r = 5 + i
        sec_mc  = sum(t.get("market_cap") or 0 for t in tlist)
        poids   = (sec_mc / total_mc * 100) if total_mc else None
        avg_sc  = _med_vals(tlist, "score_global")
        top_soc = max(tlist, key=lambda x: x.get("score_global") or 0)

        vals = [
            sec,
            len(tlist),
            _fmt_pct(poids, sign=False),
            _fmt_score(avg_sc),
            top_soc.get("company", ""),
            _fmt_mult(_med_vals(tlist, "ev_ebitda")),
            _fmt_pct(_med_vals(tlist, "ebitda_margin"), sign=False),
            _fmt_pct(_med_vals(tlist, "revenue_growth")),
            _fmt_z(_med_vals(tlist, "altman_z")),
            _signal(avg_sc),
        ]
        for j, v in enumerate(vals, 1):
            _v(ws, r, j, v)

    # A16:F23 — composition sectorielle (max 8 lignes)
    for i, (sec, tlist) in enumerate(sorted_secs[:8]):
        r = 16 + i
        sec_mc = sum(t.get("market_cap") or 0 for t in tlist)
        poids  = (sec_mc / total_mc * 100) if total_mc else 0
        avg_sc = _med_vals(tlist, "score_global")
        tend   = (
            "\u2191 Positif" if (avg_sc or 0) >= 60 else
            "\u2193 Negatif" if (avg_sc or 0) < 45 else
            "\u2192 Neutre"
        )
        vals = [
            sec,
            len(tlist),
            round(poids, 1),
            round(sec_mc, 1),
            _fmt_score(avg_sc),
            tend,
        ]
        for j, v in enumerate(vals, 1):
            _v(ws, r, j, v)

    # A24 = TOTAL
    _v(ws, 24, 1, "TOTAL")
    _v(ws, 24, 2, len(data))
    _v(ws, 24, 3, "100.0%")
    _v(ws, 24, 4, round(total_mc, 1))

    log.debug("[inject] PAR SECTEUR : %d secteurs", len(sorted_secs))


def _inject_sector_sheets(wb, data: list[dict]) -> None:
    """
    Feuilles sectorielles (TECHNOLOGY, CONSUMER DISC., FINANCIALS, INDUSTRIALS,
    HEALTHCARE, ENERGY, MATERIALS, REAL ESTATE) :
      D4 = Mkt Cap secteur  G4 = Score moyen  J4 = meilleure societe  M4 = EV/EBITDA med
      A9 = debut tableau detaille (16 cols)
      Tableau ratios : ligne 19 (TECH/CONS/FIN) ou 17 (autres), 12 ratios
    """
    sectors: dict[str, list] = defaultdict(list)
    for t in data:
        sec = _sector_short(t.get("sector") or "") or "Autres"
        sectors[sec].append(t)

    ccy = _dominant_ccy(data)

    for sec_name, sheet_name in _SECTOR_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        tlist = sectors.get(sec_name, [])
        if not tlist:
            log.debug("[inject] %s : aucune societe, feuille ignoree", sheet_name)
            continue

        ws = wb[sheet_name]
        sorted_tlist = sorted(
            tlist, key=lambda x: x.get("score_global") or 0, reverse=True
        )

        sec_mc  = sum(t.get("market_cap") or 0 for t in tlist)
        avg_sc  = _med_vals(tlist, "score_global")
        best    = sorted_tlist[0]
        med_ev  = _med_vals(tlist, "ev_ebitda")

        # KPIs ligne 4
        _v(ws, 4, 1,  len(tlist))                  # A4 — nb societes
        _v(ws, 4, 4,  round(sec_mc, 1))            # D4
        _v(ws, 4, 7,  int(avg_sc or 0))            # G4
        _v(ws, 4, 10, best.get("company", ""))      # J4
        _v(ws, 4, 13, _fmt_mult(med_ev))            # M4

        # Tableau detaille a partir de A9 (16 colonnes)
        for i, t in enumerate(sorted_tlist):
            r = 9 + i
            tccy = t.get("currency", ccy) or ccy
            vals = [
                i + 1,
                t.get("ticker", ""),
                t.get("company", ""),
                t.get("score_global"),
                t.get("score_value"),
                t.get("score_growth"),
                t.get("score_quality"),
                t.get("score_momentum"),
                _fmt_price(t.get("price"), tccy),
                _fmt_mds(t.get("market_cap"), tccy),
                _fmt_mult(t.get("ev_ebitda")),
                _fmt_mult(t.get("pe")),
                _fmt_pct(t.get("gross_margin"),  sign=False),
                _fmt_pct(t.get("ebitda_margin"), sign=False),
                _fmt_pct(t.get("revenue_growth")),
                _signal(t.get("score_global")),
            ]
            for j, v in enumerate(vals, 1):
                _v(ws, r, j, v)

        # Tableau ratios
        # ticker_row = ligne des en-tetes tickers (A col = label deja dans template)
        # valeurs a partir de ticker_row + 1
        ticker_row = _RATIO_TICKER_ROW.get(sheet_name, 18)
        n = len(sorted_tlist)
        med_col = 2 + n  # colonne mediane (tickers en B=2..col 2+n-1)

        # Tickers en B..Bn
        for j, t in enumerate(sorted_tlist, 2):
            _v(ws, ticker_row, j, t.get("ticker", ""))
        _v(ws, ticker_row, med_col, "Mediane")

        # 12 lignes ratios a partir de ticker_row + 1
        for ratio_idx, (_, ratio_key, fmt_fn) in enumerate(_RATIO_ROWS):
            r = ticker_row + 1 + ratio_idx
            vals_raw: list[float] = []
            for j, t in enumerate(sorted_tlist, 2):
                v = t.get(ratio_key)
                _v(ws, r, j, fmt_fn(v))
                if v is not None:
                    try:
                        vals_raw.append(float(v))
                    except (ValueError, TypeError):
                        pass
            med = median(vals_raw) if vals_raw else None
            _v(ws, r, med_col, fmt_fn(med) if med is not None else "\u2014")

        log.debug("[inject] %s : %d societes, tickers ligne %d",
                  sheet_name, n, ticker_row)


# ===========================================================================
# FEUILLES FROM-SCRATCH (fallback si template absent)
# ===========================================================================

def _build_dashboard(wb: Workbook, data: list[dict],
                     universe_name: str, date_str: str):
    ws = wb.create_sheet("DASHBOARD")
    N = len(data)
    ccy = _dominant_ccy(data)
    NCOLS = 20

    _title(ws, 1, f"FinSight IA  \u2022  Screening {N} soci\u00e9t\u00e9s  \u2022  {universe_name}",
           NCOLS, date_str)
    _w(ws, 2, 1,
       "Pipeline : Supabase Cache \u2192 Calculs Python/NumPy \u2192 Cours temps r\u00e9el yfinance"
       "  \u2022  Usage confidentiel  \u2022  FinSight IA v1.0",
       font=_FONT_SMALL, fill=_FILL_DARK,
       align=Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    ws.row_dimensions[2].height = 14

    best = max(data, key=lambda x: x.get("score_global") or 0)
    sectors = [t.get("sector") or "" for t in data]
    dominant_sector = max(set(sectors), key=sectors.count) if sectors else ""

    kpi_labels = [
        ("Univers analys\u00e9", f"{N} soci\u00e9t\u00e9s", 1),
        ("Meilleur score global",
         f"{best['ticker']} \u2192 {_fmt_score(best.get('score_global'))}",  5),
        ("Secteur dominant", _sector_short(dominant_sector), 9),
        ("Date d'analyse", date_str, 13),
    ]
    for label, val, col in kpi_labels:
        _w(ws, 4, col, val,   font=_FONT_BOLD, align=_ALIGN_C, fill=_FILL_GOLD)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+3)
        _w(ws, 5, col, label, font=_FONT_SMALL, align=_ALIGN_C)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+3)

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 14

    row = 7
    cats = [
        ("TOP 5  \u2022  VALUE",    "score_value",    "EV/EBITDA",   "ev_ebitda", _fmt_mult, 1),
        ("TOP 5  \u2022  GROWTH",   "score_growth",   "Cro. Rev.",   "revenue_growth", _fmt_pct, 5),
        ("TOP 5  \u2022  QUALITY",  "score_quality",  "Altman Z",    "altman_z",  _fmt_z,  10),
        ("TOP 5  \u2022  MOMENTUM", "score_momentum", "Perf. 52W",   "momentum_52w", _fmt_pct, 15),
    ]
    for label, score_key, metric_label, metric_key, fmt_fn, col in cats:
        _hdr(ws, row, col, label, span=4, fill=_FILL_NAVY)
        _w(ws, row+1, col,   "Rang",     font=_FONT_BOLD, fill=_FILL_LGRAY, align=_ALIGN_C)
        _w(ws, row+1, col+1, "Ticker",   font=_FONT_BOLD, fill=_FILL_LGRAY, align=_ALIGN_C)
        _w(ws, row+1, col+2, "Soci\u00e9t\u00e9", font=_FONT_BOLD, fill=_FILL_LGRAY, align=_ALIGN_C)
        _w(ws, row+1, col+3, metric_label, font=_FONT_BOLD, fill=_FILL_LGRAY, align=_ALIGN_C)

        top5 = sorted(
            [t for t in data if t.get(score_key) is not None],
            key=lambda x, sk=score_key: x[sk], reverse=True
        )[:5]

        for i, t in enumerate(top5, 1):
            r = row + 1 + i
            fill = _FILL_LGRAY if i % 2 == 0 else None
            _w(ws, r, col,   i,                  font=_FONT_BODY, fill=fill, align=_ALIGN_C)
            _w(ws, r, col+1, t.get("ticker",""),  font=_FONT_BOLD, fill=fill, align=_ALIGN_C)
            _w(ws, r, col+2, t.get("company",""), font=_FONT_BODY, fill=fill, align=_ALIGN_L)
            _w(ws, r, col+3, fmt_fn(t.get(metric_key)), font=_FONT_BODY, fill=fill, align=_ALIGN_R)

    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row+1].height = 14

    row = 14
    _section(ws, row, "TOP 10  \u2022  CLASSEMENT GLOBAL  \u2022  Score composite Value / Growth / Quality / Momentum", NCOLS)
    hdrs10 = ["Rang","Ticker","Soci\u00e9t\u00e9","Secteur","Score Global",
              "Score Value","Score Growth","Score Quality","Score Mom.",
              f"Cours ({_sym(ccy)})","Mkt Cap (Mds)",
              "EV/EBITDA","P/E","Marge EBITDA","Cro. Rev.","Mom. 52W",
              "Altman Z","Next Earnings"]
    for j, h in enumerate(hdrs10, 1):
        _w(ws, row+1, j, h, font=_FONT_BOLD, fill=_FILL_MGRAY, align=_ALIGN_C)

    top10 = sorted(
        [t for t in data if t.get("score_global") is not None],
        key=lambda x: x["score_global"], reverse=True
    )[:10]

    for i, t in enumerate(top10, 1):
        r = row + 1 + i
        fill = _FILL_LGRAY if i % 2 == 0 else None
        tccy = t.get("currency", ccy) or ccy
        vals = [
            i, t.get("ticker",""), t.get("company",""),
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_score(t.get("score_global")),
            _fmt_score(t.get("score_value")),
            _fmt_score(t.get("score_growth")),
            _fmt_score(t.get("score_quality")),
            _fmt_score(t.get("score_momentum")),
            _fmt_price(t.get("price"), tccy),
            _fmt_mds(t.get("market_cap"), tccy),
            _fmt_mult(t.get("ev_ebitda")),
            _fmt_mult(t.get("pe")),
            _fmt_pct(t.get("ebitda_margin"), sign=False),
            _fmt_pct(t.get("revenue_growth")),
            _fmt_pct(t.get("momentum_52w")),
            _fmt_z(t.get("altman_z")),
            _na(t.get("next_earnings")),
        ]
        for j, v in enumerate(vals, 1):
            _w(ws, r, j, v, font=_FONT_BODY, fill=fill,
               align=_ALIGN_C if j in (1,2,5,6,7,8,9,10,11,12,13,18) else _ALIGN_L)

    _set_col_widths(ws, [4,8,22,16,11,10,10,10,10,10,11,9,8,12,10,9,8,12,10,10])
    _freeze(ws, 2, 1)
    ws.sheet_view.showGridLines = False


def _build_comparables(wb: Workbook, data: list[dict], date_str: str):
    ws = wb.create_sheet("COMPARABLES")
    NCOLS = 16
    ccy = _dominant_ccy(data)
    sym = _sym(ccy)

    _title(ws, 1,
           "FinSight IA  \u2022  Analyse Comparable  \u2022  Style IB",
           NCOLS, date_str)

    hdrs = ["#","Ticker","Soci\u00e9t\u00e9","Secteur",
            f"Cours ({sym})",f"EV ({sym}Mds)",f"Rev. LTM ({sym}Mds)",
            f"EBITDA LTM ({sym}Mds)","EV/EBITDA (x)","EV/Revenue (x)",
            "EBITDA Gr. NTM (%)","EPS LTM","P/E (x)",
            "Marge Brute","Marge EBITDA","Altman Z"]
    for j, h in enumerate(hdrs, 1):
        _w(ws, 3, j, h, font=_FONT_BOLD, fill=_FILL_MGRAY, align=_ALIGN_C,
           border=_BORDER_THIN)

    sorted_data = sorted(data, key=lambda x: x.get("score_global") or 0, reverse=True)

    for i, t in enumerate(sorted_data, 1):
        r = i + 3
        fill = _FILL_LGRAY if i % 2 == 0 else None
        tccy = t.get("currency", ccy) or ccy
        row_vals = [
            i,
            t.get("ticker",""),
            t.get("company",""),
            _sector_short(t.get("sector") or "") or "\u2014",
            _fmt_price(t.get("price"), tccy),
            _fmt_mds(t.get("ev"), tccy),
            _fmt_mds(t.get("revenue_ltm"), tccy),
            _fmt_mds(t.get("ebitda_ltm"), tccy),
            _fmt_mult(t.get("ev_ebitda")),
            _fmt_mult(t.get("ev_revenue")),
            _fmt_pct(t.get("ebitda_ntm_growth")),
            _fmt_price(t.get("eps"), tccy),
            _fmt_mult(t.get("pe")),
            _fmt_pct(t.get("gross_margin"), sign=False),
            _fmt_pct(t.get("ebitda_margin"), sign=False),
            _fmt_z(t.get("altman_z")),
        ]
        for j, v in enumerate(row_vals, 1):
            _w(ws, r, j, v, font=_FONT_BODY, fill=fill,
               align=_ALIGN_C if j in (1,2,9,10,11,12,13,16) else
               _ALIGN_L if j in (3,4) else _ALIGN_R,
               border=_BORDER_THIN)

    _set_col_widths(ws, [4,8,24,16,10,9,10,10,9,9,11,9,8,11,12,9])
    _freeze(ws, 4, 3)
    ws.sheet_view.showGridLines = False


def _build_ratios_detailles(wb: Workbook, data: list[dict], date_str: str):
    ws = wb.create_sheet("RATIOS DETAILLES")
    tickers = [t["ticker"] for t in data]
    N = len(tickers)
    NCOLS = N + 2

    _title(ws, 1, "FinSight IA  \u2022  Ratios Financiers D\u00e9taill\u00e9s", NCOLS, date_str)

    _w(ws, 3, 1, "Cat\u00e9gorie / Ratio", font=_FONT_BOLD, fill=_FILL_MGRAY, align=_ALIGN_L)
    for j, t in enumerate(data, 2):
        _w(ws, 3, j, t["ticker"], font=_FONT_BOLD, fill=_FILL_MGRAY, align=_ALIGN_C)
    _w(ws, 3, N+2, "M\u00e9diane", font=_FONT_BOLD, fill=_FILL_GOLD, align=_ALIGN_C)

    def _ratio_row(r: int, label: str, key: str, fmt_fn):
        fill_row = _FILL_LGRAY if r % 2 == 0 else None
        _w(ws, r, 1, f"  {label}", font=_FONT_BODY, fill=fill_row, align=_ALIGN_L)
        vals_raw = []
        for j, t in enumerate(data, 2):
            v = t.get(key)
            txt = fmt_fn(v)
            f = _FONT_BODY
            if v is not None:
                try:
                    fv = float(v)
                    vals_raw.append(fv)
                    if key in ("revenue_growth","ebitda_ntm_growth","roe","roa","momentum_52w"):
                        f = _FONT_POS if fv >= 0 else _FONT_NEG
                except (ValueError, TypeError):
                    pass
            _w(ws, r, j, txt, font=f, fill=fill_row, align=_ALIGN_R)
        if vals_raw:
            try:
                med = median(vals_raw)
                _w(ws, r, N+2, fmt_fn(med), font=_FONT_BOLD, fill=_FILL_GOLD, align=_ALIGN_R)
            except Exception:
                _w(ws, r, N+2, "\u2014", font=_FONT_BODY, fill=_FILL_GOLD, align=_ALIGN_C)
        else:
            _w(ws, r, N+2, "\u2014", font=_FONT_BODY, fill=_FILL_GOLD, align=_ALIGN_C)

    rows_def = [
        ("  VALORISATION", None, None, True),
        ("EV/EBITDA",          "ev_ebitda",       _fmt_mult, False),
        ("EV/Revenue",         "ev_revenue",       _fmt_mult, False),
        ("P/E",                "pe",               _fmt_mult, False),
        ("EPS",                "eps",              lambda v: _fmt_price(v), False),
        ("  RENTABILIT\u00c9", None, None, True),
        ("Marge Brute",        "gross_margin",     lambda v: _fmt_pct(v, sign=False), False),
        ("Marge EBITDA",       "ebitda_margin",    lambda v: _fmt_pct(v, sign=False), False),
        ("Marge Nette",        "net_margin",       lambda v: _fmt_pct(v, sign=False), False),
        ("ROE",                "roe",              _fmt_pct, False),
        ("ROA",                "roa",              _fmt_pct, False),
        ("  LIQUIDIT\u00c9 & SOLVABILIT\u00c9", None, None, True),
        ("Current Ratio",      "current_ratio",    lambda v: _fmt_z(v), False),
        ("Net Debt/EBITDA",    "net_debt_ebitda",  _fmt_mult, False),
        ("Interest Coverage",  "interest_coverage",_fmt_cov, False),
        ("  CROISSANCE",       None, None, True),
        ("Croissance Rev. YoY","revenue_growth",   _fmt_pct, False),
        ("Croissance EBITDA NTM","ebitda_ntm_growth",_fmt_pct, False),
        ("  QUALIT\u00c9",     None, None, True),
        ("Altman Z-Score",     "altman_z",         _fmt_z, False),
        ("Beneish M-Score",    "beneish_m",        _fmt_z, False),
        ("Momentum 52W (%)",   "momentum_52w",     _fmt_pct, False),
        ("Score Global",       "score_global",     _fmt_score, False),
    ]

    r = 4
    for item in rows_def:
        label, key, fmt_fn, is_section = item
        if is_section:
            _section(ws, r, label, NCOLS)
        else:
            _ratio_row(r, label, key, fmt_fn)
        r += 1

    col_widths = [22] + [9] * N + [9]
    _set_col_widths(ws, col_widths)
    _freeze(ws, 4, 2)
    ws.sheet_view.showGridLines = False


def _build_par_secteur(wb: Workbook, data: list[dict], date_str: str):
    ws = wb.create_sheet("PAR SECTEUR")
    NCOLS = 16

    _title(ws, 1, "FinSight IA  \u2022  Vue Sectorielle & Composition", NCOLS, date_str)
    _section(ws, 3,
             "ANALYSE PAR SECTEUR  \u2022  Score, Ratios cl\u00e9s, Signal", NCOLS)

    hdrs = ["Secteur","Nb Soc.","Poids Mkt Cap","Score Moyen","Top Soci\u00e9t\u00e9",
            "EV/EBITDA M\u00e9d.","Marge EBITDA M\u00e9d.","Croissance M\u00e9d.",
            "Altman Z M\u00e9d.","Signal"]
    for j, h in enumerate(hdrs, 1):
        _w(ws, 4, j, h, font=_FONT_BOLD, fill=_FILL_MGRAY, align=_ALIGN_C)

    sectors: dict[str, list] = defaultdict(list)
    for t in data:
        sec = _sector_short(t.get("sector") or "") or "Autres"
        sectors[sec].append(t)

    total_mc = sum(t.get("market_cap") or 0 for t in data)

    def _med(lst, key):
        vals = [t.get(key) for t in lst if t.get(key) is not None]
        try:
            return median(vals) if vals else None
        except Exception:
            return None

    r = 5
    for sec, tlist in sorted(sectors.items(), key=lambda x: -sum((t.get("market_cap") or 0) for t in x[1])):
        sec_mc = sum(t.get("market_cap") or 0 for t in tlist)
        poids  = (sec_mc / total_mc * 100) if total_mc else None
        avg_sc = _med(tlist, "score_global")
        top    = max(tlist, key=lambda x: x.get("score_global") or 0)
        sig    = _signal(avg_sc)

        fill = _FILL_LGRAY if r % 2 == 0 else None
        vals = [
            sec, len(tlist),
            _fmt_pct(poids, sign=False),
            _fmt_score(avg_sc),
            top.get("company",""),
            _fmt_mult(_med(tlist, "ev_ebitda")),
            _fmt_pct(_med(tlist, "ebitda_margin"), sign=False),
            _fmt_pct(_med(tlist, "revenue_growth")),
            _fmt_z(_med(tlist, "altman_z")),
            sig,
        ]
        for j, v in enumerate(vals, 1):
            _w(ws, r, j, v, font=_FONT_BODY, fill=fill,
               align=_ALIGN_L if j in (1,5,10) else _ALIGN_C)
        r += 1

    r += 1
    _section(ws, r, "COMPOSITION SECTORIELLE  \u2022  Donn\u00e9es pour graphique", NCOLS)
    hdrs2 = ["Secteur","Nb Soci\u00e9t\u00e9s","Poids Mkt Cap (%)",
             "Mkt Cap Total (Mds)","Score Moyen","Nb Soc. score>70",
             "Nb Soc. score<50","Tendance"]
    for j, h in enumerate(hdrs2, 1):
        _w(ws, r+1, j, h, font=_FONT_BOLD, fill=_FILL_MGRAY, align=_ALIGN_C)

    total_row = r + 2
    ccy = _dominant_ccy(data)
    for sec, tlist in sorted(sectors.items(), key=lambda x: -sum((t.get("market_cap") or 0) for t in x[1])):
        sec_mc  = sum(t.get("market_cap") or 0 for t in tlist)
        poids   = (sec_mc / total_mc * 100) if total_mc else 0
        avg_sc  = _med(tlist, "score_global")
        nb_buy  = sum(1 for t in tlist if (t.get("score_global") or 0) >= 70)
        nb_sell = sum(1 for t in tlist if (t.get("score_global") or 0) < 50)
        tend    = "\u2191 Positif" if (avg_sc or 0) >= 60 else ("\u2193 N\u00e9gatif" if (avg_sc or 0) < 45 else "\u2192 Neutre")

        fill = _FILL_LGRAY if total_row % 2 == 0 else None
        for j, v in enumerate([
            sec, len(tlist), round(poids,1), round(sec_mc,1),
            _fmt_score(avg_sc), nb_buy, nb_sell, tend
        ], 1):
            _w(ws, total_row, j, v, font=_FONT_BODY, fill=fill,
               align=_ALIGN_L if j in (1,8) else _ALIGN_C)
        total_row += 1

    _w(ws, total_row, 1, "TOTAL", font=_FONT_BOLD, align=_ALIGN_L)
    _w(ws, total_row, 2, len(data), font=_FONT_BOLD, align=_ALIGN_C)
    _w(ws, total_row, 3, "100.0%", font=_FONT_BOLD, align=_ALIGN_C)
    _w(ws, total_row, 4, round(total_mc, 1), font=_FONT_BOLD, align=_ALIGN_C)

    _set_col_widths(ws, [20,8,12,14,24,10,11,12,10,14,10,10,10,10,10,10])
    ws.sheet_view.showGridLines = False


def _build_donnees_brutes(wb: Workbook, data: list[dict], date_str: str):
    ws = wb.create_sheet("DONNEES BRUTES")
    ccy = _dominant_ccy(data)

    HEADERS = [
        "Ticker","Soci\u00e9t\u00e9","Secteur","Pays",
        f"Cours ({_sym(ccy)})","Mkt Cap","EV","Rev.LTM","EBITDA LTM",
        "EV/EBITDA","EV/Rev","P/E","EPS",
        "Mg.Brute","Mg.EBITDA","Mg.Nette","Cro.Rev",
        "ROE","ROA","Current R.","ND/EBITDA","Int.Cov",
        "Altman Z","Beneish M","Mom.52W",
        "Score Val","Score Gr","Score Qual","Score Mom","Score Global",
        "Next Earn","EBITDA NTM Gr","WACC","TGR",
    ]

    _w(ws, 1, 1,
       f"DONN\u00c9ES BRUTES  \u2022  Source de v\u00e9rit\u00e9  \u2022  {date_str}",
       font=_FONT_TITLE, fill=_FILL_NAVY, align=_ALIGN_L)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    for j, h in enumerate(HEADERS, 1):
        _w(ws, 2, j, h, font=_FONT_BOLD, fill=_FILL_MGRAY, align=_ALIGN_C,
           border=_BORDER_THIN)

    for i, t in enumerate(data, 1):
        r = i + 2
        fill = _FILL_LGRAY if i % 2 == 0 else None
        tccy = t.get("currency", ccy) or ccy

        row_vals = [
            t.get("ticker",""),
            t.get("company",""),
            _sector_short(t.get("sector") or "") or "\u2014",
            t.get("country",""),
            _fmt_price(t.get("price"), tccy),
            _fmt_mds(t.get("market_cap"), tccy),
            _fmt_mds(t.get("ev"), tccy),
            _fmt_mds(t.get("revenue_ltm"), tccy),
            _fmt_mds(t.get("ebitda_ltm"), tccy),
            _fmt_mult(t.get("ev_ebitda")),
            _fmt_mult(t.get("ev_revenue")),
            _fmt_mult(t.get("pe")),
            _fmt_price(t.get("eps"), tccy),
            _fmt_pct(t.get("gross_margin"),   sign=False),
            _fmt_pct(t.get("ebitda_margin"),  sign=False),
            _fmt_pct(t.get("net_margin"),     sign=False),
            _fmt_pct(t.get("revenue_growth")),
            _fmt_pct(t.get("roe")),
            _fmt_pct(t.get("roa")),
            _fmt_z(t.get("current_ratio")),
            _fmt_mult(t.get("net_debt_ebitda")),
            _fmt_cov(t.get("interest_coverage")),
            _fmt_z(t.get("altman_z")),
            _fmt_z(t.get("beneish_m")),
            _fmt_pct(t.get("momentum_52w")),
            t.get("score_value"),
            t.get("score_growth"),
            t.get("score_quality"),
            t.get("score_momentum"),
            t.get("score_global"),
            _na(t.get("next_earnings")),
            _fmt_pct(t.get("ebitda_ntm_growth")),
            _fmt_pct(t.get("wacc"), sign=False),
            _fmt_pct(t.get("tgr"), sign=False),
        ]
        for j, v in enumerate(row_vals, 1):
            _w(ws, r, j, v, font=_FONT_BODY, fill=fill,
               align=_ALIGN_L if j in (2,3,4) else _ALIGN_C,
               border=_BORDER_THIN)

    col_w = [8,22,18,6,9,10,10,10,10,9,8,8,8,9,10,9,9,8,8,9,10,8,8,9,9,8,8,9,8,10,10,11,7,6]
    _set_col_widths(ws, col_w)
    _freeze(ws, 3, 2)
    ws.sheet_view.showGridLines = False


# ===========================================================================
# Classe publique
# ===========================================================================

class ScreeningWriter:
    """
    Genere un fichier Excel de screening :
    - Mode template  : charge assets/FinSight_IA_Screening_CAC40_v3.xlsx et
                       injecte les valeurs aux positions exactes sans modifier
                       le formatage existant.
    - Mode fallback  : cree un workbook from-scratch si le template est absent.

    Usage :
        path = ScreeningWriter.generate(tickers_data, "CAC40",
                                        "outputs/generated/screening_cac40.xlsx")

    tickers_data : liste de dicts avec les champs definis dans le brief.
    """

    _DEFAULT_TEMPLATE = "assets/FinSight_IA_Screening_CAC40_v3.xlsx"

    @staticmethod
    def generate(
        tickers_data: list[dict],
        universe_name: str,
        output_path: str,
        template_path: str = None,
    ) -> str:
        if not tickers_data:
            raise ValueError("tickers_data est vide")

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        date_str = date.today().strftime("%d/%m/%Y")

        # Resolution du template — cherche dans plusieurs racines possibles
        # (CWD peut differ de __file__ sous Streamlit)
        if template_path is not None:
            tpl = Path(template_path).resolve()
        else:
            search_roots = [
                Path(__file__).resolve().parent.parent,
                Path.cwd(),
                Path(__file__).parent.parent,
            ]
            tpl = None
            for r in search_roots:
                candidate = r / ScreeningWriter._DEFAULT_TEMPLATE
                print(f"[ScreeningWriter] candidat: {candidate} exists={candidate.exists()}", flush=True)
                if candidate.exists():
                    tpl = candidate
                    break
            if tpl is None:
                tpl = search_roots[0] / ScreeningWriter._DEFAULT_TEMPLATE

        print(f"[ScreeningWriter] template final: {tpl} exists={tpl.exists()}", flush=True)
        print(f"[ScreeningWriter] template_path arg={template_path!r}", flush=True)

        if tpl.exists():
            log.info("[ScreeningWriter] Mode template : %s", tpl.name)
            wb = load_workbook(str(tpl), keep_links=False, data_only=False)

            _inject_donnees_brutes(wb, tickers_data)
            _inject_dashboard(wb, tickers_data, universe_name, date_str)
            _inject_value(wb, tickers_data)
            _inject_growth(wb, tickers_data)
            _inject_quality(wb, tickers_data)
            _inject_momentum(wb, tickers_data)
            _inject_par_secteur(wb, tickers_data)
            _inject_sector_sheets(wb, tickers_data)

        else:
            log.warning(
                "[ScreeningWriter] Template '%s' introuvable — mode from-scratch", tpl
            )
            wb = Workbook()
            wb.remove(wb.active)

            _build_dashboard(wb, tickers_data, universe_name, date_str)
            _build_comparables(wb, tickers_data, date_str)
            _build_ratios_detailles(wb, tickers_data, date_str)
            _build_par_secteur(wb, tickers_data, date_str)
            _build_donnees_brutes(wb, tickers_data, date_str)

        wb.save(str(out))
        log.info("[ScreeningWriter] %s genere (%d tickers)",
                 out.name, len(tickers_data))
        return str(out)
