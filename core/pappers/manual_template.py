"""
Template Excel manuel — fallback pour sociétés dont les comptes ne sont pas
disponibles via Pappers (micro-entreprises, confidentialité, etc.).

Deux fonctions :
- `generate_template(output_path)` : génère un XLSX pré-rempli avec la structure
  YearAccounts (20-30 lignes clés) que l'utilisateur remplit.
- `parse_template(xlsx_path)` : lit un template rempli et renvoie YearAccounts.
"""

from __future__ import annotations

import logging
from dataclasses import fields
from pathlib import Path

from core.pappers.analytics import YearAccounts

log = logging.getLogger(__name__)


# ==============================================================================
# Lignes du template (ordre d'affichage + libellé FR + obligatoire/optionnel)
# ==============================================================================

_TEMPLATE_ROWS: list[tuple[str, str, bool]] = [
    # (attribut, libellé FR, obligatoire)
    # ═══ Compte de résultat ═══
    ("_SECTION", "COMPTE DE RÉSULTAT", False),
    ("chiffre_affaires", "Chiffre d'affaires HT", True),
    ("production_stockee", "Production stockée", False),
    ("production_immobilisee", "Production immobilisée", False),
    ("subventions_exploitation", "Subventions d'exploitation", False),
    ("achats_marchandises", "Achats de marchandises", False),
    ("variation_stocks_marchandises", "Variation stocks marchandises", False),
    ("achats_matieres_premieres", "Achats matières premières", False),
    ("variation_stocks_matieres", "Variation stocks matières", False),
    ("autres_achats_charges_externes", "Autres achats et charges externes", True),
    ("impots_taxes", "Impôts et taxes", True),
    ("salaires_traitements", "Salaires et traitements", True),
    ("charges_sociales", "Charges sociales", True),
    ("dotations_amortissements", "Dotations amortissements", True),
    ("dotations_provisions_exploitation", "Dotations provisions exploitation", False),
    ("produits_financiers", "Produits financiers", False),
    ("charges_financieres", "Charges financières", True),
    ("produits_exceptionnels", "Produits exceptionnels", False),
    ("charges_exceptionnelles", "Charges exceptionnelles", False),
    ("impots_sur_benefices", "Impôt sur les bénéfices", True),
    ("resultat_net", "Résultat net (bénéfice/perte)", True),

    # ═══ Bilan Actif ═══
    ("_SECTION", "BILAN — ACTIF", False),
    ("immobilisations_incorporelles", "Immobilisations incorporelles (net)", False),
    ("immobilisations_corporelles", "Immobilisations corporelles (net)", True),
    ("immobilisations_financieres", "Immobilisations financières (net)", False),
    ("stocks", "Stocks (net)", False),
    ("creances_clients", "Créances clients (net)", True),
    ("autres_creances", "Autres créances", False),
    ("disponibilites", "Disponibilités", True),
    ("total_actif", "TOTAL ACTIF", True),

    # ═══ Bilan Passif ═══
    ("_SECTION", "BILAN — PASSIF", False),
    ("capital_social", "Capital social", True),
    ("reserves", "Réserves (légale + statutaire + autres)", False),
    ("report_a_nouveau", "Report à nouveau", False),
    ("capitaux_propres", "TOTAL Capitaux propres", True),
    ("provisions_risques", "Provisions pour risques et charges", False),
    ("dettes_financieres", "Dettes financières (emprunts)", True),
    ("concours_bancaires", "Concours bancaires courants (découverts)", False),
    ("dettes_fournisseurs", "Dettes fournisseurs", True),
    ("dettes_fiscales_sociales", "Dettes fiscales et sociales", True),
    ("autres_dettes", "Autres dettes", False),
    ("total_passif", "TOTAL PASSIF", True),

    # ═══ Divers ═══
    ("_SECTION", "INFORMATIONS COMPLÉMENTAIRES", False),
    ("effectif_moyen", "Effectif moyen (nombre de salariés)", True),
]


# ==============================================================================
# Génération template
# ==============================================================================

def generate_template(
    output_path: str | Path,
    annee: int | None = None,
    company_name: str | None = None,
    siren: str | None = None,
) -> Path:
    """Génère un XLSX vide avec la structure YearAccounts à remplir."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError as e:
        raise RuntimeError("openpyxl requis : pip install openpyxl") from e

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Saisie comptes"

    # Styles
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1B2A4A")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="3A5288")
    required_font = Font(bold=True, color="1B2A4A")
    optional_font = Font(color="737373")
    input_fill = PatternFill("solid", fgColor="FAFAF5")
    thin_border = Border(
        left=Side(style="thin", color="E5E5E5"),
        right=Side(style="thin", color="E5E5E5"),
        top=Side(style="thin", color="E5E5E5"),
        bottom=Side(style="thin", color="E5E5E5"),
    )

    # Ligne 1 : titre
    ws["A1"] = "FinSight IA — Saisie comptes annuels"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    # Ligne 2 : métadonnées
    ws["A2"] = "Société :"
    ws["B2"] = company_name or "(à compléter)"
    ws["A3"] = "SIREN :"
    ws["B3"] = siren or "(à compléter)"
    ws["A4"] = "Année :"
    ws["B4"] = annee or 2024

    # Ligne d'entête (ligne 6)
    ws["A6"] = "Ligne comptable"
    ws["B6"] = "Valeur (€)"
    ws["C6"] = "Obligatoire"
    for col in ("A6", "B6", "C6"):
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill("solid", fgColor="E5E5E5")

    # Lignes de saisie
    current_row = 7
    for attr, label, required in _TEMPLATE_ROWS:
        if attr == "_SECTION":
            ws[f"A{current_row}"] = label
            ws[f"A{current_row}"].font = section_font
            ws[f"A{current_row}"].fill = section_fill
            ws.merge_cells(f"A{current_row}:C{current_row}")
        else:
            ws[f"A{current_row}"] = label
            ws[f"A{current_row}"].font = required_font if required else optional_font
            ws[f"B{current_row}"] = None
            ws[f"B{current_row}"].fill = input_fill
            ws[f"B{current_row}"].border = thin_border
            ws[f"C{current_row}"] = "Oui" if required else "Optionnel"
            ws[f"C{current_row}"].font = Font(color="1B2A4A" if required else "A3A3A3", size=9)
            # Cellule masquée avec l'attribut pour le parser
            ws[f"D{current_row}"] = attr
            ws[f"D{current_row}"].font = Font(color="FFFFFF", size=1)  # quasi-invisible
        current_row += 1

    # Largeur colonnes
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].hidden = True

    # Instructions finales
    current_row += 2
    ws[f"A{current_row}"] = (
        "📌 Remplissez les cellules grisées. Les lignes marquées « Oui » sont "
        "essentielles au calcul. Les autres améliorent la précision."
    )
    ws[f"A{current_row}"].font = Font(italic=True, color="737373", size=10)
    ws.merge_cells(f"A{current_row}:C{current_row}")

    wb.save(output_path)
    return output_path


# ==============================================================================
# Parsing template rempli
# ==============================================================================

def parse_template(xlsx_path: str | Path) -> YearAccounts | None:
    """Lit un template rempli et reconstruit YearAccounts."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl requis") from e

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Détecte l'année (cellule B4)
    annee_val = ws["B4"].value
    try:
        annee = int(annee_val) if annee_val else 2024
    except (ValueError, TypeError):
        annee = 2024

    # Scan les lignes : col D = attribut, col B = valeur
    extracted: dict[str, float] = {}
    for row in ws.iter_rows(min_row=7, max_row=100, values_only=True):
        if len(row) < 4:
            continue
        attr = row[3] if len(row) > 3 else None
        value = row[1]
        if not isinstance(attr, str) or attr == "_SECTION" or not attr:
            continue
        if value is None:
            continue
        try:
            extracted[attr] = float(value)
        except (ValueError, TypeError):
            continue

    if not extracted:
        return None

    # Valide que les champs appartiennent bien à YearAccounts
    valid_fields = {f.name for f in fields(YearAccounts)}
    filtered = {k: v for k, v in extracted.items() if k in valid_fields}

    return YearAccounts(annee=annee, **filtered)
