"""
tools/render_outputs.py — FinSight IA
Convertit les outputs PDF, PPTX et XLSX en images PNG pour revue visuelle.

Usage :
  python tools/render_outputs.py AAPL
  python tools/render_outputs.py AAPL --only pdf
  python tools/render_outputs.py AAPL --only pptx
  python tools/render_outputs.py AAPL --only xlsx
  python tools/render_outputs.py AAPL --only xlsx --sheet INPUT
"""
from __future__ import annotations

import argparse
import os
import sys
import tempfile
from pathlib import Path

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env", override=True)

POPPLER_PATH = os.getenv("POPPLER_PATH")
CLI_DIR = Path(__file__).parent.parent / "outputs" / "generated" / "cli_tests"


def render_pdf(pdf_path: Path, out_dir: Path, dpi: int = 150) -> list[Path]:
    """Convertit chaque page du PDF en PNG via pdf2image + Poppler."""
    from pdf2image import convert_from_path

    out_dir.mkdir(parents=True, exist_ok=True)
    images = convert_from_path(
        str(pdf_path),
        dpi=dpi,
        poppler_path=POPPLER_PATH,
    )
    paths = []
    for i, img in enumerate(images, start=1):
        p = out_dir / f"pdf_page_{i:02d}.png"
        img.save(str(p), "PNG")
        paths.append(p)
    print(f"[PDF] {len(paths)} pages -> {out_dir}")
    return paths


def render_pptx(pptx_path: Path, out_dir: Path, width_px: int = 1920) -> list[Path]:
    """Convertit chaque slide PPTX en PNG via PowerPoint COM (Office 16)."""
    import win32com.client

    out_dir.mkdir(parents=True, exist_ok=True)

    ppt = win32com.client.Dispatch("PowerPoint.Application")
    try:
        ppt.Visible = False
    except Exception:
        pass  # Certaines versions de PPT refusent Visible=False

    try:
        prs = ppt.Presentations.Open(str(pptx_path.resolve()), ReadOnly=True, Untitled=False, WithWindow=False)
        paths = []
        for i, slide in enumerate(prs.Slides, start=1):
            p = out_dir / f"slide_{i:02d}.png"
            slide.Export(str(p.resolve()), "PNG", width_px, int(width_px * 9 / 16))
            paths.append(p)
        prs.Close()
        print(f"[PPTX] {len(paths)} slides -> {out_dir}")
        return paths
    finally:
        ppt.Quit()


def _make_clean_xlsx(src: Path) -> Path:
    """
    Crée une copie data-only du xlsx (valeurs uniquement, sans liens externes ni VBA).
    Excel COM ne peut pas ouvrir les fichiers générés depuis un template avec liens.
    """
    import openpyxl

    src_wb = openpyxl.load_workbook(src, data_only=True, keep_links=False)
    clean_wb = openpyxl.Workbook()
    clean_wb.remove(clean_wb.active)  # supprime le sheet vide par défaut

    for ws_name in src_wb.sheetnames:
        src_ws  = src_wb[ws_name]
        dest_ws = clean_wb.create_sheet(ws_name)
        for row in src_ws.iter_rows():
            for cell in row:
                dest_ws[cell.coordinate] = cell.value

    tmp = Path(tempfile.mkdtemp()) / f"clean_{src.stem}.xlsx"
    clean_wb.save(tmp)
    return tmp


def render_xlsx(xlsx_path: Path, out_dir: Path, sheets: list[str] | None = None, dpi: int = 150) -> list[Path]:
    """
    Convertit chaque feuille Excel en PNG via Excel COM + export PDF intermédiaire.
    sheets : liste de noms de feuilles à rendre (None = toutes).
    """
    import win32com.client
    from pdf2image import convert_from_path

    out_dir.mkdir(parents=True, exist_ok=True)

    # Crée une copie data-only pour contourner les liens externes
    clean_path = _make_clean_xlsx(xlsx_path)

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False

    paths = []
    try:
        wb = xl.Workbooks.Open(str(clean_path), UpdateLinks=0, ReadOnly=True)

        sheet_names = [ws.Name for ws in wb.Worksheets]
        to_render = [s for s in sheet_names if sheets is None or s in sheets]

        for sheet_name in to_render:
            ws = wb.Worksheets(sheet_name)
            ws.Activate()

            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp_pdf = tmp.name

            # xlTypePDF = 0
            ws.ExportAsFixedFormat(0, tmp_pdf)

            images = convert_from_path(tmp_pdf, dpi=dpi, poppler_path=POPPLER_PATH)
            os.unlink(tmp_pdf)

            safe_name = sheet_name.replace(" ", "_").replace("/", "-")
            for j, img in enumerate(images, start=1):
                suffix = f"_p{j:02d}" if len(images) > 1 else ""
                p = out_dir / f"xlsx_{safe_name}{suffix}.png"
                img.save(str(p), "PNG")
                paths.append(p)

        wb.Close(SaveChanges=False)
        print(f"[XLSX] {len(paths)} image(s) ({len(to_render)} feuille(s)) -> {out_dir}")
        return paths
    finally:
        xl.Quit()
        try:
            os.unlink(clean_path)
        except Exception:
            pass


def render(ticker: str, only: str | None = None, sheet: str | None = None) -> dict:
    ticker = ticker.upper().replace("/", "-")
    renders_dir = CLI_DIR / "renders" / ticker

    pdf_files  = sorted(CLI_DIR.glob(f"{ticker}*report*.pdf"))
    pptx_files = sorted(CLI_DIR.glob(f"{ticker}*pitchbook*.pptx"))
    xlsx_files = sorted(CLI_DIR.glob(f"{ticker}*financials*.xlsx"))

    if not pdf_files and not pptx_files and not xlsx_files:
        print(f"Aucun output trouve pour {ticker} dans {CLI_DIR}")
        sys.exit(1)

    result = {}

    if only in (None, "pdf") and pdf_files:
        result["pdf"] = render_pdf(pdf_files[-1], renders_dir / "pdf")

    if only in (None, "pptx") and pptx_files:
        result["pptx"] = render_pptx(pptx_files[-1], renders_dir / "pptx")

    if only in (None, "xlsx") and xlsx_files:
        sheets = [sheet] if sheet else None
        result["xlsx"] = render_xlsx(xlsx_files[-1], renders_dir / "xlsx", sheets=sheets)

    return result


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("ticker")
    parser.add_argument("--only", choices=["pdf", "pptx", "xlsx"], default=None)
    parser.add_argument("--sheet", default=None, help="Nom de la feuille Excel (ex: INPUT)")
    parser.add_argument("--dpi", type=int, default=150)
    args = parser.parse_args()

    render(args.ticker, only=args.only, sheet=args.sheet)
