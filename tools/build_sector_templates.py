# -*- coding: utf-8 -*-
"""
tools/build_sector_templates.py — Génère les templates XLSX sectoriels
(INSURANCE / REIT / UTILITY) à partir du template OIL_GAS de référence.

Approche : clone + modifications ciblées sur les cellules spécifiques à chaque
secteur, sans toucher aux feuilles INPUT / SCENARIOS / DCF / DASHBOARD 2 /
STOCK_DATA qui ont la même structure pour tous.

Modifications par secteur :
1. Feuille DASHBOARD R54-R56 : label "OIL & GAS KPIs" → label sectoriel, et
   les 2 formules KPI en dessous (remplacer EV/DACF + Net Debt/EBITDAX par
   les ratios sectoriels pertinents).
2. Feuille SCENARIOS R9 : label "EBITDAX Margin (OIL_GAS)" → label secteur.
3. Feuille SECTOR_REF : ajoute une row sectorielle si elle n'existe pas
   (la row Energie/Utilities existe déjà, à réutiliser pour UTILITY).

Le writer excel_writer.py écrit les données financières via _is_rows_active etc.
qui viennent du cell_map du router. Tant que _INSURANCE_CELL_MAP etc. restent
vides (ou STANDARD), le writer utilise le layout STANDARD qui est identique à
celui de OIL_GAS sur la feuille INPUT.

USAGE :
    python tools/build_sector_templates.py

Génère les 3 fichiers dans assets/ et imprime un résumé.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

# Force stdout utf-8 pour Windows
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import openpyxl

_ASSETS = Path(__file__).resolve().parent.parent / "assets"
_BASE = _ASSETS / "FinSight_IA_Template_OIL_GAS_v1.xlsx"


# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURATION PAR SECTEUR
# ═════════════════════════════════════════════════════════════════════════════

SECTOR_CONFIG = {
    "INSURANCE": {
        "file": "FinSight_IA_Template_INSURANCE_v1.xlsx",
        # DASHBOARD R54 label + 2 KPIs
        "kpi_section_label": "MÉTRIQUES ASSURANCE (LTM)",
        "kpi1_label": "P/TBV",
        "kpi1_formula": "=IFERROR(INPUT!H97/(INPUT!H60-INPUT!H42),NA())",
        "kpi1_benchmark": "Bmk : 1,0-1,5x • Prime si >1,5x",
        "kpi2_label": "ROE",
        "kpi2_formula": "=IFERROR(INPUT!H26/INPUT!H60,NA())",
        "kpi2_benchmark": "Bmk : 10-15% stable • Reg Bâle-like",
        # SCENARIOS R9 label
        "scenario_r9_label": "Net Combined Ratio (ASSURANCE)",
        "sector_ref_name": "Assurance / Insurance",
        # RATIOS sheet valuation section header + R9-R14 remplacement
        "ratios_r7_label": "RATIOS DE VALORISATION — ASSURANCE",
        "ratios_r9_label": "P/TBV (Price / Tangible BV)",
        "ratios_r9_formula_h": "=IFERROR(INPUT!H97/(INPUT!H60-INPUT!H42),\"–\")",
        "ratios_r10_label": "P/B (Price / Book Value)",
        "ratios_r10_formula_h": "=IFERROR(INPUT!H97/INPUT!H60,\"–\")",
        "ratios_r13_label": "Solvency II Ratio",
        "ratios_r13_note": "n.d. — source SFCR",
        "ratios_r14_label": "Embedded Value / Share",
        "ratios_r14_note": "n.d. — source EEV",
        "ratios_r20_label": "EBITDA Margin",
        # Section operationnelle (R58)
        "ratios_r58_section": "MÉTRIQUES OPÉRATIONNELLES — ASSURANCE",
        "ratios_r60_label": "Combined Ratio (n.d.)",
        "ratios_r61_label": "Loss Ratio (n.d.)",
        "ratios_r62_label": "Expense Ratio (n.d.)",
        "ratios_r63_label": "Investment Yield (n.d.)",
        # DCF Brent deck label → sector-neutral
        "dcf_brent_label": "(Section Brent désactivée — non pertinente Assurance)",
        # COMPARABLES colonne G7 header (était EBITDAX LTM)
        "comparables_g7": "Book Value\n(LTM)",
        # Supprime la feuille SOTP (Sum-of-the-parts oil&gas)
        "delete_sotp": True,
    },
    "REIT": {
        "file": "FinSight_IA_Template_REIT_v1.xlsx",
        "kpi_section_label": "MÉTRIQUES REIT (LTM)",
        "kpi1_label": "P/FFO",
        "kpi1_formula": "=IFERROR(INPUT!H97/(INPUT!H26+ABS(INPUT!H16)),NA())",
        "kpi1_benchmark": "Bmk : 12-18x • Prime si qualité assets",
        "kpi2_label": "LTV",
        "kpi2_formula": "=IFERROR((INPUT!H48+INPUT!H54)/INPUT!H44,NA())",
        "kpi2_benchmark": "Bmk : <40% sain • Stress si >50%",
        "scenario_r9_label": "NOI Margin (REIT)",
        "sector_ref_name": "Immobilier / REIT",
        "ratios_r7_label": "RATIOS DE VALORISATION — REIT",
        "ratios_r9_label": "P/FFO (Price / Funds From Ops)",
        "ratios_r9_formula_h": "=IFERROR(INPUT!H97/(INPUT!H26+ABS(INPUT!H16)),\"–\")",
        "ratios_r10_label": "P/AFFO (Price / Adj. FFO)",
        "ratios_r10_formula_h": "=IFERROR(INPUT!H97/(INPUT!H26+ABS(INPUT!H16)-ABS(INPUT!H78)),\"–\")",
        "ratios_r13_label": "P/NAV (Price / Net Asset Value)",
        "ratios_r13_note": "n.d. — source expert",
        "ratios_r14_label": "Cap Rate",
        "ratios_r14_note": "n.d. — source expert",
        "ratios_r20_label": "EBITDA Margin (NOI proxy)",
        "ratios_r58_section": "MÉTRIQUES OPÉRATIONNELLES — REIT",
        "ratios_r60_label": "FFO / Share",
        "ratios_r61_label": "AFFO / Share",
        "ratios_r62_label": "LTV Ratio",
        "ratios_r63_label": "Occupancy Rate (n.d.)",
        "dcf_brent_label": "(Section Brent désactivée — non pertinente REIT)",
        "comparables_g7": "NOI\n(LTM)",
        "delete_sotp": True,
    },
    "UTILITY": {
        "file": "FinSight_IA_Template_UTILITY_v1.xlsx",
        "kpi_section_label": "MÉTRIQUES UTILITY (LTM)",
        "kpi1_label": "EV/EBITDA",
        "kpi1_formula": "=IFERROR(INPUT!H98/INPUT!H30,NA())",
        "kpi1_benchmark": "Bmk : 8-11x • Prime si croissance RAB",
        "kpi2_label": "Dividend Yield",
        "kpi2_formula": "=IFERROR(ABS(INPUT!H85)/INPUT!H97,NA())",
        "kpi2_benchmark": "Bmk : 3-5% • Clé utility",
        "scenario_r9_label": "EBITDA Margin (UTILITY)",
        "sector_ref_name": None,  # "Energie / Utilities" existe déjà
        "ratios_r7_label": "RATIOS DE VALORISATION — UTILITY",
        "ratios_r9_label": "EV/EBITDA",
        "ratios_r9_formula_h": "=IFERROR(INPUT!H98/INPUT!H30,\"–\")",
        "ratios_r10_label": "P/E (Price / Earnings)",
        "ratios_r10_formula_h": "=IFERROR(INPUT!H97/INPUT!H26,\"–\")",
        "ratios_r13_label": "EV / RAB (Regulated Asset Base)",
        "ratios_r13_note": "n.d. — source régulateur",
        "ratios_r14_label": "Dividend Yield",
        "ratios_r14_note": "",
        "ratios_r20_label": "EBITDA Margin",
        "ratios_r58_section": "MÉTRIQUES OPÉRATIONNELLES — UTILITY",
        "ratios_r60_label": "Dividend Payout Ratio",
        "ratios_r61_label": "Dividend Coverage",
        "ratios_r62_label": "CapEx / Revenue",
        "ratios_r63_label": "Allowed ROE (n.d.)",
        "dcf_brent_label": "(Section Brent désactivée — non pertinente Utility)",
        "comparables_g7": "EBITDA\n(LTM)",
        "delete_sotp": True,
    },
}


# ═════════════════════════════════════════════════════════════════════════════
# MAIN BUILD
# ═════════════════════════════════════════════════════════════════════════════

def build_template(profile: str, config: dict) -> Path:
    """Clone OIL_GAS et applique les modifications sectorielles."""
    target = _ASSETS / config["file"]

    # Clone
    shutil.copy2(_BASE, target)
    print(f"\n[{profile}] Clone : {target.name}")

    # Ouvre + modifie
    wb = openpyxl.load_workbook(str(target), keep_links=False, data_only=False)

    # ── 1. DASHBOARD R54-R56 : KPI section sectorielle ──────────────────────
    ws_dash = wb["DASHBOARD"]
    # Préserve le style de la cellule en modifiant uniquement .value
    ws_dash.cell(row=54, column=2).value = config["kpi_section_label"]
    ws_dash.cell(row=55, column=2).value = config["kpi1_label"]
    ws_dash.cell(row=55, column=3).value = config["kpi1_formula"]
    ws_dash.cell(row=55, column=4).value = config["kpi1_benchmark"]
    ws_dash.cell(row=56, column=2).value = config["kpi2_label"]
    ws_dash.cell(row=56, column=3).value = config["kpi2_formula"]
    ws_dash.cell(row=56, column=4).value = config["kpi2_benchmark"]
    # Clean OIL_GAS leftovers : E55 FCF Yield auxiliary + E56 Net Debt check
    # (non pertinents pour les autres secteurs)
    ws_dash.cell(row=55, column=5).value = None
    ws_dash.cell(row=56, column=5).value = None
    print(f"  DASHBOARD : {config['kpi_section_label']}")
    print(f"    R55 {config['kpi1_label']:12} = {config['kpi1_formula'][:50]}")
    print(f"    R56 {config['kpi2_label']:12} = {config['kpi2_formula'][:50]}")

    # ── 2. SCENARIOS R9 : label driver ──────────────────────────────────────
    ws_sc = wb["SCENARIOS"]
    # Colonne B row 9 : "EBITDAX Margin (OIL_GAS)"
    # On modifie uniquement le label, pas les ArrayFormulas qui suivent
    ws_sc.cell(row=9, column=2).value = config["scenario_r9_label"]
    print(f"  SCENARIOS R9 label : {config['scenario_r9_label']}")

    # ── 3. SECTOR_REF : ajoute row sectorielle si demandé ──────────────────
    if config.get("sector_ref_name"):
        ws_ref = wb["SECTOR_REF"]
        # Trouve la première row vide après la dernière ligne data
        new_row = ws_ref.max_row + 2  # +2 pour laisser un espacement
        # Ajoute les 11 drivers standards avec des valeurs par défaut raisonnables
        # Structure : col A = sector name, col B = driver name, col C/D/E = Pess/Real/Opt
        default_drivers = {
            "INSURANCE": [
                ("Revenue Growth Rate (Y1-Y5)", 0.01, 0.04, 0.07),
                ("Gross Margin", 0.10, 0.18, 0.25),
                ("EBITDA Margin", 0.05, 0.12, 0.18),
                ("SG&A (% Revenue)", 0.08, 0.06, 0.04),
                ("R&D (% Revenue)", 0.00, 0.00, 0.00),
                ("CapEx (% Revenue)", 0.02, 0.015, 0.01),
                ("Change in NWC (% Rev Change)", 0.02, 0.01, 0.005),
                ("Terminal Growth Rate", 0.01, 0.02, 0.03),
                ("WACC", 0.09, 0.08, 0.07),
                ("EV/EBITDA Exit Multiple", 6, 8, 11),
            ],
            "REIT": [
                ("Revenue Growth Rate (Y1-Y5)", 0.00, 0.03, 0.06),
                ("Gross Margin", 0.60, 0.70, 0.80),  # REIT high gross margin
                ("EBITDA Margin", 0.55, 0.65, 0.75),
                ("SG&A (% Revenue)", 0.05, 0.03, 0.02),
                ("R&D (% Revenue)", 0.00, 0.00, 0.00),
                ("CapEx (% Revenue)", 0.10, 0.07, 0.04),
                ("Change in NWC (% Rev Change)", 0.02, 0.01, 0.005),
                ("Terminal Growth Rate", 0.01, 0.02, 0.025),
                ("WACC", 0.08, 0.07, 0.055),
                ("EV/EBITDA Exit Multiple", 15, 18, 22),
            ],
        }
        drivers = default_drivers.get(profile, [])
        for i, (drv_name, pes, real, opt) in enumerate(drivers):
            r = new_row + i
            ws_ref.cell(row=r, column=1).value = config["sector_ref_name"]
            ws_ref.cell(row=r, column=2).value = drv_name
            ws_ref.cell(row=r, column=3).value = pes
            ws_ref.cell(row=r, column=4).value = real
            ws_ref.cell(row=r, column=5).value = opt
        print(f"  SECTOR_REF : +{len(drivers)} rows pour '{config['sector_ref_name']}'")

    # ── 4. RATIOS sheet : section VALUATION + OPERATIONAL sectorielles ──────
    ws_rat = wb["RATIOS"]
    # R7 label section valorisation
    if config.get("ratios_r7_label"):
        ws_rat.cell(row=7, column=2).value = config["ratios_r7_label"]
    # R9 (EV/DACF → sector KPI 1)
    if config.get("ratios_r9_label"):
        ws_rat.cell(row=9, column=2).value = config["ratios_r9_label"]
        ws_rat.cell(row=9, column=8).value = config["ratios_r9_formula_h"]
        # Clear D9-G9 (historical years will show "–" fallback)
        for c in range(4, 8):
            ws_rat.cell(row=9, column=c).value = "–"
    # R10 (EV/EBITDAX → sector KPI 2)
    if config.get("ratios_r10_label"):
        ws_rat.cell(row=10, column=2).value = config["ratios_r10_label"]
        ws_rat.cell(row=10, column=8).value = config["ratios_r10_formula_h"]
        for c in range(4, 8):
            ws_rat.cell(row=10, column=c).value = "–"
    # R13 & R14 (oil&gas-specific : EV/Production, Price/Reserves → sector notes)
    if config.get("ratios_r13_label"):
        ws_rat.cell(row=13, column=2).value = config["ratios_r13_label"]
        ws_rat.cell(row=13, column=8).value = config.get("ratios_r13_note", "n.d.")
    if config.get("ratios_r14_label"):
        ws_rat.cell(row=14, column=2).value = config["ratios_r14_label"]
        ws_rat.cell(row=14, column=8).value = config.get("ratios_r14_note", "n.d.")
    # R20 EBITDAX Margin → EBITDA Margin (ou label sectoriel)
    if config.get("ratios_r20_label"):
        ws_rat.cell(row=20, column=2).value = config["ratios_r20_label"]
    # R58 : section opérationnelle sectorielle
    if config.get("ratios_r58_section"):
        ws_rat.cell(row=58, column=2).value = config["ratios_r58_section"]
    # R60-R63 : labels opérationnels sectoriels
    # Pour INSURANCE/REIT : la plupart des ratios (Combined Ratio, NAV, Cap Rate)
    # ne sont pas calculables depuis yfinance → on met les labels + "n.d."
    # Pour UTILITY : les ratios sont calculables donc on remplace les formules
    for idx, key in enumerate(["ratios_r60_label", "ratios_r61_label",
                                "ratios_r62_label", "ratios_r63_label"]):
        if config.get(key):
            r = 60 + idx
            ws_rat.cell(row=r, column=2).value = config[key]
            # Clear les formules OIL_GAS originales (H60=Net Debt/EBITDAX, H61=CapEx/DD&A,
            # H62=DACF, H63=FCF Yield — non pertinents pour INSURANCE/REIT)
            if profile == "UTILITY":
                # UTILITY : formules propres
                if idx == 0:  # Dividend Payout Ratio = |Dividends| / NI
                    formula = "=IFERROR(ABS(INPUT!H85)/INPUT!H26,\"–\")"
                elif idx == 1:  # Dividend Coverage = NI / |Dividends|
                    formula = "=IFERROR(INPUT!H26/ABS(INPUT!H85),\"–\")"
                elif idx == 2:  # CapEx / Revenue
                    formula = "=IFERROR(ABS(INPUT!H78)/INPUT!H9,\"–\")"
                else:  # Allowed ROE placeholder
                    formula = "n.d."
                for c in range(5, 9):  # E, F, G, H
                    year_col = chr(ord('D') + (c - 4))  # D/E/F/G/H
                    f = formula.replace("H26", f"{year_col}26").replace(
                        "H60", f"{year_col}60").replace(
                        "H85", f"{year_col}85").replace(
                        "H78", f"{year_col}78").replace(
                        "H9", f"{year_col}9")
                    ws_rat.cell(row=r, column=c).value = f if f.startswith("=") else f
            else:
                # INSURANCE/REIT : clear + placeholder
                for c in range(5, 9):  # E-H
                    ws_rat.cell(row=r, column=c).value = "n.d."
    print(f"  RATIOS section : {config.get('ratios_r58_section', '-')}")

    # ── 4b. COMPARABLES G7 header : EBITDAX → sector-specific label ────────
    if config.get("comparables_g7"):
        ws_comp = wb["COMPARABLES"]
        ws_comp.cell(row=7, column=7).value = config["comparables_g7"]
        print(f"  COMPARABLES G7 : {config['comparables_g7'].replace(chr(10), ' ')}")

    # ── 4c. SOTP sheet : suppression pour secteurs non-oil&gas ──────────────
    if config.get("delete_sotp") and "SOTP" in wb.sheetnames:
        del wb["SOTP"]
        print(f"  SOTP sheet : supprimée (spécifique oil&gas)")

    # ── 5. DCF sheet : désactive la section Brent (B40-B46) ─────────────────
    if config.get("dcf_brent_label"):
        ws_dcf = wb["DCF"]
        ws_dcf.cell(row=40, column=2).value = config["dcf_brent_label"]
        # Clear B41-B46 brent-specific content
        for r in range(41, 47):
            for c in range(2, 10):
                cell = ws_dcf.cell(row=r, column=c)
                if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.value = None
        print(f"  DCF Brent deck : neutralisé")

    # ── 6. Save ──────────────────────────────────────────────────────────────
    wb.save(str(target))
    sz = target.stat().st_size // 1024
    print(f"  Saved : {target.name} ({sz} ko)")
    return target


def main():
    if not _BASE.exists():
        print(f"ERREUR : base template introuvable : {_BASE}")
        return 1

    print("=" * 70)
    print("  BUILD SECTOR TEMPLATES — base OIL_GAS (TTE)")
    print("=" * 70)

    for profile, config in SECTOR_CONFIG.items():
        build_template(profile, config)

    print("\n" + "=" * 70)
    print("  OK — 3 templates générés :")
    for profile, config in SECTOR_CONFIG.items():
        print(f"    assets/{config['file']}")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    sys.exit(main())
