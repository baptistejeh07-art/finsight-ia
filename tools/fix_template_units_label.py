# -*- coding: utf-8 -*-
"""
fix_template_units_label.py — Corrige le label "$000"/"(000)" → "$M"/"(M)"
dans le template Excel.

Le pipeline yfinance_source.py fait `val / 1_000_000` ligne 145, donc toutes
les valeurs financières sont en MILLIONS. Mais le template a des labels en
thousands. Conséquence : les utilisateurs lisent des chiffres ×1000 trop bas
quand ils prennent le label au pied de la lettre.

Fix : remplacer tous les labels "$000"/"(000)" par "$M"/"(M)" dans :
  - assets/TEMPLATE.xlsx (template principal)
  - assets/_LBO_WORKING_AAPL.xlsx (fichier de travail LBO)

Et remplacer "All figures in USD thousands" par "All figures in USD millions"
dans les en-têtes de feuilles.
"""
from __future__ import annotations

import sys
from pathlib import Path
from openpyxl import load_workbook


def fix_workbook(path: Path) -> dict:
    """Modifie un workbook pour corriger les labels d'unités."""
    wb = load_workbook(str(path))
    stats = {"cells_modified": 0, "sheets_touched": []}

    replacements = [
        ("($000)",         "($M)"),
        ("(000)",          "(M)"),
        ("$000",           "$M"),
        ("USD thousands",  "USD millions"),
        ("in thousands",   "in millions"),
        ("(in $000)",      "(in $M)"),
        ("($ thousands)",  "($ millions)"),
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_modified = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                original = cell.value
                new_val = original
                for old, new in replacements:
                    new_val = new_val.replace(old, new)
                if new_val != original:
                    try:
                        cell.value = new_val
                        stats["cells_modified"] += 1
                        sheet_modified = True
                    except (AttributeError, TypeError):
                        # MergedCell ou autre, on skip
                        pass
        if sheet_modified:
            stats["sheets_touched"].append(sheet_name)

    wb.save(str(path))
    return stats


def main():
    targets = [
        Path("assets/TEMPLATE.xlsx"),
        Path("assets/_LBO_WORKING_AAPL.xlsx"),
    ]

    print("=" * 60)
    print("  FIX LABELS UNITES : thousands -> millions")
    print("=" * 60)

    for path in targets:
        if not path.exists():
            print(f"  [SKIP] {path} (not found)")
            continue
        print(f"\n  {path.name}")
        stats = fix_workbook(path)
        print(f"    cells modified : {stats['cells_modified']}")
        print(f"    sheets touched : {stats['sheets_touched']}")

    print()
    print("=" * 60)
    print("  DONE")
    print("=" * 60)


if __name__ == "__main__":
    main()
